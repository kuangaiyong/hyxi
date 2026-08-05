"""Excel 生成服务 - 复用 openpyxl 样式，含舆情报告导出"""

import os
import json
from io import BytesIO
from datetime import datetime
from typing import List, Optional
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter
from app.config import settings
from app.services.post_tree import order_by_thread
from app.services.progress_manager import ProgressManager


# 复用 build_excel.py 的样式常量
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CONTENT_ALIGN = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="top")
ALT_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")


# ===== 导出报告的样式（用户唯一下载得到的那份，与流水线那份各用各的）=====

# Arial 没有中文字形，中文全靠 Excel 回退字体渲染，同一行里中英文基线和字宽对不齐
FONT_NAME = "Microsoft YaHei"

SENTIMENT_CN = {"positive": "正面", "negative": "负面", "neutral": "中立"}
UNANALYZED = "未分析"

EXPORT_TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="1E293B")
EXPORT_SECTION_FONT = Font(name=FONT_NAME, size=12, bold=True, color="2F5496")
EXPORT_HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
EXPORT_LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True)
EXPORT_BODY_FONT = Font(name=FONT_NAME, size=10)
REPLY_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

# 浅底深字。前端那三个饱和色（#10B981 等）在 Excel 里是整格上色，打印和浅色主题下都刺眼
SENTIMENT_STYLE = {
    "正面": ("D1FAE5", "065F46"),
    "负面": ("FEE2E2", "991B1B"),
    "中立": ("F3F4F6", "374151"),
}

EXPORT_WIDTHS = {
    "index": 6, "source": 16, "level": 6, "username": 14, "timestamp": 17,
    "content": 50, "translation": 50, "sentiment": 8, "intensity": 11,
    "reason": 40, "dimensions": 24,
}
EXPORT_CENTERED = {"index", "level", "timestamp", "sentiment", "intensity"}

# Excel 明细表和 CSV 共用这一份列定义，两种格式的列因此不可能走偏
EXPORT_COLUMNS = [
    ("index", "序号"), ("source", "来源"), ("level", "层级"),
    ("username", "用户名"), ("timestamp", "发布时间"),
    ("content", "原文"), ("translation", "中文翻译"),
    ("sentiment", "情感"), ("intensity", "强度"),
    ("reason", "分析理由"), ("dimensions", "涉及维度"),
]


def _star_level(value) -> int:
    """LLM 可能把 intensity 返回成字符串、None 或越界数值，星级渲染只接受 0..5"""
    if not isinstance(value, (int, float)) or value != value:
        return 0
    return round(max(0, min(5, value)))


def _parse_dutch_timestamp(ts: str):
    """帖子时间戳落盘格式为 dd-mm-yyyy HH:MM，字典序不等于时间序"""
    try:
        return datetime.strptime(ts, "%d-%m-%Y %H:%M")
    except (ValueError, TypeError):
        return None


def _save_workbook(wb, output_path: str) -> None:
    """保存工作簿。

    Windows 上用户若正用 Excel 打开同名文件，save 会抛 PermissionError，
    未捕获时前端只看到一个 500，无从判断该做什么。
    """
    try:
        wb.save(output_path)
    except PermissionError:
        raise Exception(f"无法写入 {os.path.basename(output_path)}，请先关闭已在 Excel 中打开的同名文件")


class ExcelService:
    """生成翻译 Excel 报告"""

    @staticmethod
    async def execute(
        task_id: str,
        posts: list,
        params: dict,
        progress: ProgressManager,
        step_index: int = 2,
        sources: Optional[dict] = None,
    ) -> dict:
        """生成 Excel。

        step_index 不能写死 2 —— plan 里有多个 collect 步骤时 Excel 就不在第 3 位了。
        sources 是 {source_id: {"name": ...}}，用来把来源列渲染成人看得懂的名字。
        """
        include_stats = params.get("include_stats", True)
        source_names = {sid: m.get("name", sid) for sid, m in (sources or {}).items()}
        posts = order_by_thread(posts)

        await progress.emit(task_id, "step_progress", {
            "step": step_index,
            "progress": 0.0,
            "message": "正在生成 Excel 报告...",
        })

        wb = Workbook()

        # ===== Sheet 1: 论坛帖子翻译 =====
        ws = wb.active
        ws.title = "论坛帖子翻译"

        headers = [
            ("序号", 6), ("来源", 20), ("层级", 6), ("用户名", 16), ("发布时间", 18),
            ("原文", 55), ("中文翻译", 55), ("页码", 6),
        ]
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"

        for idx, post in enumerate(posts):
            row = idx + 2
            level = int(post.get("reply_level", 0) or 0)
            # 回复行统一浅底，比隔行底纹更能一眼看出层级；不用 openpyxl 的分组，
            # 各家查看器对 outline 的行为不一致
            fill = ALT_FILL if (level > 0 or idx % 2 == 1) else PatternFill()
            sid = post.get("source", "")

            values = [
                idx + 1,
                source_names.get(sid, sid),
                level,
                ("　" * level + "└─ " if level else "") + post.get("username", ""),
                post.get("timestamp", ""),
                post.get("content", ""),
                post.get("translation", ""),
                post.get("page_number", ""),
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = THIN_BORDER
                cell.fill = fill
                if col in (1, 3, 8):
                    cell.alignment = CENTER_ALIGN
                    cell.font = Font(name="Arial", size=10)
                elif col == 4:
                    cell.font = Font(name="Arial", size=10, bold=True)
                    cell.alignment = CONTENT_ALIGN
                elif col in (6, 7):
                    cell.font = Font(name="Arial", size=10)
                    cell.alignment = CONTENT_ALIGN
                else:
                    cell.font = Font(name="Arial", size=9)
                    cell.alignment = CONTENT_ALIGN

            ws.row_dimensions[row].height = max(60, min(250, len(post.get("content", "")) * 0.4))

        await progress.emit(task_id, "step_progress", {
            "step": step_index, "progress": 0.5, "message": "帖子表已生成...",
        })

        # ===== Sheet 2: 统计信息 =====
        if include_stats:
            ws2 = wb.create_sheet("统计信息")

            # 基本统计
            stats_headers = [("统计项", 30), ("数值", 20)]
            for col, (header, width) in enumerate(stats_headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGN
                cell.border = THIN_BORDER
                ws2.column_dimensions[get_column_letter(col)].width = width

            usernames = [p.get("username", "") for p in posts]
            user_counter = Counter(usernames)
            unique_users = len(user_counter)
            pages = set(p.get("page_number", 1) for p in posts)
            timestamps = [p.get("timestamp", "") for p in posts if p.get("timestamp")]
            dated = [(_parse_dutch_timestamp(t), t) for t in timestamps]
            dated = [d for d in dated if d[0] is not None]

            stat_rows = [
                ("总帖子数", len(posts)),
                ("唯一用户数", unique_users),
                ("总页数", len(pages)),
                ("时间范围开始", min(dated)[1] if dated else "N/A"),
                ("时间范围结束", max(dated)[1] if dated else "N/A"),
            ]
            for i, (label, value) in enumerate(stat_rows):
                ws2.cell(row=i + 2, column=1, value=label).border = THIN_BORDER
                ws2.cell(row=i + 2, column=2, value=str(value)).border = THIN_BORDER
                ws2.cell(row=i + 2, column=2).alignment = CENTER_ALIGN

            # 活跃用户
            start_row = len(stat_rows) + 4
            ws2.cell(row=start_row, column=1, value="活跃用户排名").font = Font(
                name="Arial", size=12, bold=True
            )

            user_headers = [("用户", 20), ("帖子数", 10)]
            for col, (header, width) in enumerate(user_headers, 1):
                cell = ws2.cell(row=start_row + 1, column=col, value=header)
                cell.font = HEADER_FONT
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = HEADER_ALIGN
                cell.border = THIN_BORDER

            for i, (user, count) in enumerate(user_counter.most_common(15)):
                row = start_row + 2 + i
                ws2.cell(row=row, column=1, value=user).border = THIN_BORDER
                ws2.cell(row=row, column=2, value=count).border = THIN_BORDER
                ws2.cell(row=row, column=2).alignment = CENTER_ALIGN

            ws2.freeze_panes = "A2"

        # 保存
        # 带时间戳，避免同一任务重跑时把上一次的导出直接盖掉。
        # 这是**落盘名**，只在 exports 目录里做区分；下载时另算一个带来源和导出时间的
        # 中文名（results.py 的 _export_filename），别把这里的名字当成用户看到的名字
        output_name = f"hyxi_report_{task_id[:8]}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        output_path = os.path.join(settings.exports_dir, output_name)
        _save_workbook(wb, output_path)

        await progress.emit(task_id, "step_progress", {
            "step": step_index,
            "progress": 1.0,
            "message": f"Excel 已保存: {output_name}",
        })

        return {
            "file_path": output_path,
            "file_name": output_name,
        }

    # ===== 用户下载的那份报告 =====

    @staticmethod
    def build_export(rows: List[dict], meta: dict, columns: List[tuple]) -> bytes:
        """生成「概览 + 帖子明细」两张表的报告，直接返回字节流。

        **不落盘**：每次下载现算。写文件既会在 exports 里堆垃圾，两个人同时下载还会
        撞成一个在写、另一个在读。
        """
        wb = Workbook()
        ExcelService._write_overview(wb.active, rows, meta)
        ExcelService._write_details(wb.create_sheet("帖子明细"), rows, columns)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    def _section(ws, row: int, title: str) -> int:
        ws.cell(row=row, column=1, value=title).font = EXPORT_SECTION_FONT
        return row + 1

    @staticmethod
    def _table_head(ws, row: int, labels: List[str]) -> int:
        for col, label in enumerate(labels, 1):
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = EXPORT_HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        return row + 1

    @staticmethod
    def _paint_sentiment(cell, label: str) -> None:
        style = SENTIMENT_STYLE.get(label)
        if not style:
            return
        bg, fg = style
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=fg)

    @staticmethod
    def _write_overview(ws, rows: List[dict], meta: dict) -> None:
        """概览表。

        情感分布**从明细行现算**，不取 summary 里那份 —— summary 的占比只按已分析的
        帖子算，直接搬过来会和明细表对不上，一份报告里两个数打架最难解释。
        """
        ws.title = "概览"
        for col, width in zip("ABCD", (20, 46, 12, 12)):
            ws.column_dimensions[col].width = width

        ws.cell(row=1, column=1, value="HYXi 舆情分析报告").font = EXPORT_TITLE_FONT
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        row = ExcelService._section(ws, 3, "任务信息")
        total = len(rows)
        for label, value in (
            ("任务描述", meta.get("description", "")),
            ("数据来源", meta.get("sources", "")),
            ("导出时间", meta.get("exported_at", "")),
            ("帖子总数", f"{total}（其中评论 {meta.get('replies', 0)}）"),
            ("已完成舆情分析", f"{meta.get('analyzed', 0)} / {total}"),
            ("最早发布时间", meta.get("time_start") or "-"),
            ("最晚发布时间", meta.get("time_end") or "-"),
        ):
            ws.cell(row=row, column=1, value=label).font = EXPORT_LABEL_FONT
            ws.cell(row=row, column=2, value=str(value)).font = EXPORT_BODY_FONT
            row += 1

        row = ExcelService._section(ws, row + 1, "情感分布")
        row = ExcelService._table_head(ws, row, ["情感", "数量", "占比"])
        counts = Counter(r["sentiment"] for r in rows)
        for label in list(SENTIMENT_STYLE) + [UNANALYZED]:
            count = counts.get(label, 0)
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{count / total * 100:.1f}%" if total else "-")
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN
                cell.font = EXPORT_BODY_FONT
            ExcelService._paint_sentiment(ws.cell(row=row, column=1), label)
            row += 1

        # 越界值要先收敛再平均。明细列里 12 已经被画成 ★★★★★，这里照原样算就会出现
        # 「平均强度 12.0 / 5」——同一份文件里两个口径打架
        scored = [
            min(5, max(0, r["intensity"]))
            for r in rows if isinstance(r["intensity"], (int, float))
        ]
        ws.cell(row=row, column=1, value="平均强度").font = EXPORT_LABEL_FONT
        avg = round(sum(scored) / len(scored), 2) if scored else 0
        # :g 去掉整数的小数尾巴，「5.0 / 5」读着像个 bug
        ws.cell(row=row, column=2, value=f"{avg:g} / 5").font = EXPORT_BODY_FONT
        row += 2

        top_dims = (meta.get("summary") or {}).get("top_dimensions") or []
        if top_dims:
            row = ExcelService._section(ws, row, "TOP 10 关注维度")
            row = ExcelService._table_head(ws, row, ["排名", "维度", "提及次数"])
            for rank, item in enumerate(top_dims[:10], 1):
                # JSON 往返后元组变成列表，按下标取对两种形态都成立
                for col, value in enumerate((rank, item[0], item[1]), 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = THIN_BORDER
                    cell.font = EXPORT_BODY_FONT
                    cell.alignment = CONTENT_ALIGN if col == 2 else CENTER_ALIGN
                row += 1
            row += 1

        top_users = meta.get("top_users") or []
        if top_users:
            row = ExcelService._section(ws, row, "活跃用户 TOP 15")
            row = ExcelService._table_head(ws, row, ["用户", "帖子数"])
            for user, count in top_users:
                for col, value in enumerate((user, count), 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = THIN_BORDER
                    cell.font = EXPORT_BODY_FONT
                    cell.alignment = CONTENT_ALIGN if col == 1 else CENTER_ALIGN
                row += 1

    @staticmethod
    def _write_details(ws, rows: List[dict], columns: List[tuple]) -> None:
        keys = [key for key, _ in columns]
        for col, (key, label) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = EXPORT_HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = EXPORT_WIDTHS.get(key, 16)

        sentiment_col = keys.index("sentiment") + 1
        for offset, item in enumerate(rows):
            row = offset + 2
            level = item["level"]
            for col, key in enumerate(keys, 1):
                value = item[key]
                if key == "username" and level:
                    value = "　" * level + "└─ " + str(value)
                elif key == "intensity":
                    stars = _star_level(value)
                    value = "★" * stars + "☆" * (5 - stars) if stars else ""
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = THIN_BORDER
                cell.font = EXPORT_BODY_FONT
                cell.alignment = CENTER_ALIGN if key in EXPORT_CENTERED else CONTENT_ALIGN
                # 只有评论行上底色。再叠一层隔行斑马纹，两种底色互相掩盖，
                # 反而看不出哪行是评论
                if level:
                    cell.fill = REPLY_FILL
            ExcelService._paint_sentiment(ws.cell(row=row, column=sentiment_col), item["sentiment"])

            longest = max(len(item["content"] or ""), len(item["translation"] or ""))
            ws.row_dimensions[row].height = min(180, max(30, (longest // 45 + 1) * 16))

        ws.freeze_panes = "A2"
        # 让人能直接按情感、来源筛选，比翻 88 行找负面快得多
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
