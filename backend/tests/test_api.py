"""FastAPI 端点集成测试 — 使用 TestClient 真实请求"""

import os
import re
import sys
import json
import tempfile
import shutil
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from fastapi.testclient import TestClient


class TestAPIEndpointsEndToEnd:
    """API 端点端到端测试"""

    @classmethod
    def setup_class(cls):
        cls.tmpdir = tempfile.mkdtemp()
        # 设置数据目录到临时位置
        import app.config as cfg
        # 本机根目录若有 .env，api_key 会在 import 期被读入，这批用例就会全部 401
        cls._old_key = cfg.settings.api_key
        cfg.settings.api_key = ""
        cfg.settings.data_dir = cls.tmpdir
        cfg.settings.tasks_dir = os.path.join(cls.tmpdir, "tasks")
        cfg.settings.exports_dir = os.path.join(cls.tmpdir, "exports")
        os.makedirs(cfg.settings.tasks_dir, exist_ok=True)
        os.makedirs(cfg.settings.exports_dir, exist_ok=True)

        from main import app
        from app.services import storage
        cls.storage = storage
        storage.set_app_config("llm", {
            "api_key": "sk-test", "base_url": "https://api.test.com",
            "model_name": "test-model",
        })
        cls.client = TestClient(app)

    @classmethod
    def teardown_class(cls):
        import app.config as cfg
        cfg.settings.api_key = cls._old_key
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def test_health_check(self):
        resp = self.client.get("/api/health")
        assert resp.status_code == 200
        assert resp.json() == {"status": "ok"}

    def test_root_endpoint(self):
        resp = self.client.get("/")
        assert resp.status_code == 200
        assert resp.json()["service"] == "HYXi 舆情分析 API"

    def test_config_crud(self):
        # 保存配置
        resp = self.client.post("/api/v1/config", json={
            "api_key": "sk-test-key",
            "base_url": "https://api.test.com",
            "model_name": "test-model",
        })
        assert resp.status_code == 200
        assert resp.json()["is_configured"] is True

        # 获取配置(不包含 api_key)
        resp = self.client.get("/api/v1/config")
        assert resp.status_code == 200
        assert "api_key" not in resp.json()
        assert resp.json()["base_url"] == "https://api.test.com"

        # 重置配置
        resp = self.client.delete("/api/v1/config")
        assert resp.status_code == 200
        assert resp.json()["is_configured"] is False

    def test_vision_config_crud_is_isolated_from_the_llm_config(self):
        """多模态配置走 app_config 的 vision 前缀，与 llm 前缀互不影响。

        两组配置混在一起的话，改一个模型会把另一个也改掉 —— 翻译和图片理解用的
        通常不是同一个供应商。
        """
        self.client.post("/api/v1/config", json={
            "api_key": "sk-text", "base_url": "https://text.test", "model_name": "text-model",
        })
        resp = self.client.post("/api/v1/config/vision", json={
            "api_key": "sk-vision", "base_url": "https://vision.test", "model_name": "vl-model",
        })
        assert resp.status_code == 200
        assert resp.json()["is_configured"] is True

        vision = self.client.get("/api/v1/config/vision").json()
        assert vision["model_name"] == "vl-model"
        assert "api_key" not in vision
        # 文本模型没被带跑
        assert self.client.get("/api/v1/config").json()["model_name"] == "text-model"

        # 清掉视觉配置不影响文本配置
        assert self.client.delete("/api/v1/config/vision").json()["is_configured"] is False
        assert self.client.get("/api/v1/config").json()["is_configured"] is True

    def test_vision_config_defaults_when_unset(self):
        """没配置时回默认值且 is_configured=false —— 前端据此显示「纯文本分析」"""
        self.client.delete("/api/v1/config/vision")
        body = self.client.get("/api/v1/config/vision").json()
        assert body["is_configured"] is False
        assert body["base_url"] == "https://api.kimi.com/coding/v1"
        assert body["model_name"] == "kimi-for-coding"

    def test_task_list_empty_initially(self):
        resp = self.client.get("/api/v1/tasks")
        assert resp.status_code == 200
        assert resp.json()["total"] == 0

    def test_create_task_and_fetch(self):
        resp = self.client.post("/api/v1/tasks", json={
            "description": "端到端测试任务"
        })
        assert resp.status_code == 201
        task_id = resp.json()["id"]
        assert resp.json()["status"] in ("pending", "parsing")  # 取决于执行速度

        # 获取任务详情
        resp = self.client.get(f"/api/v1/tasks/{task_id}")
        assert resp.status_code == 200
        assert resp.json()["description"] == "端到端测试任务"

        # 列表包含该任务
        resp = self.client.get("/api/v1/tasks")
        assert resp.json()["total"] >= 1

    def test_task_list_includes_all(self):
        resp = self.client.get("/api/v1/tasks")
        assert resp.status_code == 200
        tasks = resp.json()["tasks"]
        assert len(tasks) >= 1

    def test_cancel_task(self):
        # 创建一个任务然后取消
        resp = self.client.post("/api/v1/tasks", json={
            "description": "待取消任务"
        })
        task_id = resp.json()["id"]

        resp = self.client.delete(f"/api/v1/tasks/{task_id}")
        assert resp.status_code in (200, 400)  # 可能已经执行完毕

    def test_schedule_presets(self):
        resp = self.client.get("/api/v1/schedules/presets")
        assert resp.status_code == 200
        presets = resp.json()["presets"]
        assert "hourly" in presets
        assert "daily" in presets

    def test_schedule_crud(self):
        # 创建
        resp = self.client.post("/api/v1/schedules", json={
            "description": "定时测试任务",
            "interval": "daily",
            "time": "09:00",
        })
        assert resp.status_code == 201
        sched_id = resp.json()["id"]
        assert resp.json()["enabled"] is True

        # 列表（检查不报错即可，next_run_time 格式因 APScheduler 版本而异）
        resp = self.client.get("/api/v1/schedules")
        assert resp.status_code == 200

        # 切换
        resp = self.client.post(f"/api/v1/schedules/{sched_id}/toggle")
        assert resp.status_code == 200
        assert resp.json()["enabled"] is False

        # 删除
        resp = self.client.delete(f"/api/v1/schedules/{sched_id}")
        assert resp.status_code == 200

    def _schedules_on_disk(self):
        import app.config as cfg
        path = os.path.join(cfg.settings.data_dir, "scheduled_tasks.json")
        if not os.path.exists(path):
            return []
        with open(path, encoding="utf-8") as f:
            return json.load(f)

    def test_malformed_time_rejected_and_not_persisted(self):
        before = len(self._schedules_on_disk())
        for bad_time in ("9", "25:00", "09:70", "abc", ""):
            resp = self.client.post("/api/v1/schedules", json={
                "description": "畸形时间",
                "interval": "daily",
                "time": bad_time,
            })
            assert resp.status_code == 422, f"{bad_time!r} 被接受了"
        # 脏配置不得落盘，否则下次启动加载时会砖化调度器
        assert len(self._schedules_on_disk()) == before

    def test_patch_rejects_unknown_interval(self):
        resp = self.client.post("/api/v1/schedules", json={
            "description": "待更新任务", "interval": "daily", "time": "09:00",
        })
        sched_id = resp.json()["id"]

        resp = self.client.patch(f"/api/v1/schedules/{sched_id}", json={"interval": "weekly"})
        assert resp.status_code == 400

        resp = self.client.patch(f"/api/v1/schedules/{sched_id}", json={"time": "99:99"})
        assert resp.status_code == 422

        # 合法更新仍然可用
        resp = self.client.patch(f"/api/v1/schedules/{sched_id}", json={"interval": "hourly"})
        assert resp.status_code == 200
        self.client.delete(f"/api/v1/schedules/{sched_id}")

    def test_patch_paused_schedule_does_not_500(self):
        resp = self.client.post("/api/v1/schedules", json={
            "description": "暂停后再编辑", "interval": "daily", "time": "09:00",
        })
        sched_id = resp.json()["id"]
        assert self.client.post(f"/api/v1/schedules/{sched_id}/toggle").json()["enabled"] is False

        # 暂停中的任务没有注册 job，remove_job 会抛 JobLookupError
        resp = self.client.patch(f"/api/v1/schedules/{sched_id}", json={"description": "改个描述"})
        assert resp.status_code == 200
        assert resp.json()["description"] == "改个描述"
        self.client.delete(f"/api/v1/schedules/{sched_id}")

    def test_404_for_nonexistent_task(self):
        resp = self.client.get("/api/v1/tasks/nonexistent-id")
        assert resp.status_code == 404

    def test_result_endpoints_do_not_500_when_result_is_none(self):
        """未成功完成的任务 result 为 None，结果类端点不得抛 AttributeError"""
        resp = self.client.post("/api/v1/tasks", json={"description": "result 为 None 的任务"})
        task_id = resp.json()["id"]
        assert self.client.get(f"/api/v1/tasks/{task_id}").json()["result"] is None

        for path, allowed in (
            ("posts", (200,)),
            ("stats", (200,)),
            ("export?format=xlsx", (404,)),
            ("export?format=csv", (404,)),
        ):
            r = self.client.get(f"/api/v1/tasks/{task_id}/{path}")
            assert r.status_code in allowed, f"{path} 返回 {r.status_code}，期望 {allowed}"

    def test_sentiment_endpoints_reject_unknown_task(self):
        """task_id 会被拼进文件路径，不存在的任务必须 404 而不是去读文件"""
        for path in ("sentiment", "export?format=xlsx"):
            r = self.client.get(f"/api/v1/tasks/no-such-task/{path}")
            assert r.status_code == 404, f"{path} 返回 {r.status_code}，期望 404"

    def test_sentiment_path_traversal_blocked(self):
        """%5C 在 Windows 上是路径分隔符，不得借此读到 data 目录之外的文件"""
        traversal = "..%5C..%5C..%5C..%5Cfrontend%5Cpackage"
        r = self.client.get(f"/api/v1/tasks/{traversal}/sentiment")
        assert r.status_code == 404
        assert "tweakers-scraper-frontend" not in r.text


class TestSourcesAPIEndToEnd:
    """数据源 CRUD 与凭据加密 — 真实 HTTP 请求 + 真实 SQLite 落盘"""

    PASSWORD = "sup3r-s3cret-pw"

    @classmethod
    def setup_class(cls):
        import app.config as cfg
        from cryptography.fernet import Fernet

        cls.cfg = cfg
        cls._old_api_key = cfg.settings.api_key
        cls._old_secret = cfg.settings.secret_key
        cfg.settings.api_key = ""
        cfg.settings.secret_key = Fernet.generate_key().decode()

        from main import app
        from app.services import storage

        # DB_PATH 是 import 时算好的常量，只改 data_dir 不生效；不重定向就会写进真实的
        # backend/data/hyxi.db —— 这个类单独跑时尤其致命
        cls.tmpdir = tempfile.mkdtemp()
        cls.storage = storage
        cls._old_db_path = storage.DB_PATH
        storage.DB_PATH = os.path.join(cls.tmpdir, "hyxi.db")
        storage.init_db()

        cls.client = TestClient(app)

    @classmethod
    def teardown_class(cls):
        cls.cfg.settings.api_key = cls._old_api_key
        cls.cfg.settings.secret_key = cls._old_secret
        cls.storage.DB_PATH = cls._old_db_path
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def _create(self, thread_id="2336074", name="测试源"):
        return self.client.post("/api/v1/sources", json={
            "collector_id": "tweakers",
            "name": name,
            "params": {"thread_id": thread_id},
        })

    def test_collector_catalog_drives_form_rendering(self):
        resp = self.client.get("/api/v1/collectors")
        assert resp.status_code == 200
        tweakers = next(c for c in resp.json() if c["id"] == "tweakers")
        assert tweakers["needs_credentials"] is False
        assert [f["name"] for f in tweakers["param_fields"]] == ["thread_id", "base_url"]
        assert [f["required"] for f in tweakers["param_fields"]] == [True, False]
        group = next(c for c in resp.json() if c["id"] == "facebook_group")
        assert group["incremental_strategy"] == "watermark"

    def test_internal_collector_hidden_from_catalog(self):
        """group_feed 是给 fixture 站点用的通用采集器（base_url 必填、没有真实站点），
        真实版本是 facebook_group。它不该出现在「新增数据源」的下拉框里，但仍要能被
        get_collector 解析 —— 既有数据源和增量回归测试都还在用它"""
        from app.collectors import get_collector

        ids = [c["id"] for c in self.client.get("/api/v1/collectors").json()]
        assert "group_feed" not in ids
        assert ids == ["tweakers", "facebook_group"]
        assert get_collector("group_feed").id == "group_feed"

    def test_source_crud(self):
        resp = self._create(name="Tweakers 主帖")
        assert resp.status_code == 201
        source = resp.json()
        assert source["collector_name"] == "Tweakers.net 论坛"
        assert source["params"]["thread_id"] == "2336074"
        assert source["has_credential"] is False

        sid = source["id"]
        assert self.client.get(f"/api/v1/sources/{sid}").json()["name"] == "Tweakers 主帖"

        resp = self.client.patch(f"/api/v1/sources/{sid}", json={"enabled": False})
        assert resp.status_code == 200
        assert resp.json()["enabled"] is False

        assert self.client.delete(f"/api/v1/sources/{sid}").status_code == 200
        assert self.client.get(f"/api/v1/sources/{sid}").status_code == 404

    def test_missing_required_param_rejected(self):
        resp = self.client.post("/api/v1/sources", json={
            "collector_id": "tweakers", "name": "缺参数", "params": {},
        })
        assert resp.status_code == 400
        assert "帖子 ID" in resp.json()["detail"]

    def test_unknown_collector_rejected(self):
        resp = self.client.post("/api/v1/sources", json={
            "collector_id": "myspace", "name": "不存在的采集器", "params": {},
        })
        assert resp.status_code == 400

    def test_undeclared_params_are_dropped(self):
        """params 会原样进 job 文件，未声明的键不能借这条路塞进采集脚本"""
        resp = self.client.post("/api/v1/sources", json={
            "collector_id": "tweakers", "name": "夹带私货",
            "params": {"thread_id": "2336074", "pacing": {"delay_min": 0}, "headless": False},
        })
        assert resp.status_code == 201
        assert set(resp.json()["params"]) == {"thread_id"}
        self.client.delete(f"/api/v1/sources/{resp.json()['id']}")

    def test_credential_never_leaves_backend_and_is_encrypted_at_rest(self):
        import sqlite3
        from app.services import storage

        sid = self._create(name="带凭据的源").json()["id"]
        try:
            resp = self.client.put(f"/api/v1/sources/{sid}/credential", json={
                "username": "tester@example.com", "password": self.PASSWORD,
            })
            assert resp.status_code == 200
            assert resp.json()["has_credential"] is True
            assert resp.json()["credential_username"] == "tester@example.com"

            # 任何出口都不得回显密码
            for path in ("/api/v1/sources", f"/api/v1/sources/{sid}"):
                body = self.client.get(path).text
                assert self.PASSWORD not in body, f"{path} 泄漏了密码明文"
                assert "secret_enc" not in body
                assert "password" not in body

            # 落盘的必须是密文
            conn = sqlite3.connect(storage.DB_PATH)
            try:
                row = conn.execute(
                    "SELECT secret_enc FROM credentials WHERE source_id = ?", (sid,)
                ).fetchone()
            finally:
                conn.close()
            assert row is not None
            assert self.PASSWORD not in row[0]

            # 但采集器启动路径能解回明文，否则登录根本没法用
            from app.services import source_service
            assert source_service.get_credential_secret(sid) == self.PASSWORD

            assert self.client.delete(f"/api/v1/sources/{sid}/credential").json()["has_credential"] is False
        finally:
            self.client.delete(f"/api/v1/sources/{sid}")

    def test_deleting_source_cascades_credential(self):
        sid = self._create(name="级联删除").json()["id"]
        self.client.put(f"/api/v1/sources/{sid}/credential", json={
            "username": "u", "password": self.PASSWORD,
        })
        self.client.delete(f"/api/v1/sources/{sid}")

        from app.services import storage
        assert storage.get_credential(sid) is None

    def test_editing_source_keeps_credential(self):
        """改名/停用不能把凭据带走。

        INSERT OR REPLACE 是「删旧行再插新行」，在 foreign_keys=ON 下会触发 credentials 的
        ON DELETE CASCADE —— 界面上点一次「保存」凭据就没了，且毫无提示。
        """
        from app.services import storage

        sid = self._create(name="改名前").json()["id"]
        try:
            self.client.put(f"/api/v1/sources/{sid}/credential", json={
                "username": "keep@example.com", "password": self.PASSWORD,
            })
            assert storage.get_credential(sid) is not None

            resp = self.client.patch(f"/api/v1/sources/{sid}", json={"name": "改名后"})
            assert resp.status_code == 200
            assert resp.json()["name"] == "改名后"
            assert resp.json()["has_credential"] is True

            resp = self.client.patch(f"/api/v1/sources/{sid}", json={"enabled": False})
            assert resp.json()["has_credential"] is True

            cred = storage.get_credential(sid)
            assert cred is not None
            assert cred["username"] == "keep@example.com"
            from app.services import source_service
            assert source_service.get_credential_secret(sid) == self.PASSWORD
        finally:
            self.client.delete(f"/api/v1/sources/{sid}")

    def test_validation_error_does_not_echo_password(self):
        """422 也不能把密码吐回来。

        FastAPI 默认把 pydantic 的 input（提交的原始值）序列化进 detail，
        一次超长校验失败就足以让明文密码进响应体、再被前端拦截器打进控制台。
        """
        sid = self._create(name="超长密码").json()["id"]
        try:
            canary = "LEAK-CANARY-" + "x" * 600
            resp = self.client.put(f"/api/v1/sources/{sid}/credential", json={
                "username": "u", "password": canary,
            })
            assert resp.status_code == 422
            assert "LEAK-CANARY" not in resp.text
            assert '"input"' not in resp.text
            # 报文仍要说清楚哪个字段错在哪，否则没法排查
            assert "password" in resp.text
        finally:
            self.client.delete(f"/api/v1/sources/{sid}")

    def test_credential_refused_without_secret_key(self):
        """没配密钥时必须报错，绝不能静默降级成明文落库"""
        sid = self._create(name="无密钥").json()["id"]
        old = self.cfg.settings.secret_key
        self.cfg.settings.secret_key = ""
        try:
            resp = self.client.put(f"/api/v1/sources/{sid}/credential", json={
                "username": "u", "password": self.PASSWORD,
            })
            assert resp.status_code == 400
            assert "TWEAKERS_SECRET_KEY" in resp.json()["detail"]

            from app.services import storage
            assert storage.get_credential(sid) is None
        finally:
            self.cfg.settings.secret_key = old
            self.client.delete(f"/api/v1/sources/{sid}")

    def test_seed_is_idempotent(self):
        from app.services import source_service

        for src in self.client.get("/api/v1/sources").json():
            self.client.delete(f"/api/v1/sources/{src['id']}")

        source_service.seed_default_sources()
        after_first = self.client.get("/api/v1/sources").json()
        assert len(after_first) == 1
        assert after_first[0]["collector_id"] == "tweakers"

        source_service.seed_default_sources()
        assert len(self.client.get("/api/v1/sources").json()) == 1

    # ===== 人工授权端点 =====
    # 这几条都不触发浏览器：真打 HTTP、真读库，只走「不该启动子进程」的那些分支。
    # 真正开浏览器的成功路径由 test_core.py 的 TestFacebookLoginEndToEnd 覆盖。

    def test_authorize_rejects_collector_that_needs_no_login(self):
        sid = self._create(name="不需要登录").json()["id"]
        try:
            resp = self.client.post(f"/api/v1/sources/{sid}/authorize")
            assert resp.status_code == 400
            assert "不需要登录" in resp.json()["detail"]
        finally:
            self.client.delete(f"/api/v1/sources/{sid}")

    def test_authorize_endpoints_404_on_unknown_source(self):
        assert self.client.post("/api/v1/sources/src_nope/authorize").status_code == 404
        assert self.client.get("/api/v1/sources/src_nope/authorize/events").status_code == 404

    def test_authorize_is_rejected_while_one_is_running(self):
        """同一个数据源不能同时开两个授权窗口"""
        from app.services import source_service

        resp = self.client.post("/api/v1/sources", json={
            "collector_id": "facebook_group", "name": "并发授权",
            "params": {"group_id": "123"},
        })
        sid = resp.json()["id"]
        # 直接把它标成「授权中」，等价于前一个授权还没结束
        source_service._authorizing.add(sid)
        try:
            resp = self.client.post(f"/api/v1/sources/{sid}/authorize")
            assert resp.status_code == 409
            assert "授权中" in resp.json()["detail"]
        finally:
            source_service._authorizing.discard(sid)
            self.client.delete(f"/api/v1/sources/{sid}")

    def test_authorization_result_is_visible_on_the_source(self):
        """授权成功后界面靠 last_auth_at 把徽标从「需重新授权」翻成「会话正常」"""
        from app.services import source_service

        resp = self.client.post("/api/v1/sources", json={
            "collector_id": "facebook_group", "name": "授权徽标",
            "params": {"group_id": "123"},
        })
        sid = resp.json()["id"]
        try:
            assert resp.json()["needs_credentials"] is True
            assert resp.json()["last_auth_at"] is None

            source_service.mark_authorized(sid)
            after = self.client.get(f"/api/v1/sources/{sid}").json()
            assert after["last_auth_at"], "授权成功没有反映到出口模型上"
            # 授权不该动凭据
            assert after["has_credential"] is False
        finally:
            self.client.delete(f"/api/v1/sources/{sid}")


class TestNestedPostsApiEndToEnd:
    """嵌套评论的出口：真实 HTTP 请求打到真实落盘的采集结果上"""

    @classmethod
    def setup_class(cls):
        import app.config as cfg

        cls.cfg = cfg
        cls._old_api_key = cfg.settings.api_key
        cfg.settings.api_key = ""

        from main import app
        from app.services import storage

        cls.tmpdir = tempfile.mkdtemp()
        cls.storage = storage
        cls._old_db_path = storage.DB_PATH
        storage.DB_PATH = os.path.join(cls.tmpdir, "hyxi.db")
        storage.init_db()
        cls.client = TestClient(app)

        # 两个来源各自一批真实结构的数据：论坛（无评论）+ 小组（主贴带评论）
        forum = [
            {"username": f"用户{i}", "timestamp": "22-05-2026 17:0{}".format(i),
             "content": f"论坛帖子{i}", "translation": "", "page_number": 1,
             "fingerprint": f"f{i}", "source": "src_forum",
             "parent_fingerprint": None, "reply_level": 0}
            for i in range(3)
        ]
        group = []
        for i in range(2):
            root_fp = f"g{i}"
            group.append({
                "username": f"楼主{i}", "timestamp": "02-06-2026 09:1{}".format(i),
                "content": f"小组主贴{i}", "translation": "", "page_number": 1,
                "fingerprint": root_fp, "source": "src_group",
                "parent_fingerprint": None, "reply_level": 0,
            })
            for j in range(2):
                group.append({
                    "username": f"回复者{i}{j}", "timestamp": "02-06-2026 10:0{}".format(j),
                    "content": f"评论{i}{j} 独特词{i}{j}", "translation": "", "page_number": 1,
                    "fingerprint": f"g{i}c{j}", "source": "src_group",
                    "parent_fingerprint": root_fp, "reply_level": 1,
                })
        storage.upsert_posts("src_forum", forum)
        storage.upsert_posts("src_group", group)

        # 任务里记下来源 id，之后即使把数据源删掉也还能按 id 读回来
        cls.task_id = "nested-e2e"
        from app.services.orchestrator import orchestrator
        orchestrator.tasks[cls.task_id] = {
            "id": cls.task_id, "status": "completed", "description": "嵌套出口",
            "plan": [], "logs": [], "progress": 1.0, "current_step": None,
            "result": {
                "total_posts": len(forum) + len(group),
                "sources": [
                    {"id": "src_forum", "name": "论坛来源", "collector_id": "tweakers",
                     "post_count": len(forum)},
                    {"id": "src_group", "name": "小组来源", "collector_id": "group_feed",
                     "post_count": len(group)},
                ],
            },
        }
        cls.orchestrator = orchestrator

    @classmethod
    def teardown_class(cls):
        cls.cfg.settings.api_key = cls._old_api_key
        cls.storage.DB_PATH = cls._old_db_path
        cls.orchestrator.tasks.pop(cls.task_id, None)
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def _posts(self, **params):
        return self.client.get(f"/api/v1/tasks/{self.task_id}/posts", params=params).json()

    def test_pagination_counts_root_posts_only(self):
        """分页粒度是主贴，评论跟着父贴走，不会被分页边界切断"""
        data = self._posts(page_size=200)
        assert data["total"] == 5          # 论坛 3 + 小组 2 个主贴
        assert len(data["posts"]) == 5
        with_replies = [p for p in data["posts"] if p["replies"]]
        assert len(with_replies) == 2
        assert all(len(p["replies"]) == 2 for p in with_replies)
        assert all(r["reply_level"] == 1 for p in with_replies for r in p["replies"])

    def test_index_is_absolute_and_never_overlaps_across_pages(self):
        """index 是扁平存储数组里的真实位置，舆情详情靠它做绝对对齐。

        一页只保证 page_size 个主贴，带上评论后条目数会超出 —— 按页内计数编号
        会让相邻两页的 index 区间重叠，详情弹窗于是显示错帖子。
        """
        page1 = self._posts(page=1, page_size=4)
        page2 = self._posts(page=2, page_size=4)

        def all_indices(payload):
            out = []
            def walk(p):
                out.append(p["index"])
                for c in p["replies"]:
                    walk(c)
            for p in payload["posts"]:
                walk(p)
            return out

        i1, i2 = all_indices(page1), all_indices(page2)
        # 主贴按时间倒序，第一页 4 个主贴是「小组主贴1、小组主贴0、论坛帖子2、论坛帖子1」，
        # 前两个各带 2 条评论
        assert len(i1) == 8, i1
        assert not (set(i1) & set(i2)), f"两页的 index 重叠了: {i1} vs {i2}"
        assert sorted(i1 + i2) == list(range(1, 10))

        # index 必须能直接反查到同一条帖子
        detail = self.client.get(
            f"/api/v1/tasks/{self.task_id}/posts/{i1[-1] - 1}"
        ).json()
        flat = [p for p in page1["posts"] for p in [p] + p["replies"]]
        assert detail["content"] == flat[-1]["content"]

    def test_roots_are_ordered_newest_first(self):
        """主贴按发表时间从新到旧。存储顺序是采集顺序，跟时间毫无关系"""
        data = self._posts(page_size=200)
        got = [p["content"] for p in data["posts"]]
        assert got == ["小组主贴1", "小组主贴0", "论坛帖子2", "论坛帖子1", "论坛帖子0"], got

        times = [p["timestamp"] for p in data["posts"]]
        assert times == sorted(times, reverse=True), times

        # 存储顺序（index）必须原封不动 —— 它是舆情结论的对齐锚点
        assert [p["index"] for p in data["posts"]] == [7, 4, 3, 2, 1]

    def test_sorting_uses_iso_not_raw_dutch_string(self):
        """落盘是 dd-mm-yyyy，直接按字符串排会变成「按日排先」

        07-01（1月7日）会排到 28-06（6月28日）前面，整个顺序错乱且看着还挺像回事。
        """
        self.storage.upsert_posts("src_forum", [{
            "username": "跨年用户", "timestamp": "07-01-2027 08:00",
            "content": "元月的帖子", "translation": "", "page_number": 1,
            "fingerprint": "jan", "source": "src_forum",
            "parent_fingerprint": None, "reply_level": 0,
        }])
        try:
            got = [p["content"] for p in self._posts(page_size=200)["posts"]]
            assert got[0] == "元月的帖子", got
        finally:
            conn = self.storage._get_conn()
            conn.execute("DELETE FROM posts WHERE fingerprint = 'jan'")
            conn.commit()
            conn.close()

    def test_posts_without_a_timestamp_sink_to_the_bottom(self):
        """早期采集读不到 tooltip 绝对时间，落盘留空。这些帖子不该霸占最前面"""
        self.storage.upsert_posts("src_forum", [{
            "username": "无时间用户", "timestamp": "",
            "content": "没有时间的帖子", "translation": "", "page_number": 1,
            "fingerprint": "notime", "source": "src_forum",
            "parent_fingerprint": None, "reply_level": 0,
        }])
        try:
            got = [p["content"] for p in self._posts(page_size=200)["posts"]]
            assert got[-1] == "没有时间的帖子", got
        finally:
            conn = self.storage._get_conn()
            conn.execute("DELETE FROM posts WHERE fingerprint = 'notime'")
            conn.commit()
            conn.close()

    def test_search_hit_on_comment_brings_back_its_root(self):
        data = self._posts(search="独特词10")
        assert data["total"] == 1
        root = data["posts"][0]
        assert root["matched"] is False, "父贴本身没命中，不该标 matched"
        assert root["content"] == "小组主贴1"
        hits = [r for r in root["replies"] if r["matched"]]
        assert [h["content"] for h in hits] == ["评论10 独特词10"]
        # 兄弟评论也一并返回，保住上下文
        assert len(root["replies"]) == 2

    def test_results_survive_source_deletion(self):
        """用户在数据源页删掉来源后，历史任务的结果不能静默变成空白"""
        # 这两个来源本来就没注册过（DB 是空的），等价于「已被删除」
        assert self.client.get("/api/v1/sources").json() == []
        data = self._posts(page_size=200)
        assert data["total"] == 5
        assert {p["source_name"] for p in data["posts"]} == {"论坛来源", "小组来源"}

        stats = self.client.get(f"/api/v1/tasks/{self.task_id}/stats").json()
        assert stats["total_posts"] == 9


class TestExportEndpointEndToEnd:
    """全站唯一的导出口 —— 真实 HTTP 打到真实落盘数据上"""

    @classmethod
    def setup_class(cls):
        import app.config as cfg

        cls.cfg = cfg
        cls.tmpdir = tempfile.mkdtemp()
        cls._old_key = cfg.settings.api_key
        cls._old_dir = cfg.settings.data_dir
        cfg.settings.api_key = ""
        cfg.settings.data_dir = cls.tmpdir

        from main import app
        from app.services import storage
        cls.storage = storage
        cls._old_db = storage.DB_PATH
        storage.DB_PATH = os.path.join(cls.tmpdir, "hyxi.db")
        storage.init_db()
        cls.client = TestClient(app)

        # 一个主贴带两条评论，再一个主贴 —— 存储顺序刻意与线程顺序不同，
        # order_by_thread 必然重排，对齐一旦写错就会被这批数据抓住
        cls.posts = [
            {"username": "楼主A", "timestamp": "02-06-2026 09:00", "content": "主贴A",
             "translation": "译A", "page_number": 1, "fingerprint": "a",
             "source": "src_x", "parent_fingerprint": None, "reply_level": 0},
            {"username": "楼主B", "timestamp": "03-06-2026 09:00", "content": "主贴B",
             "translation": "译B", "page_number": 1, "fingerprint": "b",
             "source": "src_x", "parent_fingerprint": None, "reply_level": 0},
            {"username": "回复1", "timestamp": "02-06-2026 10:00", "content": "评论A1",
             "translation": "译A1", "page_number": 1, "fingerprint": "a1",
             "source": "src_x", "parent_fingerprint": "a", "reply_level": 1},
            {"username": "回复2", "timestamp": "02-06-2026 11:00", "content": "评论A2",
             "translation": "译A2", "page_number": 1, "fingerprint": "a2",
             "source": "src_x", "parent_fingerprint": "a", "reply_level": 1},
        ]
        storage.upsert_posts("src_x", cls.posts)

        cls.task_id = "export-e2e"
        from app.services.orchestrator import orchestrator
        cls.orchestrator = orchestrator
        orchestrator.tasks[cls.task_id] = {
            "id": cls.task_id, "status": "completed", "description": "导出用任务",
            "plan": [], "logs": [], "progress": 1.0, "current_step": None,
            "result": {"total_posts": 4, "sources": [
                {"id": "src_x", "name": "小组来源:测试", "collector_id": "group_feed",
                 "post_count": 4},
            ]},
        }

    @classmethod
    def teardown_class(cls):
        cls.cfg.settings.api_key = cls._old_key
        cls.cfg.settings.data_dir = cls._old_dir
        cls.storage.DB_PATH = cls._old_db
        cls.orchestrator.tasks.pop(cls.task_id, None)
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def _write_sentiment(self, results):
        """入库时 results 仍按**扁平数组**下标给（下标来自 enumerate(all_posts)），
        存储层负责换成 (source_id, fingerprint)"""
        self.storage.save_sentiment(self.task_id, {
            "analyzed_at": "2026-08-05T10:00:00",
            "summary": {"top_dimensions": []},
            "results": results,
        }, self.posts)

    def _drop_sentiment(self):
        conn = self.storage._get_conn()
        try:
            conn.execute("DELETE FROM sentiment_runs WHERE task_id = ?", (self.task_id,))
            conn.execute("DELETE FROM sentiment_results WHERE task_id = ?", (self.task_id,))
            conn.commit()
        finally:
            conn.close()

    def _csv_rows(self):
        import csv as _csv
        from io import StringIO
        resp = self.client.get(f"/api/v1/tasks/{self.task_id}/export", params={"format": "csv"})
        assert resp.status_code == 200
        text = resp.content.decode("utf-8-sig")
        return resp, list(_csv.DictReader(StringIO(text)))

    def _filename(self, resp):
        from urllib.parse import unquote
        disposition = resp.headers["content-disposition"]
        return unquote(re.search(r"filename\*=utf-8''([^;]+)", disposition, re.I).group(1))

    def test_sentiment_follows_the_post_not_the_row_number(self):
        """**最容易写错的一处**：舆情下标对齐的是扁平数组，而明细按线程重排。

        先按下标建映射再排序才对。反过来（排完按行号取）每条帖子都会配上别人的结论，
        而表面上完全看不出异常 —— 所以这里用「每条帖子的结论里写着自己的名字」来钉死。
        """
        self._write_sentiment([
            {"sentiment": "positive", "intensity": 5, "reason_cn": "属于:主贴A", "dimensions": ["价格/性价比"]},
            {"sentiment": "negative", "intensity": 4, "reason_cn": "属于:主贴B", "dimensions": []},
            {"sentiment": "neutral", "intensity": 3, "reason_cn": "属于:评论A1", "dimensions": []},
            {"sentiment": "neutral", "intensity": 2, "reason_cn": "属于:评论A2", "dimensions": []},
        ])
        _resp, rows = self._csv_rows()

        # 排序确实发生了：存储顺序是 A,B,A1,A2，导出顺序是「主贴按时间倒序 + 评论跟着走」
        assert [r["原文"] for r in rows] == ["主贴B", "主贴A", "评论A1", "评论A2"]
        for row in rows:
            assert row["分析理由"] == f"属于:{row['原文']}", f"第 {row['序号']} 行串了：{row}"
        assert [r["情感"] for r in rows] == ["负面", "正面", "中立", "中立"]
        assert [r["层级"] for r in rows] == ["0", "0", "1", "1"]

    def test_rows_are_ordered_newest_first(self):
        """导出与页面用同一条排序规则：主贴按发表时间从新到旧，评论跟着自己的主贴走。

        主贴B(03-06) 比主贴A(02-06) 新，所以排在前面 —— 哪怕 A 的评论(10:00/11:00)
        比 B 的发表时间还晚。评论不参与主贴之间的排序。
        """
        self._drop_sentiment()
        _resp, rows = self._csv_rows()
        assert [r["原文"] for r in rows] == ["主贴B", "主贴A", "评论A1", "评论A2"]
        roots = [r["发布时间"] for r in rows if r["层级"] == "0"]
        assert roots == sorted(roots, reverse=True), roots

    def test_force_reanalyzes_posts_that_were_already_done(self):
        """增量按帖子身份跳过已分析的，改了分析口径后老帖子会永远停在旧结论上。

        实测库里 124/125 条都已分析、带图的 23 条**全部**已分析 —— 没有这个入口，
        接上图片理解之后在现有数据上一点变化都看不到。
        """
        self.storage.mark_sentiment_analyzed([
            dict(p, _processed={"sentiment_at": "2026-08-05T10:00:00"}) for p in self.posts
        ])
        try:
            # 不带 force：全都跳过
            body = self.client.post(f"/api/v1/tasks/{self.task_id}/sentiment").json()
            assert body["status"] == "completed", body
            assert "已完成舆情分析" in body["message"]

            # 带 force：4 条全部重新进入待分析
            body = self.client.post(
                f"/api/v1/tasks/{self.task_id}/sentiment?force=true"
            ).json()
            assert body["status"] == "started", body
            assert body["pending_count"] == 4, body
            assert "强制重新分析" in body["message"]
        finally:
            conn = self.storage._get_conn()
            try:
                conn.execute("UPDATE posts SET sentiment_at = NULL WHERE source_id = 'src_x'")
                conn.commit()
            finally:
                conn.close()
            self.orchestrator._sentiment_running.discard(self.task_id)

    def test_unanalyzed_posts_still_export(self):
        """只翻译没跑舆情的任务也得导得出来，否则唯一的导出口对它永远是死的"""
        self._drop_sentiment()
        resp, rows = self._csv_rows()
        assert resp.status_code == 200
        assert len(rows) == 4
        assert {r["情感"] for r in rows} == {"未分析"}
        assert {r["强度"] for r in rows} == {""}
        # 原文译文照常给全
        assert [r["中文翻译"] for r in rows] == ["译B", "译A", "译A1", "译A2"]

    def test_partial_sentiment_does_not_overrun(self):
        """增量分析进行到一半时 results 比 posts 短，越界的按未分析处理"""
        self._write_sentiment([
            {"sentiment": "positive", "intensity": 5, "reason_cn": "属于:主贴A", "dimensions": []},
            {"sentiment": "negative", "intensity": 4, "reason_cn": "属于:主贴B", "dimensions": []},
        ])
        _resp, rows = self._csv_rows()
        by_content = {r["原文"]: r for r in rows}
        assert by_content["主贴A"]["情感"] == "正面"
        assert by_content["主贴B"]["情感"] == "负面"
        assert by_content["评论A1"]["情感"] == "未分析"
        assert by_content["评论A2"]["情感"] == "未分析"

    def test_filename_carries_source_and_export_time(self):
        """文件名要能自证是哪个来源、哪次任务、什么时候导的"""
        from datetime import datetime

        for fmt in ("xlsx", "csv"):
            resp = self.client.get(f"/api/v1/tasks/{self.task_id}/export", params={"format": fmt})
            assert resp.status_code == 200
            name = self._filename(resp)
            assert name.startswith("HYXi舆情_分析报告_"), name
            assert name.endswith(f".{fmt}"), name
            # 来源名里的冒号是 Windows 非法字符，必须被剔掉
            assert "小组来源测试" in name, name
            assert ":" not in name, name
            assert self.task_id[:8] in name, name
            assert f"{datetime.now():%Y%m%d}" in name, name

    def test_multiple_sources_collapse_to_a_count(self):
        """来源一多就不逐个列，否则文件名能长到没法看"""
        task = self.orchestrator.tasks[self.task_id]
        original = task["result"]["sources"]
        task["result"]["sources"] = original + [
            {"id": "src_y", "name": "论坛来源", "collector_id": "tweakers",
             "post_count": 4},
        ]
        try:
            resp = self.client.get(f"/api/v1/tasks/{self.task_id}/export", params={"format": "csv"})
            assert "2个来源" in self._filename(resp)
        finally:
            task["result"]["sources"] = original

    def test_xlsx_is_a_real_workbook_with_both_sheets(self):
        from io import BytesIO
        from openpyxl import load_workbook

        self._write_sentiment([
            {"sentiment": "positive", "intensity": 5, "reason_cn": "好", "dimensions": ["价格/性价比"]},
        ])
        resp = self.client.get(f"/api/v1/tasks/{self.task_id}/export", params={"format": "xlsx"})
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.content))
        # 这批帖子都没有配图，所以不该多出一张「配图」表
        assert wb.sheetnames == ["概览", "帖子明细"]
        ws = wb["帖子明细"]
        assert ws.max_row == 5                       # 表头 + 4 条
        # 按列名定位，别写死列号 —— 加一列就得改一遍测试，改漏了还查不出来
        col = {c.value: c.column for c in ws[1]}
        assert {"原文", "中文翻译", "图片描述", "配图", "情感"} <= set(col)
        # 结论写给扁平数组第 0 条（主贴A），导出按时间倒序后它落在第 3 行
        assert ws.cell(3, col["原文"]).value == "主贴A"
        assert ws.cell(3, col["情感"]).value == "正面"

    def test_format_must_be_one_of_the_two(self):
        for bad in ("json", "pdf", "XLS", ""):
            resp = self.client.get(f"/api/v1/tasks/{self.task_id}/export", params={"format": bad})
            assert resp.status_code == 400, f"format={bad!r} 返回 {resp.status_code}"

    def test_default_format_is_xlsx(self):
        resp = self.client.get(f"/api/v1/tasks/{self.task_id}/export")
        assert resp.status_code == 200
        assert self._filename(resp).endswith(".xlsx")

    def test_old_download_endpoints_are_gone(self):
        """四个旧下载口已合并，留着等于让用户拿到不含舆情的半份数据"""
        for path in ("download", "export/csv", "export/json", "sentiment/download"):
            resp = self.client.get(f"/api/v1/tasks/{self.task_id}/{path}")
            assert resp.status_code == 404, f"{path} 仍然可用（{resp.status_code}）"


class TestImagesInExportAndApiEndToEnd:
    """纯图帖走完整条出口：真 HTTP → 真 SQLite → 真图片文件 → 真 xlsx / csv"""

    @classmethod
    def setup_class(cls):
        import app.config as cfg
        from PIL import Image as PILImage

        cls.cfg = cfg
        cls.tmpdir = tempfile.mkdtemp()
        cls._old_key = cfg.settings.api_key
        cls._old_dir = cfg.settings.data_dir
        cfg.settings.api_key = ""
        cfg.settings.data_dir = cls.tmpdir

        from main import app
        from app.services import storage
        cls.storage = storage
        cls._old_db = storage.DB_PATH
        storage.DB_PATH = os.path.join(cls.tmpdir, "hyxi.db")
        storage.init_db()
        cls.client = TestClient(app)

        media = os.path.join(cls.tmpdir, "media", "src_p")
        os.makedirs(media, exist_ok=True)
        PILImage.new("RGB", (600, 400), (20, 120, 90)).save(os.path.join(media, "shot_0.png"))

        cls.posts = [
            {"username": "koen", "timestamp": "08-08-2026 12:37", "content": "",
             "translation": "", "page_number": 1, "fingerprint": "pic",
             "source": "src_p", "parent_fingerprint": None, "reply_level": 0,
             "images": ["src_p/shot_0.png"], "image_desc": "安装检查报告，总分 88"},
            {"username": "bob", "timestamp": "08-08-2026 13:00", "content": "Zelfde hier.",
             "translation": "我也一样。", "page_number": 1, "fingerprint": "txt",
             "source": "src_p", "parent_fingerprint": "pic", "reply_level": 1},
        ]
        storage.upsert_posts("src_p", cls.posts)

        cls.task_id = "img-export-e2e"
        from app.services.orchestrator import orchestrator
        cls.orchestrator = orchestrator
        orchestrator.tasks[cls.task_id] = {
            "id": cls.task_id, "status": "completed", "description": "带图任务",
            "plan": [], "logs": [], "progress": 1.0, "current_step": None,
            "result": {"total_posts": 2, "sources": [
                {"id": "src_p", "name": "Facebook", "collector_id": "facebook_group",
                 "post_count": 2},
            ]},
        }

    @classmethod
    def teardown_class(cls):
        cls.cfg.settings.api_key = cls._old_key
        cls.cfg.settings.data_dir = cls._old_dir
        cls.storage.DB_PATH = cls._old_db
        cls.orchestrator.tasks.pop(cls.task_id, None)
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def test_image_only_post_survives_collection(self):
        """一个字都没有、只有一张图的帖子必须进得了库 —— 曾经在入库口被静默丢掉"""
        stored = self.storage.load_posts(["src_p"])
        assert [p["fingerprint"] for p in stored] == ["pic", "txt"]
        assert stored[0]["images"] == ["src_p/shot_0.png"]

    def test_posts_api_exposes_the_image_description(self):
        """页面靠它解释「这条没有正文的帖子凭什么得出这个结论」"""
        resp = self.client.get(f"/api/v1/tasks/{self.task_id}/posts")
        assert resp.status_code == 200
        root = resp.json()["posts"][0]
        assert root["images"] == ["src_p/shot_0.png"]
        assert root["image_desc"] == "安装检查报告，总分 88"

    def test_xlsx_carries_the_picture(self):
        from io import BytesIO
        from openpyxl import load_workbook
        from app.services.excel_service import IMAGE_SHEET

        resp = self.client.get(f"/api/v1/tasks/{self.task_id}/export",
                               params={"format": "xlsx"})
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.content))
        assert wb.sheetnames == ["概览", "帖子明细", IMAGE_SHEET]
        assert len(wb["帖子明细"]._images) == 1, "明细行里没有缩略图"
        assert len(wb[IMAGE_SHEET]._images) == 1, "配图表里没有大图"

        col = {c.value: c.column for c in wb["帖子明细"][1]}
        cell = wb["帖子明细"].cell(2, col["图片描述"])
        assert "安装检查报告" in str(cell.value)
        assert cell.hyperlink and IMAGE_SHEET in cell.hyperlink.location

    def test_csv_lists_the_image_path_as_text(self):
        import csv as _csv
        from io import StringIO

        resp = self.client.get(f"/api/v1/tasks/{self.task_id}/export",
                               params={"format": "csv"})
        assert resp.status_code == 200
        rows = list(_csv.DictReader(StringIO(resp.content.decode("utf-8-sig"))))
        assert rows[0]["配图"] == "src_p/shot_0.png", "CSV 里得留个能找到原图的线索"
        assert rows[0]["图片描述"] == "安装检查报告，总分 88"
        assert rows[1]["配图"] == ""


class TestStatsTimeRangeEndToEnd:
    """统计里的时间区间必须真的是最早和最晚"""

    @classmethod
    def setup_class(cls):
        import app.config as cfg

        cls.cfg = cfg
        cls._old_api_key = cfg.settings.api_key
        cfg.settings.api_key = ""

        from main import app
        from app.services import storage

        cls.tmpdir = tempfile.mkdtemp()
        cls.storage = storage
        cls._old_db_path = storage.DB_PATH
        storage.DB_PATH = os.path.join(cls.tmpdir, "hyxi.db")
        storage.init_db()
        cls.client = TestClient(app)

        # 存储顺序刻意不等于时间顺序 —— 信息流按时间倒序渲染，增量又往后追加，
        # 真实落盘文件本来就是乱的。最早的那条排在最后，最晚的排在中间。
        # 日期还跨了月：01-07 和 28-06 按落盘的 dd-mm-yyyy 字符串比大小会得出
        # 「7 月 1 日早于 6 月 28 日」，所以必须先归一化成 ISO 再比。
        posts = [
            {"username": "中", "timestamp": "30-06-2026 12:00", "content": "中间",
             "fingerprint": "t2", "source": "src_x", "page_number": 1,
             "parent_fingerprint": None, "reply_level": 0},
            {"username": "晚", "timestamp": "01-07-2026 08:00", "content": "最晚",
             "fingerprint": "t3", "source": "src_x", "page_number": 1,
             "parent_fingerprint": None, "reply_level": 0},
            {"username": "无时间", "timestamp": "", "content": "时间提取失败",
             "fingerprint": "t4", "source": "src_x", "page_number": 1,
             "parent_fingerprint": None, "reply_level": 0},
            {"username": "早", "timestamp": "28-06-2026 09:00", "content": "最早",
             "fingerprint": "t1", "source": "src_x", "page_number": 1,
             "parent_fingerprint": None, "reply_level": 0},
        ]
        # seq 按插入顺序给，读回来仍是这个乱序 —— 排序必须由 /stats 自己做
        storage.upsert_posts("src_x", posts)

        cls.task_id = "range-e2e"
        from app.services.orchestrator import orchestrator
        orchestrator.tasks[cls.task_id] = {
            "id": cls.task_id, "status": "completed", "description": "时间区间",
            "plan": [], "logs": [], "progress": 1.0, "current_step": None,
            "result": {
                "total_posts": len(posts),
                "sources": [{"id": "src_x", "name": "乱序来源", "collector_id": "group_feed",
                             "post_count": len(posts)}],
            },
        }
        cls.orchestrator = orchestrator

    @classmethod
    def teardown_class(cls):
        cls.cfg.settings.api_key = cls._old_api_key
        cls.storage.DB_PATH = cls._old_db_path
        cls.orchestrator.tasks.pop(cls.task_id, None)
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def test_time_range_is_min_and_max_not_first_and_last(self):
        """取的是最早/最晚，不是数组的首尾。

        实测线上出现过 time_range_start=2026-07-28、time_range_end=2026-07-10，
        开始比结束晚了 18 天 —— 因为直接取了 timestamps[0] 和 timestamps[-1]，
        而存储顺序是抓取顺序，与时间早晚无关。空时间戳不参与（真实数据里有一批
        旧版提取器留下的空时间，让它当上「最早」会把整个区间拉垮）。
        """
        stats = self.client.get(f"/api/v1/tasks/{self.task_id}/stats").json()

        assert stats["time_range_start"] == "2026-06-28 09:00"
        assert stats["time_range_end"] == "2026-07-01 08:00"
        assert stats["time_range_start"] <= stats["time_range_end"], "开始晚于结束"


class TestMediaEndpointEndToEnd:
    """正文图回读端点 —— 鉴权 + 路径穿越"""

    @classmethod
    def setup_class(cls):
        import app.config as cfg

        cls.cfg = cfg
        cls.tmpdir = tempfile.mkdtemp()
        cls._old_dir = cfg.settings.data_dir
        cls._old_key = cfg.settings.api_key
        cfg.settings.data_dir = cls.tmpdir
        cfg.settings.api_key = ""

        os.makedirs(os.path.join(cls.tmpdir, "media", "src_a"), exist_ok=True)
        with open(os.path.join(cls.tmpdir, "media", "src_a", "pic.png"), "wb") as f:
            f.write(b"\x89PNG\r\n\x1a\n" + b"0" * 32)
        # 采集器落图的目录旁边就是明文存着 LLA API Key 的 config.json
        with open(os.path.join(cls.tmpdir, "config.json"), "w", encoding="utf-8") as f:
            json.dump({"api_key": "sk-must-not-leak"}, f)

        from main import app
        cls.client = TestClient(app)

    @classmethod
    def teardown_class(cls):
        cls.cfg.settings.data_dir = cls._old_dir
        cls.cfg.settings.api_key = cls._old_key
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def test_serves_existing_image(self):
        resp = self.client.get("/api/v1/media/src_a/pic.png")
        assert resp.status_code == 200
        assert resp.content.startswith(b"\x89PNG")

    def test_missing_file_is_404(self):
        assert self.client.get("/api/v1/media/src_a/nope.png").status_code == 404

    def test_path_traversal_is_blocked(self):
        """`../../config.json` 里是明文 LLM API Key —— 穿越出去就等于把它送人。

        不回 403 而回 404：403 等于承认那个位置有东西。

        **必须用 URL 编码形式**：裸的 `../` 会被 httpx 在客户端就规范化掉，根本到不了
        端点，拿它当用例等于什么都没测（实测把校验整段禁用，裸 `../` 那版照样绿）。
        `%2e%2e%2f` 才会被框架解码后原样交到 rel_path 手里。
        """
        for evil in (
            "%2e%2e%2fconfig.json",
            "..%2fconfig.json",
            "src_a%2f..%2f..%2fconfig.json",
            "src_a/..%2f..%2fconfig.json",
        ):
            resp = self.client.get(f"/api/v1/media/{evil}")
            assert resp.status_code == 404, evil
            assert b"sk-must-not-leak" not in resp.content, evil

    def test_requires_api_key_when_configured(self):
        """<img> 带不了请求头，所以必须支持 ?api_key=（与 SSE 同一个口子）"""
        self.cfg.settings.api_key = "s3cr3t"
        try:
            assert self.client.get("/api/v1/media/src_a/pic.png").status_code == 401
            ok = self.client.get("/api/v1/media/src_a/pic.png?api_key=s3cr3t")
            assert ok.status_code == 200
        finally:
            self.cfg.settings.api_key = ""


class TestApiKeyAuthEndToEnd:
    """共享密钥认证 — 未配置时放行，配置后拦截"""

    @classmethod
    def setup_class(cls):
        cls.tmpdir = tempfile.mkdtemp()
        import app.config as cfg
        cls.cfg = cfg
        cls._old_key = cfg.settings.api_key

        from main import app
        cls.client = TestClient(app)

    @classmethod
    def teardown_class(cls):
        cls.cfg.settings.api_key = cls._old_key
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def test_open_when_key_not_configured(self):
        """漏配环境变量不该让既有部署整个不可用"""
        self.cfg.settings.api_key = ""
        assert self.client.get("/api/v1/tasks").status_code == 200

    def test_rejected_without_key(self):
        self.cfg.settings.api_key = "s3cr3t"
        try:
            resp = self.client.get("/api/v1/tasks")
            assert resp.status_code == 401
            resp = self.client.get("/api/v1/tasks", headers={"X-API-Key": "wrong"})
            assert resp.status_code == 401
        finally:
            self.cfg.settings.api_key = ""

    def test_accepted_with_key(self):
        self.cfg.settings.api_key = "s3cr3t"
        try:
            resp = self.client.get("/api/v1/tasks", headers={"X-API-Key": "s3cr3t"})
            assert resp.status_code == 200
            # 浏览器 EventSource 无法自定义请求头，只能走 query
            resp = self.client.get("/api/v1/tasks?api_key=s3cr3t")
            assert resp.status_code == 200
        finally:
            self.cfg.settings.api_key = ""

    def test_health_and_root_stay_public(self):
        """健康检查被监控系统调用，不能要求密钥"""
        self.cfg.settings.api_key = "s3cr3t"
        try:
            assert self.client.get("/api/health").status_code == 200
            assert self.client.get("/").status_code == 200
        finally:
            self.cfg.settings.api_key = ""

    def test_write_endpoints_also_protected(self):
        """覆写 LLM 密钥、植入定时任务这类写操作必须一并挡住"""
        self.cfg.settings.api_key = "s3cr3t"
        try:
            resp = self.client.post("/api/v1/config", json={
                "api_key": "sk-hijack", "base_url": "https://evil.test", "model_name": "x",
            })
            assert resp.status_code == 401
            resp = self.client.post("/api/v1/schedules", json={
                "description": "植入的定时任务", "interval": "daily", "time": "09:00",
            })
            assert resp.status_code == 401
        finally:
            self.cfg.settings.api_key = ""
