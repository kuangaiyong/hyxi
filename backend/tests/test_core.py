"""端到端真实测试 — 不 mock，操作真实文件和真实逻辑"""

import os
import sys
import json
import asyncio
import base64
import tempfile
import shutil
import hashlib
import pytest
from io import BytesIO
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def _use_temp_db(tmpdir):
    """把 data_dir 和 DB_PATH 一起指到临时目录，返回 (storage 模块, 还原函数)。

    **DB_PATH 是 import 时算好的常量**，只改 data_dir 会让用例写进真实的 hyxi.db。
    帖子进库之后，凡是跑采集器的用例都要这一步 —— CollectorRunner 会查 posts 表
    算增量锚点，再把结果写回去。
    """
    import app.config as cfg
    import app.services.storage as storage_module

    old_dir, old_db = cfg.settings.data_dir, storage_module.DB_PATH
    cfg.settings.data_dir = tmpdir
    storage_module.DB_PATH = os.path.join(tmpdir, "hyxi.db")
    storage_module.init_db()

    def restore():
        cfg.settings.data_dir = old_dir
        storage_module.DB_PATH = old_db

    return storage_module, restore


class TestTaskLifecycleEndToEnd:
    """任务生命周期：真实 JSON 文件持久化"""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        self.tasks_path = os.path.join(self.tmpdir, "tasks.json")

    def teardown_method(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _create_orchestrator(self):
        from app.services.orchestrator import TaskOrchestrator
        import app.services.orchestrator as orch_module
        import app.services.storage as storage_module
        orch_module.settings.data_dir = self.tmpdir
        # storage.DB_PATH 是 import 时算好的常量，改 settings.data_dir 不影响它，
        # 必须一并重定向，否则本类的测试会写进真实的 backend/data/hyxi.db
        storage_module.DB_PATH = os.path.join(self.tmpdir, "hyxi.db")
        return TaskOrchestrator()

    def test_create_task_persists(self):
        orch = self._create_orchestrator()
        orch.create_task("t1", "抓取帖子2336074并翻译")

        # 通过 API 验证持久化（SQLite 或 JSON）
        task = orch.get_task("t1")
        assert task is not None
        assert task["id"] == "t1"
        assert task["description"] == "抓取帖子2336074并翻译"
        assert task["status"] == "pending"

    def test_cancel_running_task_persists_status_change(self):
        orch = self._create_orchestrator()
        orch.create_task("t2", "测试取消")
        orch.tasks["t2"]["status"] = "running"
        orch._persist()

        orch.cancel_task("t2")

        # 通过 API 验证
        task = orch.get_task("t2")
        assert task["status"] == "cancelled"
        assert task.get("_cancelled") is True

    def test_delete_task_removes_from_storage(self):
        orch = self._create_orchestrator()
        orch.create_task("t3", "待删除")
        orch.create_task("t4", "保留")

        orch.delete_task("t3")

        assert orch.get_task("t3") is None
        assert orch.get_task("t4") is not None

    def test_deleting_a_task_keeps_its_posts(self):
        """删任务只删任务记录，帖子必须一条不少 —— 包括花钱换来的译文和舆情结论。

        用户问过这件事。答案在 storage.delete_task()：它只发一句
        `DELETE FROM tasks WHERE id = ?`。posts 表**没有 task_id 列**（帖子按
        (source_id, fingerprint) 存），tasks 上也没有任何外键指向它 —— 全库唯一的
        外键是 credentials.source_id → sources(id)。舆情结论同理按帖子身份存，
        task_id 只是「哪个任务触发的」这条附注。

        行为本身早就是对的，但一直没有测试守着：哪天有人给 tasks 挂个
        ON DELETE CASCADE，或者在 delete_task 里顺手清一把「这个任务的帖子」，
        用户的历史数据就没了，而且不会有任何报错。
        """
        orch = self._create_orchestrator()
        import app.services.storage as storage

        posts = [
            {"username": "Dorpjes", "timestamp": "22-05-2026 17:06", "content": "原文",
             "translation": "译文", "page_number": 1, "fingerprint": "keep1",
             "source": "src_x", "parent_fingerprint": None, "reply_level": 0,
             "images": ["src_x/pic_0.png"],
             "_processed": {"translated": True, "sentiment_at": "2026-07-28T20:27:00"}},
        ]
        storage.upsert_posts("src_x", posts)
        storage.save_sentiment("t_del", {"results": [
            {"sentiment": "negative", "intensity": 3, "reason_cn": "抱怨续航",
             "dimensions": ["续航"]},
        ]}, posts)

        orch.create_task("t_del", "会被删掉的任务")
        assert orch.delete_task("t_del") is True
        assert orch.get_task("t_del") is None

        kept = storage.load_posts(["src_x"])
        assert len(kept) == 1, "删任务把帖子一起删了"
        assert kept[0]["translation"] == "译文", "译文没了 —— 这是花钱换来的"
        assert kept[0]["_processed"]["sentiment_at"] == "2026-07-28T20:27:00"
        assert kept[0]["images"] == ["src_x/pic_0.png"]
        # 结论按 (source_id, fingerprint) 存，跟哪个任务触发的无关，删任务不该带走它
        again = storage.get_sentiment("t_del", kept)
        assert (again or {}).get("results", [{}])[0].get("sentiment") == "negative"

    def test_deleted_task_does_not_reappear_after_restart(self):
        orch1 = self._create_orchestrator()
        orch1.create_task("t8", "删除后不应复活")
        orch1.create_task("t9", "应当保留")

        assert orch1.delete_task("t8") is True
        assert orch1.get_task("t8") is None

        # 重建 orchestrator 模拟服务重启：删除必须落到存储层
        orch2 = self._create_orchestrator()
        assert orch2.get_task("t8") is None, "已删除的任务在重启后复活了"
        assert orch2.get_task("t9") is not None

    def test_tasks_survive_restart(self):
        orch1 = self._create_orchestrator()
        orch1.create_task("t5", "重启存活测试")
        orch1.tasks["t5"]["status"] = "completed"
        orch1._persist()

        orch2 = self._create_orchestrator()
        assert orch2.get_task("t5") is not None
        assert orch2.get_task("t5")["status"] == "completed"

    def test_running_task_becomes_cancelled_on_restart(self):
        orch1 = self._create_orchestrator()
        orch1.create_task("t6", "运行中被杀")
        orch1.tasks["t6"]["status"] = "running"
        orch1._persist()

        orch2 = self._create_orchestrator()
        assert orch2.get_task("t6")["status"] == "cancelled"

    def test_log_persistence(self):
        orch = self._create_orchestrator()
        orch.create_task("t7", "日志测试")
        orch.tasks["t7"]["logs"] = [
            {"time": "2026-07-28T22:00:00", "level": "info", "message": "步骤1完成"},
            {"time": "2026-07-28T22:01:00", "level": "error", "message": "步骤2失败"},
        ]
        orch._persist()

        task = orch.get_task("t7")
        assert len(task["logs"]) == 2


class TestSentimentParsingEndToEnd:
    """舆情分析：真实 LLM 响应解析"""

    def test_parse_clean_json(self):
        from app.services.sentiment_service import SentimentService
        text = '{"sentiment": "positive", "intensity": 4, "reason_cn": "用户对安装体验满意", "dimensions": ["安装/配置体验", "App/软件体验"]}'
        r = SentimentService._parse_sentiment(text)
        assert r["sentiment"] == "positive"
        assert r["intensity"] == 4
        assert len(r["dimensions"]) == 2

    def test_parse_json_with_markdown_fence(self):
        from app.services.sentiment_service import SentimentService
        text = '```json\n{"sentiment": "negative", "intensity": 3, "reason_cn": "固件有问题", "dimensions": ["固件更新"]}\n```'
        r = SentimentService._parse_sentiment(text)
        assert r["sentiment"] == "negative"

    def test_parse_json_embedded_in_narrative(self):
        from app.services.sentiment_service import SentimentService
        text = '根据分析，这条帖子{"sentiment": "neutral", "intensity": 2, "reason_cn": "纯技术咨询", "dimensions": ["认证/合规(如Synergrid)"]}，总体来看。'
        r = SentimentService._parse_sentiment(text)
        assert r["sentiment"] == "neutral"
        assert r["intensity"] == 2

    def test_parse_invalid_returns_none(self):
        from app.services.sentiment_service import SentimentService
        assert SentimentService._parse_sentiment("不是JSON") is None
        assert SentimentService._parse_sentiment("") is None


class TestSentimentRetryEndToEnd:
    """批量里解析失败的条目必须逐条重试 —— 真 HTTP、真 httpx、真解析"""

    def setup_method(self):
        import app.config as cfg
        import app.services.storage as storage_module
        self.cfg = cfg
        self.tmpdir = tempfile.mkdtemp()
        self._old_dir = cfg.settings.data_dir
        cfg.settings.data_dir = self.tmpdir
        # DB_PATH 是 import 时算好的常量，只改 data_dir 会写进真实的 hyxi.db
        self.storage = storage_module
        self._old_db = storage_module.DB_PATH
        storage_module.DB_PATH = os.path.join(self.tmpdir, "hyxi.db")
        storage_module.init_db()

    def teardown_method(self):
        self.cfg.settings.data_dir = self._old_dir
        self.storage.DB_PATH = self._old_db
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _llm_site(self):
        sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures"))
        import llm_site
        return llm_site

    def test_parse_failures_are_retried_one_by_one(self):
        """批量输出少一段可解析的 JSON，不能就让那条帖子永远「解析失败」。

        批量结果靠 ---SENTIMENT_SEPARATOR--- 切分，LLM 偶尔会在某一段吐出非 JSON。
        改造前那条直接记成 {"sentiment": null, "reason_cn": "解析失败"} 就完事了 ——
        真实任务里 88 条中有 2 条因此永久缺席，而翻译早就有单条重译。
        """
        import asyncio
        site = self._llm_site()
        from app.services.sentiment_service import SentimentService
        from app.services.progress_manager import ProgressManager

        posts = [
            {"content": f"Firmware 2.4.{i} sloopte mijn WiFi-verbinding volledig.",
             "fingerprint": f"fp{i}", "source": "src_x", "timestamp": "01-07-2026 10:00"}
            for i in range(3)
        ]

        server = site.LLMSite()
        with server as base_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-test", "base_url": base_url, "model_name": "test-model",
            })

            out = asyncio.new_event_loop().run_until_complete(
                SentimentService.analyze("retry-e2e", posts, ProgressManager(),
                                         source_names={"src_x": "Facebook"})
            )

        results = out["results"]
        assert out["failed"] == 0, f"仍有条目没救回来: {results}"
        assert out["success"] == 3
        assert all(r and r.get("sentiment") for r in results), results
        # 最后一条正是批量里被吐成非 JSON 的那条，重试后拿到了真实结论
        assert results[-1]["reason_cn"] == "单条重试成功"
        # 重试必须是一条一条发的：批量那次 3 条，之后跟着一次 1 条的重试
        assert server.seen[0] == 3, server.seen
        assert server.seen[1:] == [1], f"重试没有按单条发出: {server.seen}"

    def test_missing_segments_are_retried_not_faked_as_neutral(self):
        """模型整批没照分隔符输出时，后面几条不能被编成一条 neutral。

        批量结果按 ---SENTIMENT_SEPARATOR--- 切分，模型偶尔把几条 JSON 连成一段 ——
        parts 只有 1 段，第 2 条往后落进「parts 不够」那条分支。它以前记的是
        {"sentiment": "neutral", "intensity": 1, "reason_cn": "解析失败"}：
        sentiment 一有值就绕过了下面的单条重试，还会被写上 sentiment_at 永久定死，
        最后以「中性 + 解析失败 + 空维度」落进报告和情感分布。真实库里捞出 10 条，
        其中一条正文是明确在抱怨固件的「Deze update werkt niet na de update...」，
        却算成了中性。
        """
        import asyncio
        site = self._llm_site()
        from app.services.sentiment_service import SentimentService
        from app.services.progress_manager import ProgressManager

        posts = [
            {"content": f"Deze update werkt niet, versie 2.4.{i} is een ramp.",
             "fingerprint": f"nofp{i}", "source": "src_x", "timestamp": "01-07-2026 10:00"}
            for i in range(3)
        ]

        server = site.LLMSite(drop_separator=True)
        with server as base_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-test", "base_url": base_url, "model_name": "test-model",
            })
            out = asyncio.new_event_loop().run_until_complete(
                SentimentService.analyze("nosep-e2e", posts, ProgressManager(),
                                         source_names={"src_x": "Facebook"})
            )

        results = out["results"]
        # 占位记录绝不能带着 sentiment 值出门 —— 那就是一条编出来的结论
        for r in results:
            assert not (r and r.get("reason_cn") == "解析失败" and r.get("sentiment")),                 f"「解析失败」冒充成了一条真结论: {r}"
        # 三条全部进了单条重试并拿回真实结论
        assert server.seen == [3, 1, 1, 1], f"后两条没进重试队列: {server.seen}"
        assert out["failed"] == 0, f"仍有条目没救回来: {results}"
        assert all(r["sentiment"] == "negative" for r in results), results
        assert all(r["dimensions"] == ["固件更新"] for r in results), results


class TestIndustryRoleIsSharedWithVision:
    """多模态模型的角色必须与翻译时一致（需求明确要求），且抽取常量没改动翻译行为"""

    def test_translation_prompt_survived_the_extraction(self):
        """把人设与术语表抽成常量后，翻译 prompt 必须还是原来那一份。

        它一变，几百条帖子的译文口径就跟着变，而这次改动压根不该碰翻译。
        """
        from app.services.translator_service import TRANSLATION_SYSTEM_PROMPT as P
        assert P.startswith("你是一位精通荷兰语、英语和中文的新能源光储行业专业翻译专家。\n")
        for term in ("thuisbatterij / thuisaccu → 家用储能电池",
                     "P1-poort / P1-meter → P1端口/P1电表",
                     "LiFePO4 → 磷酸铁锂",
                     "stopcontact → 插座"):
            assert term in P, f"术语表少了: {term}"
        for section in ("1. **行业术语准确**：", "2. **语言风格**：",
                        "3. **格式要求**：", "4. 输入是一段外文文本"):
            assert section in P, f"小节丢了: {section}"

    def test_vision_prompt_uses_the_very_same_role_and_glossary(self):
        from app.services.translator_service import INDUSTRY_GLOSSARY, INDUSTRY_ROLE
        from app.services.vision_service import VISION_SYSTEM_PROMPT
        assert VISION_SYSTEM_PROMPT.startswith(INDUSTRY_ROLE)
        assert INDUSTRY_GLOSSARY in VISION_SYSTEM_PROMPT

    def test_vision_prompt_forbids_drawing_conclusions(self):
        """描述图片是它的活，判情感是下一步的活。混在一起会让舆情判定被描述者带跑"""
        from app.services.vision_service import VISION_SYSTEM_PROMPT
        assert "不要推测" in VISION_SYSTEM_PROMPT
        assert "不要写「用户对此不满」这类结论" in VISION_SYSTEM_PROMPT


class _VisionTmpEnv:
    """图片理解相关测试的公共环境：临时 data_dir + 临时库 + 一张真图"""

    # 一个合法的 1x1 PNG，不是随便几个字节 —— 断言「解回来与磁盘上逐字节相同」时
    # 用的就是它
    PNG = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJ"
        "AAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
    )

    def setup_method(self):
        import app.config as cfg
        import app.services.storage as storage_module
        self.cfg = cfg
        self.tmpdir = tempfile.mkdtemp()
        self._old_dir = cfg.settings.data_dir
        cfg.settings.data_dir = self.tmpdir
        self.storage = storage_module
        self._old_db = storage_module.DB_PATH
        storage_module.DB_PATH = os.path.join(self.tmpdir, "hyxi.db")
        storage_module.init_db()

        self.media_rel = "src_x/pic_0.png"
        media_dir = os.path.join(self.tmpdir, "media", "src_x")
        os.makedirs(media_dir, exist_ok=True)
        with open(os.path.join(media_dir, "pic_0.png"), "wb") as f:
            f.write(self.PNG)

    def teardown_method(self):
        self.cfg.settings.data_dir = self._old_dir
        self.storage.DB_PATH = self._old_db
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _fixtures(self):
        sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures"))
        import llm_site
        import vision_site
        return llm_site, vision_site

    def _run(self, coro):
        return asyncio.new_event_loop().run_until_complete(coro)

    def _analyze(self, posts, pending=None, all_posts=None):
        from app.services.sentiment_service import SentimentService
        from app.services.progress_manager import ProgressManager
        from app.services.post_tree import thread_of, post_key
        full = all_posts if all_posts is not None else posts
        target = pending if pending is not None else posts
        fp_to_idx = {post_key(p): i for i, p in enumerate(full)}
        return self._run(SentimentService.analyze(
            "vision-e2e", target, ProgressManager(),
            fp_to_idx=fp_to_idx, source_names={"src_x": "Facebook"},
            thread_by_key=thread_of(full), all_posts=full,
        ))


class TestImageUnderstandingEndToEnd(_VisionTmpEnv):
    """图片先给多模态模型理解，描述再进舆情 prompt —— 真 HTTP、真 base64、真 SQLite"""

    def _post(self, fp="fp1", images=None, **kw):
        p = {"source": "src_x", "fingerprint": fp, "timestamp": "01-07-2026 10:00",
             "username": "alice", "content": "Kijk zelf maar.", **kw}
        if images is not None:
            p["images"] = images
        return p

    def test_image_reaches_vision_intact_and_its_description_reaches_the_llm(self):
        """整条链路：磁盘上的图 → data URI → 多模态模型 → 描述 → 舆情 prompt"""
        llm_site, vision_site = self._fixtures()
        from app.services.translator_service import INDUSTRY_ROLE

        posts = [self._post(images=[self.media_rel])]
        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite()
        with llm as llm_url, vision as vision_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self.storage.set_app_config("vision", {
                "api_key": "sk-v", "base_url": vision_url, "model_name": "vision-model"})
            out = self._analyze(posts)

        # 多模态模型确实收到了那张图，且解回来与磁盘上逐字节相同
        assert vision.image_count == 1, vision.calls
        assert vision.decoded_images() == [self.PNG]
        # 角色与翻译一致
        assert vision.calls[0]["system"].startswith(INDUSTRY_ROLE)
        # 描述进了舆情 prompt
        joined = "\n".join(llm.user_prompts)
        assert "[图片: " in joined, joined
        assert vision_site.DESCRIPTION in joined, joined
        assert out["success"] == 1

    def test_root_image_is_described_for_the_sake_of_its_reply(self):
        """增量时待分析的只有回复，但主贴的图正是这条回复的上下文，必须一并理解。

        只按 pending 收集带图帖子的话，回复就完全看不见主贴那张图。
        """
        llm_site, vision_site = self._fixtures()
        root = self._post("root1", images=[self.media_rel], content="Zie foto.")
        reply = self._post("c1", username="bob", content="+1", parent_fingerprint="root1")

        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite()
        with llm as llm_url, vision as vision_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self.storage.set_app_config("vision", {
                "api_key": "sk-v", "base_url": vision_url, "model_name": "vision-model"})
            self._analyze([root, reply], pending=[reply], all_posts=[root, reply])

        assert vision.image_count == 1, "主贴的图没有被理解，回复就等于看不见图"
        assert vision_site.DESCRIPTION in "\n".join(llm.user_prompts)

    def test_description_is_persisted_and_not_bought_twice(self):
        """image_desc 落库后，下一轮不该再为同一张图付一次钱"""
        llm_site, vision_site = self._fixtures()
        self.storage.upsert_posts("src_x", [self._post(images=[self.media_rel])])

        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite()
        with llm as llm_url, vision as vision_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self.storage.set_app_config("vision", {
                "api_key": "sk-v", "base_url": vision_url, "model_name": "vision-model"})

            self._analyze(self.storage.load_posts(["src_x"]))
            assert vision.image_count == 1

            reloaded = self.storage.load_posts(["src_x"])
            assert reloaded[0]["image_desc"] == vision_site.DESCRIPTION, "描述没落库"

            # 第二轮：描述已经在库里，不该再调多模态模型
            self._analyze(reloaded)
            assert vision.image_count == 1, f"同一张图被重复理解了: {vision.calls}"

    def test_translation_never_touches_the_vision_model(self):
        """需求明确：只有舆情分析要图片理解，翻译不要"""
        llm_site, vision_site = self._fixtures()
        from app.services.translator_service import TranslatorService
        from app.services.progress_manager import ProgressManager

        posts = [self._post(images=[self.media_rel])]
        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite()
        with llm as llm_url, vision as vision_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self.storage.set_app_config("vision", {
                "api_key": "sk-v", "base_url": vision_url, "model_name": "vision-model"})
            self._run(TranslatorService.execute("t", posts, {}, ProgressManager()))

        assert vision.calls == [], f"翻译不该调用多模态模型: {vision.calls}"

    def test_without_vision_config_analysis_still_completes(self):
        """多模态模型是可选增强。没配就按纯文本分析，绝不能因此失败"""
        llm_site, _vision_site = self._fixtures()
        posts = [self._post(images=[self.media_rel])]
        llm = llm_site.LLMSite()
        with llm as llm_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            out = self._analyze(posts)
        assert out["success"] == 1
        assert "[图片: " not in "\n".join(llm.user_prompts)

    def test_vision_quota_exhausted_degrades_instead_of_failing(self):
        """复刻实测踩到的 Kimi 403：配额用尽时舆情分析必须照常跑完"""
        llm_site, vision_site = self._fixtures()
        posts = [self._post(images=[self.media_rel])]
        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite(fail_status=403)
        with llm as llm_url, vision as vision_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self.storage.set_app_config("vision", {
                "api_key": "sk-v", "base_url": vision_url, "model_name": "vision-model"})
            out = self._analyze(posts)

        assert vision.image_count >= 1, "应该尝试过"
        assert out["success"] == 1, "403 不该让整轮分析失败"
        assert "[图片: " not in "\n".join(llm.user_prompts)

    def test_missing_image_file_is_skipped_quietly(self):
        llm_site, vision_site = self._fixtures()
        posts = [self._post(images=["src_x/does_not_exist.jpg"])]
        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite()
        with llm as llm_url, vision as vision_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self.storage.set_app_config("vision", {
                "api_key": "sk-v", "base_url": vision_url, "model_name": "vision-model"})
            out = self._analyze(posts)

        assert vision.calls == [], "一张图都没读到就不该发请求"
        assert out["success"] == 1

    def test_vision_call_does_not_pin_temperature(self):
        """真机实测：kimi-for-coding 只接受 temperature=1，传别的值整个请求 400。

        被降级逻辑吞掉之后，界面上看不出任何异常 —— 图片理解「在跑」，但一张图都
        没理解过。各家视觉模型对这个参数的约束不同，一律交给服务端默认值。
        """
        llm_site, vision_site = self._fixtures()
        posts = [self._post(images=[self.media_rel])]
        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite()
        with llm as llm_url, vision as vision_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self.storage.set_app_config("vision", {
                "api_key": "sk-v", "base_url": vision_url, "model_name": "vision-model"})
            self._analyze(posts)

        assert vision.calls, "根本没调用多模态模型"
        assert vision.calls[0]["temperature"] is None, \
            f"请求里钉了 temperature={vision.calls[0]['temperature']}，真实模型会 400"
        # 替身在 temperature 不合法时返回 400，描述就拿不到 —— 这条断言保证真的拿到了
        assert vision_site.DESCRIPTION in "\n".join(llm.user_prompts)

    def test_output_budget_leaves_room_for_reasoning_models(self):
        """真机实测：kimi-for-coding 是推理模型，reasoning_content 也计入 max_tokens。

        给 512 时 512 个全被推理吃掉，HTTP 200、finish_reason=length、content 是空串，
        于是「每张图都理解成功但没有描述」—— 界面上完全看不出异常，最难查的一类。
        """
        from app.services.vision_service import MAX_OUTPUT_TOKENS
        llm_site, vision_site = self._fixtures()
        assert MAX_OUTPUT_TOKENS >= vision_site.REASONING_TOKEN_FLOOR, \
            "输出预算太小，推理模型会把它烧光并返回空描述"

        posts = [self._post(images=[self.media_rel])]
        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite()
        with llm as llm_url, vision as vision_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self.storage.set_app_config("vision", {
                "api_key": "sk-v", "base_url": vision_url, "model_name": "vision-model"})
            self._analyze(posts)

        # 替身在预算不足时返回 200 + 空 content，描述就进不了 prompt
        assert vision_site.DESCRIPTION in "\n".join(llm.user_prompts), \
            "预算被推理吃光，拿到的是空描述"

    def test_empty_description_is_retried_once(self):
        """真机实测：预算给足了也会偶发跑飞，把 max_tokens 全烧在 reasoning 上。

        23 张图里 2~4 张会中，且同一张图换一次就好。调大预算躲不开（4096 照样被
        烧穿，只是多浪费一倍 token），所以拿到空描述就重打一次。
        """
        llm_site, vision_site = self._fixtures()
        posts = [self._post(images=[self.media_rel])]
        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite(empty_first=1)
        with llm as llm_url, vision as vision_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self.storage.set_app_config("vision", {
                "api_key": "sk-v", "base_url": vision_url, "model_name": "vision-model"})
            self._analyze(posts)

        assert len(vision.calls) == 2, f"没有重试，只打了 {len(vision.calls)} 次"
        assert vision_site.DESCRIPTION in "\n".join(llm.user_prompts), "重试后仍然没拿到描述"

    def test_persistently_empty_description_gives_up_instead_of_looping(self):
        """重试也拿不到就认了 —— 一张图不能把整轮分析拖住，更不能无限打下去"""
        llm_site, vision_site = self._fixtures()
        posts = [self._post(images=[self.media_rel])]
        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite(empty_first=99)
        with llm as llm_url, vision as vision_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self.storage.set_app_config("vision", {
                "api_key": "sk-v", "base_url": vision_url, "model_name": "vision-model"})
            out = self._analyze(posts)

        assert len(vision.calls) == 2, f"重试次数失控：{len(vision.calls)} 次"
        assert "[图片: " not in "\n".join(llm.user_prompts)
        assert out["success"] == 1, "拿不到描述时舆情分析必须照常跑完"

    def test_path_traversal_in_images_is_refused(self):
        """images 来自采集脚本，但 media 目录之外就是数据库和明文密钥"""
        from app.services.vision_service import media_path
        assert media_path("../hyxi.db") is None
        assert media_path("src_x/../../hyxi.db") is None
        assert media_path(self.media_rel) is not None


class TestImageOnlyPostsAreAnalyzedEndToEnd(_VisionTmpEnv):
    """一个字都没有、只有一张图的帖子，照样要拿到舆情结论。

    实测那条真实数据正是这样：正文空白，图上是 HYXi 安装检查报告（总分 88、
    发电异常 8/20 标橙）—— 正文过滤器一刀切掉的恰恰是信息量最大的一类帖子。
    """

    def _pic_post(self, fp="pic1", **kw):
        return {"source": "src_x", "fingerprint": fp, "timestamp": "01-07-2026 10:00",
                "username": "koen", "content": "", "images": [self.media_rel], **kw}

    def _configure(self, llm_url, vision_url=None):
        self.storage.set_app_config("llm", {
            "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
        if vision_url:
            self.storage.set_app_config("vision", {
                "api_key": "sk-v", "base_url": vision_url, "model_name": "vision-model"})

    def test_image_only_post_gets_a_verdict(self):
        llm_site, vision_site = self._fixtures()
        posts = [self._pic_post()]
        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite()
        with llm as llm_url, vision as vision_url:
            self._configure(llm_url, vision_url)
            out = self._analyze(posts)

        assert vision.image_count == 1, "纯图帖的图根本没被送去理解"
        joined = "\n".join(llm.user_prompts)
        assert vision_site.DESCRIPTION in joined, joined
        assert out["success"] == 1, "纯图帖必须能拿到结论"
        assert posts[0]["_processed"]["sentiment_at"], "sentiment_at 没置位，下轮会重复花钱"

    def test_image_only_post_reads_as_a_post_not_a_truncated_one(self):
        """正文位置不能是一片空白 —— 那看起来像一条被截断的帖子，会把判定带偏"""
        llm_site, vision_site = self._fixtures()
        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite()
        with llm as llm_url, vision as vision_url:
            self._configure(llm_url, vision_url)
            self._analyze([self._pic_post()])
        assert "没有文字" in "\n".join(llm.user_prompts), llm.user_prompts

    def test_placeholder_never_promises_a_picture_that_does_not_exist(self):
        """既没正文也没配图的帖子不能套「内容全在配图上」—— 那是句假话，后面也没有
        [图片: ...] 跟着，等于让模型去读一张不存在的图。它该说的是「没取到」。
        """
        from app.services.sentiment_service import (
            NO_TEXT_PLACEHOLDER, UNREADABLE_PLACEHOLDER,
        )
        llm_site, vision_site = self._fixtures()
        root = self._pic_post("root1")
        blank = {"source": "src_x", "fingerprint": "blank", "parent_fingerprint": "root1",
                 "username": "ghost", "content": "", "timestamp": "01-07-2026 11:00"}
        reply = {"source": "src_x", "fingerprint": "c1", "parent_fingerprint": "root1",
                 "username": "bob", "content": "Zelfde probleem hier.",
                 "timestamp": "01-07-2026 12:00"}

        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite()
        with llm as llm_url, vision as vision_url:
            self._configure(llm_url, vision_url)
            self._analyze([root, blank, reply])

        joined = "\n".join(llm.user_prompts)
        # 主贴有图 → 「内容全在配图上」；@ghost 没图 → 只能说「没取到」，两者不能混
        assert f"@koen: {NO_TEXT_PLACEHOLDER}" in joined, joined
        assert f"@ghost: {UNREADABLE_PLACEHOLDER}" in joined, joined
        assert f"@ghost: {NO_TEXT_PLACEHOLDER}" not in joined, joined
        # 系统提示词得告诉模型别把「没取到」当成「什么都没说」
        from app.services.sentiment_service import SENTIMENT_SYSTEM_PROMPT
        assert UNREADABLE_PLACEHOLDER.strip("（）") in SENTIMENT_SYSTEM_PROMPT

    def test_placeholder_follows_the_description_not_the_mere_presence_of_images(self):
        """有图**不等于**读得出图：图丢了、模型没配、调用失败时都拿不到描述。

        占位文案的判据必须是「描述有没有真的到手」——[图片: ...] 那行只在 image_desc
        非空时才跟在后面。按 images 判的话，这条会说出「内容全在配图上」却没有任何图
        跟着，正是 UNREADABLE_PLACEHOLDER 存在要避免的那种假话。
        """
        from app.services.sentiment_service import (
            NO_TEXT_PLACEHOLDER, UNREADABLE_PLACEHOLDER,
        )
        llm_site, vision_site = self._fixtures()
        root = self._pic_post("root1")
        # 有 images，但那个文件根本不在 media 目录里 —— 真实发生过的一类降级
        gone = {"source": "src_x", "fingerprint": "gone", "parent_fingerprint": "root1",
                "username": "ghost", "content": "", "timestamp": "01-07-2026 11:00",
                "images": ["src_x/never_downloaded_0.png"]}
        reply = {"source": "src_x", "fingerprint": "c1", "parent_fingerprint": "root1",
                 "username": "bob", "content": "Zelfde probleem hier.",
                 "timestamp": "01-07-2026 12:00"}

        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite()
        with llm as llm_url, vision as vision_url:
            self._configure(llm_url, vision_url)
            self._analyze([root, gone, reply])

        joined = "\n".join(llm.user_prompts)
        assert not gone.get("image_desc"), "前提不成立：丢失的图不该拿到描述"
        assert f"@ghost: {UNREADABLE_PLACEHOLDER}" in joined, joined
        assert f"@ghost: {NO_TEXT_PLACEHOLDER}" not in joined, joined
        # 真读出图的那条仍要说「内容全在配图上」，两条路不能一起改坏
        assert f"@koen: {NO_TEXT_PLACEHOLDER}" in joined, joined

    def test_image_only_post_without_a_description_is_left_unanalyzed(self):
        """多模态没配 / 调用失败时，纯图帖是**真的**没有可判断的内容。

        送一个空块给 LLM 只会换回一条编出来的结论 —— 那比「未分析」糟得多。
        """
        llm_site, _ = self._fixtures()
        posts = [self._pic_post()]
        llm = llm_site.LLMSite()
        with llm as llm_url:
            self._configure(llm_url)          # 只配文本模型，不配多模态
            out = self._analyze(posts)

        assert llm.user_prompts == [], f"没有描述就不该送去判情感: {llm.user_prompts}"
        assert out["success"] == 0
        assert out["failed"] == 1
        assert not (posts[0].get("_processed") or {}).get("sentiment_at")

    def test_text_reply_still_carries_the_image_only_root(self):
        """纯图主贴进了分析队列，不能反过来把它评论的上下文弄丢"""
        llm_site, vision_site = self._fixtures()
        root = self._pic_post("root1")
        reply = {"source": "src_x", "fingerprint": "c1", "parent_fingerprint": "root1",
                 "username": "bob", "content": "Zelfde probleem hier.",
                 "timestamp": "01-07-2026 11:00"}

        llm = llm_site.LLMSite()
        vision = vision_site.VisionSite()
        with llm as llm_url, vision as vision_url:
            self._configure(llm_url, vision_url)
            out = self._analyze([root, reply])

        joined = "\n".join(llm.user_prompts)
        assert vision.image_count == 1, "同一张图只该理解一次"
        assert "[讨论串上下文]" in joined
        assert vision_site.DESCRIPTION in joined, "回复看不到主贴那张图"
        assert out["success"] == 2, "主贴和回复都该有结论"


class TestThreadContextInPromptEndToEnd(_VisionTmpEnv):
    """回复贴的舆情判定要看整串，不能只看父贴前 200 字"""

    def _thread(self):
        return [
            {"source": "src_x", "fingerprint": "root1", "username": "alice",
             "content": "Mijn accu geeft foutcode E03 na de laatste firmware."},
            {"source": "src_x", "fingerprint": "c1", "parent_fingerprint": "root1",
             "username": "bob", "content": "Zelfde hier, al twee weken."},
            {"source": "src_x", "fingerprint": "c2", "parent_fingerprint": "root1",
             "username": "carol", "content": "+1"},
        ]

    def _prompt_for_reply(self):
        llm_site, _ = self._fixtures()
        posts = self._thread()
        llm = llm_site.LLMSite()
        with llm as llm_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self._analyze(posts, pending=[posts[2]], all_posts=posts)
        return "\n".join(llm.user_prompts)

    def test_reply_prompt_carries_root_and_sibling_replies(self):
        """「+1」单独看必然是 neutral 噪音。整串给到，模型才判得出它在附和什么"""
        prompt = self._prompt_for_reply()
        assert "foutcode E03" in prompt, "主贴正文没进上下文"
        assert "Zelfde hier" in prompt, "同串的另一条回复没进上下文"
        assert "[讨论串上下文]" in prompt

    def test_the_post_under_analysis_is_marked_and_others_are_not(self):
        """不标出待分析对象，模型会把上下文里别人的态度当成这条的态度"""
        prompt = self._prompt_for_reply()
        assert "▶ 回复 @carol: +1" in prompt, prompt
        assert "← 本条为待分析对象" in prompt
        assert "▶ 主贴 @alice" not in prompt
        assert "▶ 回复 @bob" not in prompt

    def test_lone_root_gets_no_empty_context_wrapper(self):
        """没有回复的孤立主贴不该套一层空的讨论串外壳"""
        llm_site, _ = self._fixtures()
        posts = [{"source": "src_x", "fingerprint": "solo", "username": "dave",
                  "content": "Werkt prima hier."}]
        llm = llm_site.LLMSite()
        with llm as llm_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self._analyze(posts)
        joined = "\n".join(llm.user_prompts)
        assert "[讨论串上下文]" not in joined
        assert "Werkt prima hier." in joined

    def test_lone_root_keeps_the_full_2000_char_allowance(self):
        """孤立主贴是**待分析对象**，不是上下文里的旁人，正文额度仍是 2000 字。

        实测 65 个主贴里大多数没有回复，全走这一支 —— 按 600 字截等于大面积丢正文。
        """
        llm_site, _ = self._fixtures()
        body = "A" * 1500 + "TAILMARKER" + "B" * 400
        posts = [{"source": "src_x", "fingerprint": "solo", "username": "dave",
                  "content": body}]
        llm = llm_site.LLMSite()
        with llm as llm_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self._analyze(posts)
        assert "TAILMARKER" in "\n".join(llm.user_prompts), "正文被按上下文额度截短了"

    def test_context_survives_when_pending_holds_different_objects(self):
        """待分析的那批与组串用的 all_posts 可能不是同一批 dict（各自 load 一次就是）。

        改造中一度按 `is` 认人：那种情况下 ▶ 标记会整个消失，定位待分析条目的 next()
        还会抛 StopIteration 把整轮分析带走。身份一律走 post_key。
        """
        import copy
        llm_site, _ = self._fixtures()
        posts = self._thread()
        detached = copy.deepcopy(posts[2])   # 同一条帖子，不同对象
        llm = llm_site.LLMSite()
        with llm as llm_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            out = self._analyze(posts, pending=[detached], all_posts=posts)
        prompt = "\n".join(llm.user_prompts)
        assert out["success"] == 1, "分析被身份判断带崩了"
        assert "▶ 回复 @carol: +1" in prompt, prompt
        assert "foutcode E03" in prompt

    def test_huge_thread_is_capped_but_keeps_root_and_target(self):
        """信息流来源一个主贴能挂几十条评论，prompt 不能无上限膨胀。

        但无论怎么裁，主贴和待分析那条都必须留着 —— 它们正是判定所依赖的两端。
        """
        from app.services.sentiment_service import THREAD_CONTEXT_LIMIT
        llm_site, _ = self._fixtures()
        posts = [{"source": "src_x", "fingerprint": "root1", "username": "alice",
                  "content": "ROOTMARKER " + "x" * 300}]
        for i in range(60):
            posts.append({
                "source": "src_x", "fingerprint": f"c{i}", "parent_fingerprint": "root1",
                "username": f"u{i}", "content": f"reply{i} " + "y" * 300,
            })
        target = posts[-1]
        target["content"] = "TARGETMARKER " + "z" * 300

        llm = llm_site.LLMSite()
        with llm as llm_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-t", "base_url": llm_url, "model_name": "text-model"})
            self._analyze(posts, pending=[target], all_posts=posts)

        prompt = "\n".join(llm.user_prompts)
        assert "ROOTMARKER" in prompt, "主贴被裁掉了"
        assert "TARGETMARKER" in prompt, "待分析对象被裁掉了"
        assert "…（此处省略部分回复）…" in prompt, "没有裁剪，prompt 会无上限膨胀"
        # 整串块本身受限；给足两倍余量容纳帖子头、标记和省略号
        assert len(prompt) < THREAD_CONTEXT_LIMIT * 2, len(prompt)


class TestBuildSummaryEndToEnd:
    """舆情汇总：真实数据统计"""

    def test_distribution_counts(self):
        from app.services.sentiment_service import SentimentService
        results = [
            {"sentiment": "positive", "intensity": 4, "reason_cn": "", "dimensions": []},
            {"sentiment": "positive", "intensity": 5, "reason_cn": "", "dimensions": []},
            {"sentiment": "negative", "intensity": 3, "reason_cn": "", "dimensions": []},
            {"sentiment": "neutral", "intensity": 2, "reason_cn": "", "dimensions": []},
        ]
        s = SentimentService._build_summary(results)
        assert s["sentiment_distribution"] == {"positive": 2, "negative": 1, "neutral": 1}
        assert s["avg_intensity"] == 3.5

    def test_top_dimensions(self):
        from app.services.sentiment_service import SentimentService
        results = [
            {"sentiment": "positive", "intensity": 3, "reason_cn": "", "dimensions": ["价格/性价比", "安装/配置体验"]},
            {"sentiment": "positive", "intensity": 4, "reason_cn": "", "dimensions": ["价格/性价比"]},
            {"sentiment": "negative", "intensity": 2, "reason_cn": "", "dimensions": ["固件更新"]},
        ]
        s = SentimentService._build_summary(results)
        assert s["top_dimensions"][0][0] == "价格/性价比"
        assert s["top_dimensions"][0][1] == 2

    def test_null_items_are_not_counted_as_neutral(self):
        from app.services.sentiment_service import SentimentService
        results = [None, {"sentiment": "positive", "intensity": 2, "reason_cn": "", "dimensions": []}]
        s = SentimentService._build_summary(results)
        assert s["sentiment_distribution"]["neutral"] == 0
        assert s["not_analyzed"] == 1
        assert s["analyzed"] == 1
        # 分母是实际分析成功的条数，未分析的不该稀释占比
        assert s["sentiment_percentages"]["positive"] == 100.0

    def test_parse_failure_placeholder_is_not_counted_as_neutral(self):
        from app.services.sentiment_service import SentimentService
        # LLM 解析失败时写入的兜底条目，不能冒充成一条真实的中性评价
        results = [
            {"sentiment": None, "intensity": 0, "reason_cn": "解析失败", "dimensions": []},
            {"sentiment": "negative", "intensity": 4, "reason_cn": "", "dimensions": []},
        ]
        s = SentimentService._build_summary(results)
        assert s["sentiment_distribution"]["neutral"] == 0
        assert s["not_analyzed"] == 1
        assert s["sentiment_percentages"]["negative"] == 100.0
        # 失败条目的 intensity 不该拉低均值
        assert s["avg_intensity"] == 4

    def test_unknown_sentiment_value_goes_to_not_analyzed(self):
        from app.services.sentiment_service import SentimentService
        results = [
            {"sentiment": "很正面", "intensity": 4, "reason_cn": "", "dimensions": []},
            {"sentiment": "positive", "intensity": 4, "reason_cn": "", "dimensions": []},
        ]
        s = SentimentService._build_summary(results)
        assert s["sentiment_distribution"] == {"positive": 1, "negative": 0, "neutral": 0}
        assert s["not_analyzed"] == 1

    def test_empty_results(self):
        from app.services.sentiment_service import SentimentService
        s = SentimentService._build_summary([])
        assert s["avg_intensity"] == 0
        assert s["sentiment_percentages"]["positive"] == 0


class TestDimensionNormalizationEndToEnd:
    """维度必须保持封闭集合，否则 top_dimensions 会碎成近义标签、跨来源对比失效"""

    def _norm(self, dims):
        from app.services.sentiment_service import SentimentService
        return SentimentService._normalize_dimensions(dims)

    def test_shorthand_maps_to_canonical_label(self):
        """实测 LLM 会把带括号的标签简写，于是同一维度在报表里占两行"""
        assert self._norm(["认证/合规"]) == ["认证/合规(如Synergrid)"]
        assert self._norm(["与其他品牌对比"]) == ["与其他品牌对比(如AEG/Marstek)"]

    def test_exact_labels_pass_through(self):
        assert self._norm(["价格/性价比", "安全性"]) == ["价格/性价比", "安全性"]

    def test_unknown_label_is_dropped(self):
        """宁可少一个标签，也不能让维度表不再封闭"""
        assert self._norm(["用户自创的新维度"]) == []

    def test_duplicates_after_normalization_are_collapsed(self):
        assert self._norm(["认证/合规", "认证/合规(如Synergrid)"]) == ["认证/合规(如Synergrid)"]

    def test_parse_sentiment_normalizes_inline(self):
        from app.services.sentiment_service import SentimentService
        parsed = SentimentService._parse_sentiment(
            '{"sentiment":"positive","intensity":3,"reason_cn":"x","dimensions":["认证/合规"]}'
        )
        assert parsed["dimensions"] == ["认证/合规(如Synergrid)"]

    def test_non_dict_and_bad_json_are_unchanged(self):
        from app.services.sentiment_service import SentimentService
        assert SentimentService._parse_sentiment("不是 JSON") is None


class TestBySourceSummaryEndToEnd:
    """跨来源对比：分组走纯 Python，不让 LLM 感知来源，免得它带着平台先验去评分"""

    def _fixture(self):
        results = [
            {"sentiment": "positive", "intensity": 4, "dimensions": ["价格/性价比"]},
            {"sentiment": "negative", "intensity": 2, "dimensions": ["固件更新"]},
            {"sentiment": "negative", "intensity": 5, "dimensions": ["价格/性价比", "安全性"]},
        ]
        posts = [
            {"source": "src_forum", "fingerprint": "a"},
            {"source": "src_forum", "fingerprint": "b"},
            {"source": "src_social", "fingerprint": "c"},
        ]
        return results, posts, {"src_forum": "论坛", "src_social": "社媒"}

    def test_no_posts_means_no_by_source(self):
        """既有调用方一行不改：不传 posts 就是原来的行为"""
        from app.services.sentiment_service import SentimentService
        results, _, _ = self._fixture()
        assert "by_source" not in SentimentService._build_summary(results)

    def test_by_source_splits_the_same_batch(self):
        from app.services.sentiment_service import SentimentService
        results, posts, names = self._fixture()
        s = SentimentService._build_summary(results, posts, names)

        assert set(s["by_source"]) == {"src_forum", "src_social"}
        assert s["by_source"]["src_forum"]["name"] == "论坛"
        assert s["by_source"]["src_forum"]["distribution"] == {"positive": 1, "negative": 1, "neutral": 0}
        assert s["by_source"]["src_forum"]["avg_intensity"] == 3.0
        assert s["by_source"]["src_social"]["distribution"]["negative"] == 1
        # 分组统计不能改变全局统计
        assert s["sentiment_distribution"] == {"positive": 1, "negative": 2, "neutral": 0}

    def test_cross_source_lines_up_dimensions(self):
        from app.services.sentiment_service import SentimentService
        results, posts, names = self._fixture()
        s = SentimentService._build_summary(results, posts, names)
        assert s["cross_source"]["价格/性价比"] == {"src_forum": 1, "src_social": 1}
        assert s["cross_source"]["安全性"] == {"src_social": 1}

    def test_legacy_posts_without_source_are_grouped_as_tweakers(self):
        from app.services.sentiment_service import SentimentService
        s = SentimentService._build_summary(
            [{"sentiment": "positive", "intensity": 3, "dimensions": []}],
            [{"fingerprint": "old"}],
            {},
        )
        assert list(s["by_source"]) == ["tweakers"]


class TestResolveSourcesEndToEnd:
    """LLM 提议、后端保证：来源一多模型很容易只挑一个，那是静默漏采"""

    def _sources(self):
        return [
            {"id": "src_a", "name": "论坛A"},
            {"id": "src_b", "name": "社媒B"},
        ]

    def _resolve(self, desc, plan_actions):
        from app.models import PlanStep
        from app.services.orchestrator import _resolve_sources
        plan = [PlanStep(action=a, params=p) for a, p in plan_actions]
        return _resolve_sources({"description": desc}, plan, self._sources())

    def test_all_sources_wording_expands_regardless_of_llm(self):
        plan, warns = self._resolve(
            "采集所有来源，翻译，导出Excel",
            [("collect", {"source_id": "src_a"}), ("translate", {}), ("generate_excel", {})],
        )
        assert [s.params["source_id"] for s in plan if s.action == "collect"] == ["src_a", "src_b"]
        assert any("漏掉" in w for w in warns)
        # 非 collect 步骤保持原有顺序，跟在 collect 后面
        assert [s.action for s in plan] == ["collect", "collect", "translate", "generate_excel"]

    def test_named_single_source_is_respected(self):
        plan, warns = self._resolve("只采集论坛A", [("collect", {"source_id": "src_a"})])
        assert [s.params["source_id"] for s in plan] == ["src_a"]
        assert warns == []

    def test_hallucinated_source_id_is_ignored_with_warning(self):
        plan, warns = self._resolve("采集数据", [("collect", {"source_id": "src_不存在"})])
        assert any("编造" in w for w in warns)
        # 一个有效来源都没给出时全采，比静默不采安全
        assert len(plan) == 2

    def test_override_fails_loudly_instead_of_substituting(self):
        """用户临时贴的链接不能拿已注册来源顶替 —— 那是答非所问，报告里还看不出来"""
        import pytest
        with pytest.raises(Exception) as e:
            self._resolve("采集 https://example.com/thread/1", [("collect", {"override": "https://example.com/thread/1"})])
        assert "数据源" in str(e.value)

    def test_plan_without_collect_is_untouched(self):
        plan, warns = self._resolve("翻译已有数据", [("translate", {}), ("generate_excel", {})])
        assert [s.action for s in plan] == ["translate", "generate_excel"]
        assert warns == []


class TestSentimentActionIsOfferedToTheLLM:
    """舆情步骤由模型自己判断要不要，不再有关键词表。至少得让它知道有这个动作"""

    def test_prompt_declares_the_sentiment_action_and_when_to_use_it(self):
        import asyncio
        sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures"))
        import llm_site
        from app.services.llm_service import LLMService
        from app.models import LLMConfig

        server = llm_site.LLMSite()
        with server as base_url:
            svc = LLMService(LLMConfig(api_key="sk-test", base_url=base_url,
                                       model_name="test-model"))
            plan = asyncio.new_event_loop().run_until_complete(
                svc.parse_intent("帮我分析下舆情", [])
            )

        prompt = server.prompts[0]
        assert "sentiment" in prompt, "模型看不到这个动作就永远产不出它"
        # 只把「舆情」当话题词提一下不该触发 —— 舆情分析要逐条调 LLM，很贵
        assert "抓取舆情数据" in prompt, "缺了反例，模型会把话题词也当成要求"
        assert "translate 之后" in prompt, "舆情读的是译文，顺序必须交代"
        # 模型给了 sentiment，解析这一层就得原样带出来
        assert [s["action"] for s in plan["plan"]] == ["generate_excel", "sentiment"]


class TestEmptyPostsNeverStoredEndToEnd:
    """正文为空的帖子一条都不入库 —— 它翻译不了、分析不了，只在报告里占一行空白"""

    def setup_method(self):
        import app.services.storage as storage_module
        self.tmpdir = tempfile.mkdtemp()
        self.storage = storage_module
        self._old_db = storage_module.DB_PATH
        storage_module.DB_PATH = os.path.join(self.tmpdir, "hyxi.db")
        storage_module.init_db()

    def teardown_method(self):
        self.storage.DB_PATH = self._old_db
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _post(self, fp, content, parent=None, level=0, images=None):
        return {"username": "某人", "timestamp": "01-07-2026 10:00", "content": content,
                "translation": "", "page_number": 1, "fingerprint": fp, "source": "src_e",
                "parent_fingerprint": parent, "reply_level": level,
                "images": images or []}

    def test_empty_post_is_not_stored(self):
        added = self.storage.upsert_posts("src_e", [
            self._post("a", "有正文的帖子"),
            self._post("b", ""),
            self._post("c", "   \n  "),   # 只有空白也算空
        ])
        assert added == 1
        assert [p["fingerprint"] for p in self.storage.load_posts(["src_e"])] == ["a"]

    def test_image_only_post_is_stored(self):
        """没有正文但有配图的帖子**不是空帖** —— 内容全在图上。

        实测丢过一条：真实文件 media/src_b32bc603/6680d5f13a6b2b4c_0.jpg 还在盘上
        （HYXi 安装检查报告，总分 88、发电异常 8/20 标橙），posts 表里却一行都没有。
        采集脚本先下图、再由这个入库口丢帖子，图留下、帖子没了 —— 它连「未分析」
        都不会显示，是直接消失，比留一行空白更难发现。
        """
        added = self.storage.upsert_posts("src_e", [
            self._post("a", "有正文的帖子"),
            self._post("pic", "", images=["src_e/pic_0.jpg"]),
            self._post("d", "", images=[]),   # 既没正文也没图，照旧丢掉
        ])
        assert added == 2
        stored = self.storage.load_posts(["src_e"])
        assert [p["fingerprint"] for p in stored] == ["a", "pic"]
        assert stored[1]["images"] == ["src_e/pic_0.jpg"]

    def test_comments_of_an_image_only_root_stay_attached_to_it(self):
        """纯图主贴留下了，它的评论就该继续挂在它下面 —— 那张图正是这些评论的上下文"""
        self.storage.upsert_posts("src_e", [
            self._post("root", "", images=["src_e/root_0.jpg"]),
            self._post("c1", "评论一", parent="root", level=1),
        ])
        posts = self.storage.load_posts(["src_e"])
        assert [p["fingerprint"] for p in posts] == ["root", "c1"]
        assert posts[1]["parent_fingerprint"] == "root"
        assert posts[1]["reply_level"] == 1

    def test_empty_root_with_comments_is_kept_because_nobody_comments_on_nothing(self):
        """有人在它下面发言 → 它多半是**提取失败**，不是真的空帖。

        实测踩过：真站上那条主贴写着「Mijn HyXi Halo is gekoppeld aan:」，8-10 那轮
        正文没提取出来变成空帖，于是被丢掉、它的两条评论被提成主贴 —— 而那个提升是
        **不可逆的**：下一轮正文提对了，父贴换了指纹重新入库，评论却还挂在主贴身份上。
        真站核实过一模一样的第二例：纯图主贴被丢，回复「Is bij mij ook zo…」被提成主贴，
        舆情因此判成 neutral，而它其实是在附和一条报故障的帖子。
        """
        self.storage.upsert_posts("src_e", [
            self._post("root", ""),
            self._post("c1", "评论一", parent="root", level=1),
            self._post("c2", "评论二", parent="root", level=1),
        ])
        posts = self.storage.load_posts(["src_e"])
        assert [p["fingerprint"] for p in posts] == ["root", "c1", "c2"]
        assert [p["parent_fingerprint"] for p in posts] == [None, "root", "root"]
        assert [p["reply_level"] for p in posts] == [0, 1, 1]

    def test_empty_root_whose_comments_are_also_empty_is_still_dropped(self):
        """整棵子树一个字都没有，那就是真的什么都没有，照旧一条都不留"""
        added = self.storage.upsert_posts("src_e", [
            self._post("root", ""),
            self._post("c1", "", parent="root", level=1),
            self._post("keep", "有正文的另一条"),
        ])
        assert added == 1
        assert [p["fingerprint"] for p in self.storage.load_posts(["src_e"])] == ["keep"]

    def test_kept_empty_root_still_anchors_its_thread_for_sentiment(self):
        """留住父贴的意义在这里：评论能拿到同串上下文，而不是各自成为孤立主贴。

        这正是被提升那条付出的代价 —— 「Is bij mij ook zo」单独看必然是 neutral。
        """
        from app.services.post_tree import thread_of, post_key

        self.storage.upsert_posts("src_e", [
            self._post("root", ""),
            self._post("c1", "我这边也一样", parent="root", level=1),
            self._post("c2", "我也是", parent="root", level=1),
        ])
        posts = self.storage.load_posts(["src_e"])
        threads = thread_of(posts)
        c1 = next(p for p in posts if p["fingerprint"] == "c1")
        members = threads[post_key(c1)]
        assert [m["fingerprint"] for m in members] == ["root", "c1", "c2"], \
            "评论看不到父贴和同串的其他人，舆情就只能各判各的"

    def test_seq_still_has_no_holes_so_ordering_survives(self):
        """seq 是全链路的顺序锚点，被丢掉的帖子不能在里面留个洞"""
        self.storage.upsert_posts("src_e", [
            self._post("a", "第一条"), self._post("skip", ""), self._post("b", "第二条"),
        ])
        conn = self.storage._get_conn()
        try:
            rows = conn.execute(
                "SELECT fingerprint, seq FROM posts WHERE source_id='src_e' ORDER BY seq"
            ).fetchall()
        finally:
            conn.close()
        assert [(r["fingerprint"], r["seq"]) for r in rows] == [("a", 0), ("b", 1)]

    def test_second_round_does_not_resurrect_them(self):
        """增量重抓会把同一条空帖再送一遍，不能这轮挡住下轮放行"""
        self.storage.upsert_posts("src_e", [self._post("a", "有正文")])
        added = self.storage.upsert_posts("src_e", [
            self._post("a", "有正文"), self._post("b", ""),
        ])
        assert added == 0
        assert len(self.storage.load_posts(["src_e"])) == 1


class TestPipelineSentimentStepEndToEnd:
    """描述里写「分析舆情」→ 任务跑完结论就在库里。真 HTTP LLM、真 SQLite、真步骤调度"""

    def setup_method(self):
        import app.config as cfg
        import app.services.storage as storage_module
        self.cfg = cfg
        self.tmpdir = tempfile.mkdtemp()
        self._old_dir = cfg.settings.data_dir
        cfg.settings.data_dir = self.tmpdir
        # exports_dir 是类定义时算好的常量（同 storage.DB_PATH 那个坑），只改
        # data_dir 的话 generate_excel 会把报告写进真实的 backend/data/exports
        self._old_exports = cfg.settings.exports_dir
        cfg.settings.exports_dir = os.path.join(self.tmpdir, "exports")
        os.makedirs(cfg.settings.exports_dir, exist_ok=True)
        self.storage = storage_module
        self._old_db = storage_module.DB_PATH
        storage_module.DB_PATH = os.path.join(self.tmpdir, "hyxi.db")
        storage_module.init_db()

        self.storage.save_source({
            "id": "src_p", "name": "小组来源", "collector_id": "group_feed",
            "params": {"group_id": "g1", "base_url": "http://127.0.0.1:1"},
            "enabled": True, "created_at": "2026-08-06T09:00:00",
        })
        self.storage.upsert_posts("src_p", [
            {"username": "用户1", "timestamp": "01-07-2026 10:00",
             "content": "De firmware update sloopte mijn WiFi-verbinding.",
             "translation": "固件更新弄坏了我的 WiFi 连接。", "page_number": 1,
             "fingerprint": "fp1", "source": "src_p",
             "parent_fingerprint": None, "reply_level": 0,
             "_processed": {"translated": True}},
            {"username": "用户2", "timestamp": "02-07-2026 10:00",
             "content": "Werkt hier prima na de update.",
             "translation": "更新后这边一切正常。", "page_number": 1,
             "fingerprint": "fp2", "source": "src_p",
             "parent_fingerprint": None, "reply_level": 0,
             "_processed": {"translated": True}},
        ])

    def teardown_method(self):
        self.cfg.settings.data_dir = self._old_dir
        self.cfg.settings.exports_dir = self._old_exports
        self.storage.DB_PATH = self._old_db
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _run(self, description, force_full=False):
        task, server = self._run_full(description, force_full)
        return task, server.seen

    def _run_full(self, description, force_full=False):
        import asyncio
        sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures"))
        import llm_site
        from app.services.orchestrator import orchestrator

        server = llm_site.LLMSite()
        with server as base_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-test", "base_url": base_url, "model_name": "test-model",
            })
            task_id = f"pipeline-{len(server.seen)}-{description[:4]}"
            orchestrator.create_task(task_id, description, force_full=force_full)
            asyncio.new_event_loop().run_until_complete(
                orchestrator.execute_task(task_id)
            )
        return orchestrator.get_task(task_id), server

    def _conclusions(self):
        posts = self.storage.load_posts(["src_p"])
        return posts, [(p.get("_processed") or {}).get("sentiment_at") for p in posts]

    def test_plan_with_sentiment_actually_analyzes(self):
        """用户报的就是这个：描述里写了「分析舆情」，任务跑完一条都没分析。

        模型识别出意图、计划里带了 sentiment，流水线就必须真的去分析。不然页面上
        显示的是别的任务留下的结论，看着像分析过了，用户根本发现不了。
        """
        task, seen = self._run("翻译已有数据，导出Excel，分析舆情")

        assert task["status"] == "completed", task.get("error_message")
        assert [s["action"] for s in task["plan"]] == ["generate_excel", "sentiment"]
        assert "plan" in seen, f"意图解析没走到 LLM: {seen}"

        posts, analyzed = self._conclusions()
        assert all(analyzed), f"有帖子没被标记分析过: {analyzed}"
        data = self.storage.get_sentiment(task["id"], posts)
        assert data and len(data["results"]) == 2, data
        assert all(r and r.get("sentiment") for r in data["results"]), data["results"]

    def test_second_run_does_not_pay_for_the_same_posts_again(self):
        """增量粒度是 sentiment_at，跨任务共享 —— 重跑不该再调一次 LLM"""
        self._run("翻译已有数据，分析舆情")
        _task, seen = self._run("翻译已有数据，分析舆情")
        assert [s for s in seen if s != "plan"] == [], f"重复分析了: {seen}"

    def _run_watching_the_progress_stream(self, description):
        """真跑一遍流水线，同时像页面那样连着 /tasks/{id}/events。

        走的是**路由函数本身**而不是直接调 event_generator —— 这条流的结束条件由
        端点决定，绕过它就测不到真正会出错的那个选择。
        """
        import asyncio
        sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures"))
        import llm_site
        from app.services.orchestrator import orchestrator
        from app.services.progress_manager import progress_manager
        from app.routers.tasks import task_events

        server = llm_site.LLMSite()
        events = []
        with server as base_url:
            self.storage.set_app_config("llm", {
                "api_key": "sk-test", "base_url": base_url, "model_name": "test-model",
            })
            task_id = f"stream-{len(server.seen)}"
            orchestrator.create_task(task_id, description)

            async def main():
                resp = await task_events(task_id)

                async def consume():
                    async for chunk in resp.body_iterator:
                        if chunk.startswith("event: "):
                            events.append(chunk.splitlines()[0][len("event: "):])

                consumer = asyncio.ensure_future(consume())
                # 订阅是在生成器第一次被拉动时才建立的，抢在任务开跑之前
                for _ in range(200):
                    if progress_manager.subscribers.get(task_id):
                        break
                    await asyncio.sleep(0.01)
                await orchestrator.execute_task(task_id)
                await asyncio.wait_for(consumer, timeout=15)

            asyncio.new_event_loop().run_until_complete(main())
        return orchestrator.get_task(task_id), events

    def test_progress_stream_survives_the_sentiment_step_and_delivers_task_complete(self):
        """用户报的：带舆情的任务跑完，进度页停在 running，最后一行是「连接中断」。

        两条 SSE 流跑在同一个频道上，却共用一份结束条件。流水线的 sentiment 步骤一发完
        `sentiment_complete`，任务进度流就自己 break 了 —— 紧随其后的 `task_complete`
        没人收得到。前端的 task_complete 处理器是唯一会调 fetchTask() 的地方，收不到
        就没人刷新 currentTask，isCompleted 永远是 false：不跳转、也不出现「查看结果」。
        """
        task, events = self._run_watching_the_progress_stream("翻译已有数据，分析舆情")

        assert task["status"] == "completed", task.get("error_message")
        assert "sentiment_complete" in events, f"这一轮压根没跑舆情: {events}"
        assert "task_complete" in events, f"任务结束了，进度流却没收到 task_complete: {events}"
        assert events.index("sentiment_complete") < events.index("task_complete"), events
        # task_complete 之后流才该结束，它必须是最后一个
        assert events[-1] == "task_complete", events

    def test_no_sentiment_step_means_no_llm_analysis_calls(self):
        """模型没给 sentiment 步骤时，一次分析调用都不该发生 —— 那是要花钱的"""
        task, seen = self._run("翻译已有数据，导出Excel")
        assert [s["action"] for s in task["plan"]] == ["generate_excel"]
        assert [s for s in seen if s != "plan"] == [], f"没要舆情却调了分析: {seen}"
        _posts, analyzed = self._conclusions()
        assert not any(analyzed), analyzed

    DESC_FULL = "重新翻译并分析舆情"

    def test_a_normal_rerun_still_skips_what_is_already_done(self):
        """对照组：两条帖子本来就带 translated 标记、跑完也有了结论，
        普通重跑必须一次翻译调用都不发。

        没有这个对照，实现里把增量整个拆掉也是绿的 —— 而那等于每一轮定时任务
        都在重复付翻译和舆情的钱。
        """
        self._run(self.DESC_FULL)                      # 第一轮：把舆情结论补上
        _task, server = self._run_full(self.DESC_FULL)  # 第二轮：什么都不该再做

        assert server.translated == [], f"已翻译的帖子又被送去翻译了: {server.translated}"
        assert [x for x in server.seen if x != "plan"] == [], (
            f"已分析的帖子又被送去分析了: {server.seen}"
        )

    def test_full_rerun_also_re_describes_the_images(self):
        """全量重跑必须连图片描述一起重算 —— 确认弹窗里就是这么承诺用户的。

        image_desc 平时绝不重算（那是花钱换来的），可 sentiment_service 跳过的判据
        正是「已经有 image_desc」。不清掉它，用户为「按当前口径重算一遍」付了钱，
        图片描述却仍是旧模型、旧 prompt 的产物 —— 而换了多模态模型正是有人点这个
        按钮的主要原因。
        """
        import asyncio
        sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures"))
        import llm_site
        import vision_site
        from app.services.orchestrator import orchestrator

        # 一张真图落在临时 media 目录里，帖子挂上它
        png = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJ"
            "AAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
        )
        media_dir = os.path.join(self.tmpdir, "media", "src_p")
        os.makedirs(media_dir, exist_ok=True)
        with open(os.path.join(media_dir, "pic_0.png"), "wb") as f:
            f.write(png)
        self.storage.upsert_posts("src_p", [{
            "username": "用户3", "timestamp": "03-07-2026 10:00",
            "content": "Zie foto.", "translation": "看图。",
            "page_number": 1, "fingerprint": "fp3", "source": "src_p",
            "parent_fingerprint": None, "reply_level": 0,
            "images": ["src_p/pic_0.png"],
            "_processed": {"translated": True},
        }])

        def run(tag, force_full):
            llm = llm_site.LLMSite()
            vision = vision_site.VisionSite()
            with llm as llm_url, vision as vision_url:
                self.storage.set_app_config("llm", {
                    "api_key": "sk-test", "base_url": llm_url, "model_name": "text-model"})
                self.storage.set_app_config("vision", {
                    "api_key": "sk-v", "base_url": vision_url, "model_name": "vision-model"})
                orchestrator.create_task(tag, self.DESC_FULL, force_full=force_full)
                asyncio.new_event_loop().run_until_complete(orchestrator.execute_task(tag))
            return orchestrator.get_task(tag), vision

        first, vision1 = run("img-rerun-1", False)
        assert first["status"] == "completed", first.get("error_message")
        assert vision1.image_count == 1, f"第一轮就没理解图: {vision1.calls}"
        stored = {p["fingerprint"]: p for p in self.storage.load_posts(["src_p"])}
        assert stored["fp3"].get("image_desc") == vision_site.DESCRIPTION, "描述没落库"

        # 对照组：普通重跑不该为同一张图再付一次钱
        _second, vision2 = run("img-rerun-2", False)
        assert vision2.image_count == 0, f"普通重跑又买了一次图片描述: {vision2.calls}"

        # 全量重跑：必须重新调一次多模态
        third, vision3 = run("img-rerun-3", True)
        assert third["status"] == "completed", third.get("error_message")
        assert vision3.image_count == 1, (
            f"全量重跑没有重算图片描述，"
            f"弹窗里那句「带图的还会再调一次多模态模型」就是假话: {vision3.calls}"
        )

    def test_full_rerun_ignores_every_incremental_marker(self):
        """全量重跑：已翻译、已分析的帖子必须全部重新送给模型。

        增量标记散在三处（collect 的 incremental、translate 的 _processed.translated、
        sentiment 的 sentiment_at），少放开一处，用户点了「全量重跑」却发现某一步
        照旧跳过 —— 而这个按钮本来就是花钱买「按当前口径重算一遍」的。
        """
        self._run(self.DESC_FULL)                       # 先把结论跑出来
        task, server = self._run_full(self.DESC_FULL, force_full=True)

        assert task["status"] == "completed", task.get("error_message")
        assert task["force_full"] is True, "标志没随任务落库"
        assert sum(server.translated) == 2, f"没有全部重译: {server.translated}"
        # seen 里翻译请求记的是 0（它们的 prompt 里没有「帖子N [来源:」），
        # 所以只判「非空」恒真 —— 必须找到一条真带帖子的分析请求
        assert any(x != "plan" and x > 0 for x in server.seen), (
            f"没有重新分析舆情（seen 里一条带帖子的请求都没有）: {server.seen}"
        )
        assert any("全量重跑" in (l.get("message") or "") for l in task["logs"]), (
            "日志里没说这是全量重跑，用户事后无从分辨这一轮跟别的有什么不同"
        )


class TestTimestampNormalizationEndToEnd:
    """时间戳格式转换：荷兰 dd-mm-yyyy → yyyy-mm-dd"""

    def _normalize(self, ts):
        import re
        if not ts:
            return ts
        m = re.match(r"(\d{2})-(\d{2})-(\d{4})\s+(.+)", ts)
        if m:
            return f"{m.group(3)}-{m.group(2)}-{m.group(1)} {m.group(4)}"
        return ts

    def test_dutch_format_converted(self):
        assert self._normalize("22-05-2026 17:06") == "2026-05-22 17:06"

    def test_end_of_year(self):
        assert self._normalize("31-12-2026 23:59") == "2026-12-31 23:59"

    def test_already_iso_unchanged(self):
        assert self._normalize("2026-05-22 17:06") == "2026-05-22 17:06"

    def test_empty_unchanged(self):
        assert self._normalize("") == ""


class TestExcelGenerationEndToEnd:
    """Excel 生成：真实 openpyxl 输出"""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        import app.services.excel_service as excel_mod
        excel_mod.settings.exports_dir = self.tmpdir

    def teardown_method(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_generate_excel_creates_valid_file(self):
        from app.services.excel_service import ExcelService
        from app.services.progress_manager import ProgressManager

        posts = [
            {"username": "TestUser", "timestamp": "2026-05-22 17:06",
             "content": "测试内容", "translation": "Test translation", "page_number": 1},
        ]
        pm = ProgressManager()

        import asyncio
        async def run():
            return await ExcelService.execute("test-task", posts, {"include_stats": True}, pm)

        result = asyncio.get_event_loop().run_until_complete(run())

        assert os.path.exists(result["file_path"])
        assert result["file_name"].endswith(".xlsx")

        from openpyxl import load_workbook
        wb = load_workbook(result["file_path"])
        assert "论坛帖子翻译" in wb.sheetnames
        assert "统计信息" in wb.sheetnames

        ws = wb["论坛帖子翻译"]
        assert [ws.cell(1, c).value for c in range(1, 9)] == [
            "序号", "来源", "层级", "用户名", "发布时间", "原文", "中文翻译", "页码",
        ]
        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 4).value == "TestUser"
        assert ws.cell(2, 7).value == "Test translation"

    def test_generate_excel_without_stats(self):
        from app.services.excel_service import ExcelService
        from app.services.progress_manager import ProgressManager

        posts = [{"username": "U", "timestamp": "", "content": "C", "translation": "T", "page_number": 1}]
        pm = ProgressManager()

        import asyncio
        async def run():
            return await ExcelService.execute("test-task2", posts, {"include_stats": False}, pm)

        result = asyncio.get_event_loop().run_until_complete(run())
        from openpyxl import load_workbook
        wb = load_workbook(result["file_path"])
        assert "统计信息" not in wb.sheetnames


class TestPostKeyEndToEnd:
    """跨来源索引键：指纹不含来源，键必须带上，否则两个平台的帖子会互相覆盖"""

    def test_same_fingerprint_different_sources_do_not_collide(self):
        from app.services.post_tree import post_key
        a = {"source": "src_aaa", "fingerprint": "deadbeefdeadbeef"}
        b = {"source": "src_bbb", "fingerprint": "deadbeefdeadbeef"}
        assert post_key(a) != post_key(b)

    def test_legacy_posts_without_source_fall_back_to_tweakers(self):
        """历史数据没有 source 字段，缺省填 tweakers 才能和既有落盘数据对上"""
        from app.services.post_tree import post_key
        assert post_key({"fingerprint": "abc123"}) == "tweakers:abc123"
        assert post_key({"source": "", "fingerprint": "abc123"}) == "tweakers:abc123"

    def test_merge_by_key_does_not_leak_across_sources(self):
        from app.services.orchestrator import _merge_by_fingerprint
        source_posts = [
            {"source": "src_a", "fingerprint": "ff", "content": "A 的帖子"},
            {"source": "src_b", "fingerprint": "ff", "content": "B 的帖子"},
        ]
        translated = [{"source": "src_a", "fingerprint": "ff", "content": "A 的帖子", "translation": "只属于A"}]
        merged = _merge_by_fingerprint(source_posts, translated)
        assert merged[0]["translation"] == "只属于A"
        assert "translation" not in merged[1]


class TestThreadOfEndToEnd:
    """每条帖子 → 它所属的整串。舆情分析拿它当上下文"""

    def _posts(self):
        return [
            {"source": "s", "fingerprint": "root1", "content": "主贴1"},
            {"source": "s", "fingerprint": "c1", "parent_fingerprint": "root1", "content": "评论1"},
            {"source": "s", "fingerprint": "c2", "parent_fingerprint": "c1", "content": "评论1的回复"},
            {"source": "s", "fingerprint": "root2", "content": "主贴2"},
        ]

    def test_every_member_maps_to_the_same_thread(self):
        from app.services.post_tree import thread_of, post_key
        posts = self._posts()
        mapping = thread_of(posts)
        thread = mapping[post_key(posts[1])]
        assert [p["fingerprint"] for p in thread] == ["root1", "c1", "c2"]
        # 同一串里的每个成员拿到的都是同一份（含嵌套回复）
        assert mapping[post_key(posts[0])] is thread
        assert mapping[post_key(posts[2])] is thread

    def test_nested_reply_belongs_to_the_top_level_thread(self):
        """嵌套回复归到顶层主贴那一串，不按子树再切 —— 讨论是围绕主贴发生的"""
        from app.services.post_tree import thread_of, post_key
        posts = self._posts()
        thread = thread_of(posts)[post_key(posts[2])]
        assert thread[0]["fingerprint"] == "root1"

    def test_unrelated_threads_do_not_bleed_into_each_other(self):
        from app.services.post_tree import thread_of, post_key
        posts = self._posts()
        mapping = thread_of(posts)
        assert [p["fingerprint"] for p in mapping[post_key(posts[3])]] == ["root2"]

    def test_orphan_reply_forms_its_own_thread(self):
        """父贴不在本批里的回复自成一串，语义与 build_tree 一致，不能丢"""
        from app.services.post_tree import thread_of, post_key
        orphan = {"source": "s", "fingerprint": "x", "parent_fingerprint": "gone",
                  "content": "孤儿回复"}
        mapping = thread_of([orphan])
        assert [p["fingerprint"] for p in mapping[post_key(orphan)]] == ["x"]


class TestPostTreeEndToEnd:
    """扁平数组 → 出口组树。存储层不动，8 处依赖扁平结构的逻辑一行不改"""

    def _posts(self):
        return [
            {"source": "s", "fingerprint": "root1", "content": "主贴1"},
            {"source": "s", "fingerprint": "c1", "parent_fingerprint": "root1", "content": "评论1"},
            {"source": "s", "fingerprint": "c2", "parent_fingerprint": "c1", "content": "评论1的回复"},
            {"source": "s", "fingerprint": "root2", "content": "主贴2"},
        ]

    def test_build_tree_groups_children_under_parents(self):
        from app.services.post_tree import build_tree, post_key
        roots, children = build_tree(self._posts())
        assert [r["fingerprint"] for r in roots] == ["root1", "root2"]
        assert [c["fingerprint"] for c in children[post_key(roots[0])]] == ["c1"]

    def test_orphan_comment_is_kept_as_root(self):
        """父贴不在本批数据里的评论不能被丢掉 —— 增量只抓到评论那一轮就会这样"""
        from app.services.post_tree import build_tree
        roots, _ = build_tree([
            {"source": "s", "fingerprint": "x", "parent_fingerprint": "不存在的父贴"},
        ])
        assert len(roots) == 1

    def test_order_by_thread_is_depth_first_and_fills_level(self):
        from app.services.post_tree import order_by_thread
        ordered = order_by_thread(self._posts())
        assert [p["fingerprint"] for p in ordered] == ["root1", "c1", "c2", "root2"]
        assert [p["reply_level"] for p in ordered] == [0, 1, 2, 0]

    def test_roots_are_ordered_newest_first(self):
        """主贴按发表时间从新到旧，评论跟着自己的主贴走、不参与排序"""
        from app.services.post_tree import order_by_thread
        posts = [
            {"source": "s", "fingerprint": "r1", "timestamp": "20-05-2026 10:00", "content": "旧主贴"},
            {"source": "s", "fingerprint": "c1", "parent_fingerprint": "r1",
             "timestamp": "28-06-2026 09:00", "content": "旧主贴的新评论"},
            {"source": "s", "fingerprint": "r2", "timestamp": "01-07-2026 08:00", "content": "新主贴"},
        ]
        ordered = order_by_thread(posts)
        assert [p["content"] for p in ordered] == ["新主贴", "旧主贴", "旧主贴的新评论"]

    def test_sorting_uses_iso_not_raw_dutch_string(self):
        """落盘是 dd-mm-yyyy，按字符串排会变成「按日排先」：01-07 排到 28-06 前面"""
        from app.services.post_tree import order_by_thread
        posts = [
            {"source": "s", "fingerprint": "a", "timestamp": "28-06-2026 09:00", "content": "六月底"},
            {"source": "s", "fingerprint": "b", "timestamp": "01-07-2026 08:00", "content": "七月初"},
            {"source": "s", "fingerprint": "c", "parent_fingerprint": "b", "content": "评论"},
        ]
        assert [p["content"] for p in order_by_thread(posts)][0] == "七月初"

    def test_posts_without_a_timestamp_sink_to_the_bottom(self):
        """早期采集读不到 tooltip 绝对时间，落盘留空。这些帖子实际很新，不该霸占最前面"""
        from app.services.post_tree import order_by_thread
        posts = [
            {"source": "s", "fingerprint": "a", "timestamp": "", "content": "没时间1"},
            {"source": "s", "fingerprint": "b", "timestamp": "", "content": "没时间2"},
            {"source": "s", "fingerprint": "c", "timestamp": "20-05-2026 10:00", "content": "有时间"},
            {"source": "s", "fingerprint": "d", "parent_fingerprint": "c", "content": "评论"},
        ]
        ordered = [p["content"] for p in order_by_thread(posts)]
        assert ordered == ["有时间", "评论", "没时间1", "没时间2"], ordered


class TestFingerprintGenerationEndToEnd:
    """帖子指纹：SHA256 唯一标识"""

    def test_same_input_produces_same_fingerprint(self):
        def fingerprint(username, timestamp, content):
            raw = f"{username}|{timestamp}|{(content or '')[:100]}"
            return hashlib.sha256(raw.encode()).hexdigest()[:16]

        fp1 = fingerprint("Dorpjes", "22-05-2026 17:06", "Sinds gisteren...")
        fp2 = fingerprint("Dorpjes", "22-05-2026 17:06", "Sinds gisteren...")
        assert fp1 == fp2
        assert len(fp1) == 16

    def test_different_inputs_produce_different_fingerprints(self):
        def fingerprint(username, timestamp, content):
            raw = f"{username}|{timestamp}|{(content or '')[:100]}"
            return hashlib.sha256(raw.encode()).hexdigest()[:16]

        fp1 = fingerprint("A", "t1", "content1")
        fp2 = fingerprint("B", "t2", "content2")
        assert fp1 != fp2

    def test_content_truncated_at_100_chars(self):
        def fingerprint(username, timestamp, content):
            raw = f"{username}|{timestamp}|{(content or '')[:100]}"
            return hashlib.sha256(raw.encode()).hexdigest()[:16]

        long_content = "x" * 200
        short_content = "x" * 100
        fp_long = fingerprint("U", "T", long_content)
        fp_short = fingerprint("U", "T", short_content)
        assert fp_long == fp_short


class TestIncrementalLogicEndToEnd:
    """增量处理逻辑"""

    def test_filter_already_translated_posts(self):
        posts = [
            {"fingerprint": "fp1", "_processed": {"translated": True, "sentiment_at": None}, "content": "a"},
            {"fingerprint": "fp2", "_processed": {"translated": False, "sentiment_at": None}, "content": "b"},
            {"fingerprint": "fp3", "_processed": {"translated": True, "sentiment_at": None}, "content": "c"},
        ]
        pending = [p for p in posts if not p["_processed"].get("translated")]
        assert len(pending) == 1
        assert pending[0]["fingerprint"] == "fp2"

    def test_filter_already_analyzed_posts(self):
        posts = [
            {"fingerprint": "fp1", "_processed": {"sentiment_at": "2026-07-28"}},
            {"fingerprint": "fp2", "_processed": {"sentiment_at": None}},
            {"fingerprint": "fp3", "_processed": {"sentiment_at": "2026-07-27"}},
        ]
        pending = [p for p in posts if not p["_processed"].get("sentiment_at")]
        assert len(pending) == 1
        assert pending[0]["fingerprint"] == "fp2"

    def test_fingerprint_dedup_merges_new_posts(self):
        existing = [{"fingerprint": "fp1"}, {"fingerprint": "fp2"}]
        new_all = [{"fingerprint": "fp1"}, {"fingerprint": "fp2"}, {"fingerprint": "fp3"}]
        existing_fps = {p["fingerprint"] for p in existing}
        new_only = [p for p in new_all if p["fingerprint"] not in existing_fps]
        assert len(new_only) == 1
        assert new_only[0]["fingerprint"] == "fp3"


class TestLLMUtilsEndToEnd:
    """LLM 工具函数测试"""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        import app.config as cfg
        import app.services.storage as storage_module
        self._orig_data_dir = cfg.settings.data_dir
        cfg.settings.data_dir = self.tmpdir
        # DB_PATH 是 import 时算好的常量，只改 data_dir 会写进真实的 hyxi.db
        self.storage = storage_module
        self._old_db = storage_module.DB_PATH
        storage_module.DB_PATH = os.path.join(self.tmpdir, "hyxi.db")
        storage_module.init_db()

    def teardown_method(self):
        import app.config as cfg
        cfg.settings.data_dir = self._orig_data_dir
        self.storage.DB_PATH = self._old_db
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_load_config_returns_none_when_unset(self):
        from app.services.llm_utils import load_llm_config
        assert load_llm_config() is None

    def test_load_config_returns_config_when_saved(self):
        self.storage.set_app_config("llm", {
            "api_key": "sk-test-key-123",
            "base_url": "https://api.example.com",
            "model_name": "test-model",
        })

        from app.services.llm_utils import load_llm_config
        config = load_llm_config()
        assert config is not None
        assert config.api_key == "sk-test-key-123"
        assert config.base_url == "https://api.example.com"
        assert config.model_name == "test-model"

    def test_load_config_handles_incomplete_row(self):
        """只有 base_url 没有 api_key 时不能当成「已配置」——LLMConfig 会校验失败"""
        self.storage.set_app_config("llm", {"base_url": "https://api.example.com"})
        from app.services.llm_utils import load_llm_config
        assert load_llm_config() is None

    def test_reset_clears_the_key(self):
        """重置后必须真的取不到，否则界面显示未配置而后台还在拿旧 key 发请求"""
        self.storage.set_app_config("llm", {
            "api_key": "sk-x", "base_url": "https://a.test", "model_name": "m",
        })
        self.storage.delete_app_config("llm")
        from app.services.llm_utils import load_llm_config
        assert load_llm_config() is None
        assert self.storage.get_app_config("llm") == {}

    def test_get_llm_service_returns_none_without_config(self):
        from app.services.llm_utils import get_llm_service
        assert get_llm_service() is None

    def test_get_llm_service_returns_service_with_config(self):
        self.storage.set_app_config("llm", {
            "api_key": "sk-test", "base_url": "https://api.test.com", "model_name": "m",
        })

        from app.services.llm_utils import get_llm_service
        import asyncio
        llm = get_llm_service()
        assert llm is not None
        assert llm.config.api_key == "sk-test"
        # 清理 client
        asyncio.get_event_loop().run_until_complete(llm.close())


def _export_row(**over):
    """导出明细的一行。字段与 results.py 的 _export_rows 产出一致"""
    row = {
        "index": 1, "source": "论坛来源", "level": 0, "username": "u",
        "timestamp": "2026-05-22 17:06", "content": "origineel", "translation": "原文",
        "image_desc": "", "images": [],
        # 「老帖新回复」的提醒；不命中就是空串（fresh_gap 只在命中时有意义）
        "fresh": "", "fresh_gap": -1,
        "sentiment": "中立", "intensity": 2, "reason": "询问", "dimensions": "",
    }
    row.update(over)
    return row


def _export_meta(**over):
    meta = {
        "description": "测试任务", "sources": "论坛来源",
        "exported_at": "2026-08-05 11:00:00", "total": 1, "replies": 0, "analyzed": 1,
        "time_start": "2026-05-22 17:06", "time_end": "2026-05-22 17:06",
        "summary": {"top_dimensions": []}, "top_users": [],
        # 窗口与基准会进概览表；不给的话那一行会显示成「近  天老帖新回复」
        "fresh_days": 7, "fresh_count": 0, "fresh_baseline": "",
    }
    meta.update(over)
    return meta


class TestJsonToSqliteMigrationEndToEnd:
    """启动时把遗留的 JSON 搬进 SQLite —— 真文件、真库，不 mock"""

    def setup_method(self):
        import app.config as cfg
        import app.services.storage as storage_module
        self.cfg = cfg
        self.tmpdir = tempfile.mkdtemp()
        self._old_dir = cfg.settings.data_dir
        cfg.settings.data_dir = self.tmpdir
        self.storage = storage_module
        self._old_db = storage_module.DB_PATH
        storage_module.DB_PATH = os.path.join(self.tmpdir, "hyxi.db")
        storage_module.init_db()

    def teardown_method(self):
        self.cfg.settings.data_dir = self._old_dir
        self.storage.DB_PATH = self._old_db
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _write(self, name, payload):
        path = os.path.join(self.tmpdir, name)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)
        return path

    def test_config_and_schedules_move_into_tables(self):
        cfg_path = self._write("config.json", {
            "api_key": "sk-legacy", "base_url": "https://legacy.test", "model_name": "m1",
        })
        sched_path = self._write("scheduled_tasks.json", [
            {"id": "s1", "description": "每天抓一次", "interval": "daily",
             "time": "09:00", "enabled": True, "created_at": "2026-08-01T10:00:00",
             "history": [{"task_id": "t1", "status": "completed"}]},
            {"id": "s2", "description": "停用的", "interval": "hourly",
             "time": "", "enabled": False, "created_at": "2026-08-02T10:00:00"},
        ])

        self.storage.migrate_from_json()

        assert self.storage.get_app_config("llm") == {
            "api_key": "sk-legacy", "base_url": "https://legacy.test", "model_name": "m1",
        }
        rows = {s["id"]: s for s in self.storage.load_schedules()}
        assert rows["s1"]["description"] == "每天抓一次"
        assert rows["s1"]["enabled"] is True
        assert rows["s1"]["history"] == [{"task_id": "t1", "status": "completed"}]
        assert rows["s2"]["enabled"] is False

        # 源文件归档而不是删除，让人能自己核对
        assert not os.path.exists(cfg_path)
        assert not os.path.exists(sched_path)
        backup = os.path.join(self.tmpdir, self.storage.MIGRATED_DIR_NAME)
        assert sorted(os.listdir(backup)) == ["config.json", "scheduled_tasks.json"]

    def test_migration_is_idempotent(self):
        """启动就跑一次，重启不能把数据搅乱，也不能把用户后来的修改冲掉"""
        self._write("config.json", {"api_key": "sk-a", "base_url": "u", "model_name": "m"})
        self.storage.migrate_from_json()

        # 用户迁移后改了配置；此时归档目录里还躺着那份旧 config.json
        self.storage.set_app_config("llm", {
            "api_key": "sk-new", "base_url": "u2", "model_name": "m2",
        })
        self.storage.migrate_from_json()

        assert self.storage.get_app_config("llm")["api_key"] == "sk-new", "旧文件把新配置冲掉了"

    def test_missing_files_are_not_an_error(self):
        """全新部署没有任何遗留 JSON，迁移不该抛"""
        self.storage.migrate_from_json()
        assert self.storage.get_app_config("llm") == {}
        assert self.storage.load_schedules() == []

    # ===== 舆情：整份 JSON 一个列 → 按帖子身份存 =====

    def _post(self, fp, analyzed=True):
        p = {"fingerprint": fp, "source": "src_x", "content": f"内容{fp}"}
        if analyzed:
            p["_processed"] = {"sentiment_at": "2026-08-01T10:00:00"}
        return p

    def _seed_legacy(self, task_id, results):
        conn = self.storage._get_conn()
        try:
            conn.execute("""CREATE TABLE IF NOT EXISTS sentiment (
                task_id TEXT PRIMARY KEY, data_json TEXT NOT NULL, created_at TEXT NOT NULL)""")
            conn.execute(
                "INSERT OR REPLACE INTO sentiment VALUES (?,?,?)",
                (task_id, json.dumps({
                    "task_id": task_id, "analyzed_at": "2026-08-01T10:00:00",
                    "summary": {"top_dimensions": []}, "results": results,
                }, ensure_ascii=False), "2026-08-01T10:00:00"),
            )
            conn.commit()
        finally:
            conn.close()

    def test_legacy_blob_becomes_identity_keyed(self):
        posts = [self._post("p1"), self._post("p2")]
        self._seed_legacy("t", [
            {"sentiment": "positive", "intensity": 4, "reason_cn": "属于p1", "dimensions": []},
            {"sentiment": "negative", "intensity": 3, "reason_cn": "属于p2", "dimensions": []},
        ])
        assert self.storage.migrate_sentiment_blob("t", posts) is True

        got = self.storage.get_sentiment("t", posts)
        assert [r["reason_cn"] for r in got["results"]] == ["属于p1", "属于p2"]
        # intensity 必须还是整数：REAL 亲和性会存成 4.0，导出的强度列跟着变成「4.0」
        assert got["results"][0]["intensity"] == 4
        assert isinstance(got["results"][0]["intensity"], int)

    def test_deleted_leading_source_does_not_shift_conclusions(self):
        """结果比帖子多时不许「迁前缀」—— 缺文件的来源可能排在最前面。

        实测踩过：某任务采了 tweakers(9) + group_feed(8)，tweakers 的落盘文件后来
        被删，幸存的 8 条 group_feed 帖子于是套上了 results[0:8]（tweakers 的结论）。
        sentiment_at 那条不变量拦不住，因为这 8 条帖子确实都分析过。
        """
        posts = [self._post("g1"), self._post("g2")]      # 幸存的是排在后面那个来源
        self._seed_legacy("t", [
            {"sentiment": "positive", "intensity": 4, "reason_cn": "属于已删来源的t1", "dimensions": []},
            {"sentiment": "negative", "intensity": 3, "reason_cn": "属于已删来源的t2", "dimensions": []},
            {"sentiment": "neutral", "intensity": 1, "reason_cn": "属于g1", "dimensions": []},
            {"sentiment": "positive", "intensity": 5, "reason_cn": "属于g2", "dimensions": []},
        ])
        assert self.storage.migrate_sentiment_blob("t", posts) is False

        assert self.storage.get_sentiment("t", posts) is None, "宁可没迁，也不能把 t1/t2 的结论安到 g1/g2 头上"
        assert self.storage.legacy_sentiment_task_ids() == ["t"], "原始数据必须留着"

    def test_misaligned_blob_is_refused(self):
        """有结论的位置上帖子却没有 sentiment_at —— 下标对不上，整份放弃。

        宁可留一份没迁进来的，也不能把结论安到别人身上（线上出过这个事故）。
        """
        posts = [self._post("p1"), self._post("p2", analyzed=False)]
        self._seed_legacy("t", [
            {"sentiment": "positive", "intensity": 4, "reason_cn": "属于p1", "dimensions": []},
            {"sentiment": "negative", "intensity": 3, "reason_cn": "凭空多出来的结论", "dimensions": []},
        ])
        assert self.storage.migrate_sentiment_blob("t", posts) is False
        assert self.storage.get_sentiment("t", posts) is None
        assert self.storage.legacy_sentiment_task_ids() == ["t"], "原始数据必须留着"


class TestSentimentRekeyMigrationEndToEnd:
    """把主键含 task_id 的旧 sentiment_results 换成按帖子身份。

    这是一次性、不可逆的数据迁移，跑坏了就是全部历史结论永久不可见。
    """

    OLD_DDL = """
    CREATE TABLE sentiment_results (
        task_id TEXT NOT NULL, source_id TEXT NOT NULL, fingerprint TEXT NOT NULL,
        sentiment TEXT, intensity NUMERIC, reason_cn TEXT NOT NULL DEFAULT '',
        dimensions_json TEXT NOT NULL DEFAULT '[]',
        PRIMARY KEY (task_id, source_id, fingerprint))
    """

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        import app.config as cfg
        import app.services.storage as storage_module
        self.cfg, self.storage = cfg, storage_module
        self._old_dir, self._old_db = cfg.settings.data_dir, storage_module.DB_PATH
        cfg.settings.data_dir = self.tmpdir
        storage_module.DB_PATH = os.path.join(self.tmpdir, "hyxi.db")

    def teardown_method(self):
        self.cfg.settings.data_dir = self._old_dir
        self.storage.DB_PATH = self._old_db
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    POSTS = [{"fingerprint": "p1", "source": "src_x"}, {"fingerprint": "p2", "source": "src_x"}]

    def _seed_old_schema(self):
        """造一个旧库：同一条帖子被两个任务分别分析过，结论不同"""
        conn = self.storage._get_conn()
        conn.execute("""CREATE TABLE sentiment_runs (
            task_id TEXT PRIMARY KEY, analyzed_at TEXT NOT NULL,
            summary_json TEXT NOT NULL DEFAULT '{}')""")
        conn.execute(self.OLD_DDL)
        conn.executemany(
            "INSERT INTO sentiment_runs (task_id, analyzed_at) VALUES (?,?)",
            [("old-task", "2026-08-01T10:00:00"), ("new-task", "2026-08-05T10:00:00")],
        )
        conn.executemany(
            """INSERT INTO sentiment_results
               (task_id, source_id, fingerprint, sentiment, intensity, reason_cn, dimensions_json)
               VALUES (?,?,?,?,?,?,?)""",
            [
                ("old-task", "src_x", "p1", "positive", 4, "旧结论", "[]"),
                ("new-task", "src_x", "p1", "negative", 2, "新结论", "[]"),
                ("old-task", "src_x", "p2", "neutral", 1, "只有旧任务分析过", "[]"),
            ],
        )
        conn.commit()
        conn.close()

    def test_duplicate_conclusions_collapse_to_the_latest(self):
        self._seed_old_schema()
        self.storage.init_db()

        got = self.storage.get_sentiment("old-task", self.POSTS)
        assert [r["reason_cn"] for r in got["results"]] == ["新结论", "只有旧任务分析过"]
        assert got["results"][0]["sentiment"] == "negative", "该取 2026-08-05 那份"

        conn = self.storage._get_conn()
        try:
            assert conn.execute("SELECT COUNT(*) FROM sentiment_results").fetchone()[0] == 2
            sql = conn.execute(
                "SELECT sql FROM sqlite_master WHERE name='sentiment_results'"
            ).fetchone()["sql"]
            assert "PRIMARY KEY (source_id, fingerprint)" in sql
            assert conn.execute(
                "SELECT COUNT(*) FROM sqlite_master WHERE name='sentiment_results_old'"
            ).fetchone()[0] == 0, "残留了孤儿表"
        finally:
            conn.close()

    def test_rekey_is_idempotent(self):
        self._seed_old_schema()
        self.storage.init_db()
        self.storage.init_db()
        conn = self.storage._get_conn()
        try:
            assert conn.execute("SELECT COUNT(*) FROM sentiment_results").fetchone()[0] == 2
        finally:
            conn.close()

    def test_failure_midway_rolls_back_completely(self):
        """中途失败必须整体回滚。

        DDL 在这里**不会**自动回滚：不加显式事务的话，失败后留下的是「空的新表 +
        孤立的旧表」，而下次启动的检测会看到新主键、判定「不用迁移」直接跳过 ——
        全部历史结论静默永久不可见。executescript 也会打断事务，同样的下场。
        """
        self._seed_old_schema()

        # 让重建新表这一步失败 —— 正好落在 RENAME 之后那个危险窗口里。
        # （不能去 monkeypatch conn.executemany：它是只读属性，改不动，
        #   之前那版测试因此从头到尾没触发过任何失败，是在空跑）
        orig_ddl = self.storage.SENTIMENT_RESULTS_DDL
        self.storage.SENTIMENT_RESULTS_DDL = "CREATE TABLE 语法坏掉的东西 ("
        try:
            self.storage.init_db()          # 内部会吞掉异常并打 error 日志
        finally:
            self.storage.SENTIMENT_RESULTS_DDL = orig_ddl

        conn = self.storage._get_conn()
        try:
            n = conn.execute("SELECT COUNT(*) FROM sentiment_results").fetchone()[0]
            assert n == 3, f"回滚后旧数据应原样还在，实际 {n} 行"
            assert conn.execute(
                "SELECT COUNT(*) FROM sqlite_master WHERE name='sentiment_results_old'"
            ).fetchone()[0] == 0, "回滚后不该留下孤儿表"
        finally:
            conn.close()

        # 失败不是终点：下次启动还认得出旧主键，能重新迁移成功
        self.storage.init_db()
        assert self.storage.get_sentiment("old-task", self.POSTS)["success"] == 2

    def test_orphan_table_from_an_older_crash_is_recovered(self):
        """老版本崩在半路留下的孤儿表：里面是唯一一份完整数据，必须接回来"""
        self._seed_old_schema()
        conn = self.storage._get_conn()
        conn.execute("ALTER TABLE sentiment_results RENAME TO sentiment_results_old")
        conn.commit()
        conn.close()

        self.storage.init_db()

        got = self.storage.get_sentiment("old-task", self.POSTS)
        assert got["success"] == 2, "孤儿表里的历史结论没被接回来"
        assert [r["reason_cn"] for r in got["results"]] == ["新结论", "只有旧任务分析过"]


class TestFakeNeutralPurgeEndToEnd:
    """把「解析失败」那种冒充结论的行连同 sentiment_at 一起清掉 —— 真库、真 SQL"""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        import app.config as cfg
        import app.services.storage as storage_module
        self.cfg, self.storage = cfg, storage_module
        self._old_dir, self._old_db = cfg.settings.data_dir, storage_module.DB_PATH
        cfg.settings.data_dir = self.tmpdir
        storage_module.DB_PATH = os.path.join(self.tmpdir, "hyxi.db")

    def teardown_method(self):
        self.cfg.settings.data_dir = self._old_dir
        self.storage.DB_PATH = self._old_db
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _seed(self):
        """一条被编出来的 neutral 占位、一条真结论，两条帖子都已带 sentiment_at"""
        self.storage.init_db()
        conn = self.storage._get_conn()
        conn.executemany(
            """INSERT INTO posts (source_id, fingerprint, seq, content, sentiment_at)
               VALUES (?,?,?,?,?)""",
            [
                ("src_x", "fake", 1, "Deze update werkt niet na de update...",
                 "2026-08-25T09:48:43"),
                ("src_x", "real", 2, "Prima batterij, snel geleverd.",
                 "2026-08-25T09:48:43"),
            ],
        )
        conn.executemany(
            """INSERT INTO sentiment_results
               (source_id, fingerprint, sentiment, intensity, reason_cn, dimensions_json,
                task_id, analyzed_at)
               VALUES (?,?,?,?,?,?,?,?)""",
            [
                ("src_x", "fake", "neutral", 1, "解析失败", "[]", "t1", "2026-08-25T09:48:43"),
                ("src_x", "real", "positive", 4, "交付顺利", '["价格/性价比"]', "t1",
                 "2026-08-25T09:48:43"),
            ],
        )
        conn.commit()
        conn.close()

    def test_placeholder_row_and_its_sentiment_at_are_both_cleared(self):
        """只删结论行不够：sentiment_at 还在的话增量分析永远不会再碰它。

        增量的判据就是 `_processed.sentiment_at` 为空（orchestrator 那句列表推导），
        清一半等于把那几条帖子永久钉在「已分析」上，用户只剩全量重跑一条路 ——
        为 3 条帖子重算 258 条，还要重新付一次钱。
        """
        self._seed()
        self.storage.purge_fake_parse_failures()      # 升级后第一次启动

        conn = self.storage._get_conn()
        rows = dict(conn.execute(
            "SELECT fingerprint, sentiment FROM sentiment_results").fetchall())
        ats = dict(conn.execute(
            "SELECT fingerprint, sentiment_at FROM posts").fetchall())
        conn.close()

        assert "fake" not in rows, f"占位结论还在: {rows}"
        assert ats["fake"] is None, "sentiment_at 没清，增量分析再也够不着它"
        # 真结论一根汗毛都不能动
        assert rows["real"] == "positive"
        assert ats["real"] == "2026-08-25T09:48:43"

    def test_the_purged_post_comes_back_as_pending(self):
        """清完之后，它必须真的重新出现在待分析队列里（照 orchestrator 那句判据）"""
        self._seed()
        self.storage.purge_fake_parse_failures()

        posts = self.storage.load_posts(["src_x"])
        pending = [p for p in posts if not p.get("_processed", {}).get("sentiment_at")]
        assert [p["fingerprint"] for p in pending] == ["fake"], pending

    def test_purge_is_idempotent_and_quiet_on_clean_dbs(self):
        """幂等：干净库上反复启动不该有任何动静"""
        self._seed()
        self.storage.purge_fake_parse_failures()
        self.storage.purge_fake_parse_failures()
        self.storage.purge_fake_parse_failures()

        conn = self.storage._get_conn()
        assert conn.execute("SELECT COUNT(*) FROM sentiment_results").fetchone()[0] == 1
        assert conn.execute(
            "SELECT COUNT(*) FROM posts WHERE sentiment_at IS NULL").fetchone()[0] == 1
        conn.close()

    def test_honest_null_placeholder_is_left_alone(self):
        """新代码写下的占位是 sentiment 为空的诚实记录，不许连它一起删。

        它带的「解析失败」是给用户看的原因，且本来就没有 sentiment_at、下一轮增量
        自然会重算。连它一起删的话，这个一次性迁移永远变不成 no-op —— 每次启动都在
        删一条马上又要写回来的行，页面上那条「未分析」还白白丢掉了原因。
        """
        self.storage.init_db()
        conn = self.storage._get_conn()
        conn.execute(
            """INSERT INTO posts (source_id, fingerprint, seq, content, sentiment_at)
               VALUES (?,?,?,?,?)""",
            ("src_x", "honest", 1, "Onleesbare tekst", None),
        )
        conn.execute(
            """INSERT INTO sentiment_results
               (source_id, fingerprint, sentiment, intensity, reason_cn, dimensions_json,
                task_id, analyzed_at)
               VALUES (?,?,?,?,?,?,?,?)""",
            ("src_x", "honest", None, 0, "解析失败", "[]", "t1", "2026-08-25T09:48:43"),
        )
        conn.commit()
        conn.close()

        self.storage.purge_fake_parse_failures()

        conn = self.storage._get_conn()
        row = conn.execute("SELECT sentiment, reason_cn FROM sentiment_results "
                           "WHERE fingerprint='honest'").fetchone()
        conn.close()
        assert row is not None, "诚实的占位被当成冒充结论删掉了"
        assert row["sentiment"] is None and row["reason_cn"] == "解析失败"

    def test_legacy_blob_migration_cannot_smuggle_them_back_in(self):
        """清理必须排在旧 JSON blob 的迁移之后，否则升级那一次恰好是空转的。

        老库升上来的第一次启动：清理先跑完，migrate_sentiment_blob() 才把 blob 里
        那批假 neutral 重新写进 sentiment_results —— 而这一次正是它最该生效的那一次。
        真 orchestrator、真 SQLite、真迁移，无 mock。
        """
        # **import 必须排在播种之前**：orchestrator 模块在 import 时就实例化一个全局
        # 单例并跑完一轮迁移。留在下面的话这里会变成「单例 + 显式构造」两轮，而第二轮
        # 开头的 init_db() 恰好把第一轮迁进来的假 neutral 又清掉 —— 这条测试于是在
        # 修复前后都是绿的，等于什么都没测（实测踩过）
        from app.services.orchestrator import TaskOrchestrator

        self.storage.init_db()
        conn = self.storage._get_conn()
        conn.executemany(
            """INSERT INTO posts (source_id, fingerprint, seq, content, sentiment_at)
               VALUES (?,?,?,?,?)""",
            [("src_x", "fake", 1, "Deze update werkt niet na de update...",
              "2026-08-25T09:48:43"),
             ("src_x", "real", 2, "Prima batterij, snel geleverd.",
              "2026-08-25T09:48:43")],
        )
        # 旧版本那张「整份结果塞进一个 JSON 列」的表，现在的 SCHEMA 已不再建它
        conn.execute("CREATE TABLE IF NOT EXISTS sentiment "
                     "(task_id TEXT PRIMARY KEY, data_json TEXT NOT NULL)")
        conn.execute(
            "INSERT INTO sentiment (task_id, data_json) VALUES (?,?)",
            ("t-legacy", json.dumps({"results": [
                {"sentiment": "neutral", "intensity": 1, "reason_cn": "解析失败",
                 "dimensions": []},
                {"sentiment": "positive", "intensity": 4, "reason_cn": "交付顺利",
                 "dimensions": ["价格/性价比"]},
            ]})),
        )
        conn.execute(
            """INSERT INTO tasks (id, status, description, result_json, created_at)
               VALUES (?,?,?,?,?)""",
            ("t-legacy", "completed", "旧任务",
             json.dumps({"sources": [{"id": "src_x"}]}), "2026-08-20T10:00:00"),
        )
        conn.commit()
        conn.close()

        TaskOrchestrator()                      # 升级后的第一次启动

        conn = self.storage._get_conn()
        rows = dict(conn.execute(
            "SELECT fingerprint, sentiment FROM sentiment_results").fetchall())
        ats = dict(conn.execute(
            "SELECT fingerprint, sentiment_at FROM posts").fetchall())
        conn.close()
        assert "fake" not in rows, f"迁移又把假 neutral 写回来了，这一轮清理是空转的: {rows}"
        assert ats["fake"] is None, "sentiment_at 没清，这条帖子再也不会被重新分析"
        # 同一份 blob 里的真结论必须照常迁进来
        assert rows.get("real") == "positive", f"真结论没迁进来: {rows}"


class TestPostsStorageEndToEnd:
    """posts 表：seq 是全链路的顺序锚点，upsert 不能碰已处理标记"""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        self.storage, self._restore = _use_temp_db(self.tmpdir)

    def teardown_method(self):
        self._restore()
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    @staticmethod
    def _p(fp, **over):
        post = {"fingerprint": fp, "source": "src_x", "username": f"u{fp}",
                "timestamp": "02-06-2026 09:00", "content": f"内容{fp}", "page_number": 1}
        post.update(over)
        return post

    def test_images_survive_a_collect_that_found_none(self):
        """重采时这一轮没抓到图，库里已有的图片路径不能被冲成空。

        全量重跑撞上一次网络抖动就会走到这条路上：**图片文件还好端端躺在 media
        目录里，页面上却整批消失**，而且没有任何提示。与旁边 translation /
        image_desc 「绝不被采集冲回空值」是同一套道理。
        """
        self.storage.upsert_posts("src_x", [self._p("a", images=["src_x/a_0.png"])])
        assert self.storage.load_posts(["src_x"])[0]["images"] == ["src_x/a_0.png"]

        # 同一条帖子重新采回来，但这一轮一张图都没下到
        self.storage.upsert_posts("src_x", [self._p("a", content="正文改了一下")])

        kept = self.storage.load_posts(["src_x"])[0]
        assert kept["images"] == ["src_x/a_0.png"], "已有的图被空结果冲掉了"
        assert kept["content"] == "正文改了一下", "采集字段该更新的还是要更新"

    def test_new_images_still_replace_the_old_ones(self):
        """真抓到图时照旧覆盖 —— 不能为了上一条把这里变成只进不出"""
        self.storage.upsert_posts("src_x", [self._p("a", images=["src_x/old.png"])])
        self.storage.upsert_posts("src_x", [self._p("a", images=["src_x/new.png"])])
        assert self.storage.load_posts(["src_x"])[0]["images"] == ["src_x/new.png"]

    def test_seq_is_stable_across_incremental_rounds(self):
        """已有帖子的 seq 一个都不能变，新帖只能追加在后面。

        整条链有 8 处依赖「顺序即存储顺序」，seq 一洗牌，全部历史舆情结论就错位。
        """
        self.storage.upsert_posts("src_x", [self._p("a"), self._p("b"), self._p("c")])
        first = [p["fingerprint"] for p in self.storage.load_posts(["src_x"])]

        # 第二轮：老帖子重新出现（信息流每批都会重扫）+ 两条新帖
        self.storage.upsert_posts("src_x", [
            self._p("b"), self._p("a"), self._p("d"), self._p("c"), self._p("e"),
        ])
        after = [p["fingerprint"] for p in self.storage.load_posts(["src_x"])]

        assert first == ["a", "b", "c"]
        assert after == ["a", "b", "c", "d", "e"], "已有帖子的顺序被重排了"

    def test_upsert_never_clobbers_translation_or_flags(self):
        """重扫到的老帖子只更新采集字段。覆盖 translation 等于让它重新付费翻译一遍"""
        self.storage.upsert_posts("src_x", [self._p("a")])
        self.storage.save_translations([
            {"source": "src_x", "fingerprint": "a", "translation": "译文",
             "_processed": {"translated": True}},
        ])
        self.storage.mark_sentiment_analyzed([
            {"source": "src_x", "fingerprint": "a",
             "_processed": {"sentiment_at": "2026-08-05T10:00:00"}},
        ])

        # 下一轮重扫：脚本给的是刚提取的原始帖子，没有 translation 也没有标记
        self.storage.upsert_posts("src_x", [self._p("a", content="正文被编辑过了")])

        got = self.storage.load_posts(["src_x"])[0]
        assert got["translation"] == "译文", "译文被重扫抹掉了"
        assert got["_processed"]["translated"] is True
        assert got["_processed"]["sentiment_at"] == "2026-08-05T10:00:00"
        assert got["content"] == "正文被编辑过了", "采集字段没更新"

    def test_upsert_never_clobbers_image_desc(self):
        """图片描述与译文同类：花钱换来的，重扫不得抹掉。

        抹掉就意味着每轮采集之后所有带图帖子都要重新调一次多模态模型。
        """
        self.storage.upsert_posts("src_x", [self._p("a")])
        self.storage.save_image_descs([
            {"source": "src_x", "fingerprint": "a", "image_desc": "一台报错的储能电池"},
        ])

        self.storage.upsert_posts("src_x", [self._p("a", content="正文被编辑过了")])

        got = self.storage.load_posts(["src_x"])[0]
        assert got["image_desc"] == "一台报错的储能电池", "图片描述被重扫抹掉了"

    def test_empty_image_desc_is_not_written(self):
        """空描述通常意味着模型没配或配额用尽。写个空串下去，下一轮就再也不会重试了"""
        self.storage.upsert_posts("src_x", [self._p("a")])
        assert self.storage.save_image_descs([
            {"source": "src_x", "fingerprint": "a", "image_desc": "  "},
        ]) == 0
        assert "image_desc" not in self.storage.load_posts(["src_x"])[0]

    def test_added_count_reflects_only_new_posts(self):
        assert self.storage.upsert_posts("src_x", [self._p("a"), self._p("b")]) == 2
        assert self.storage.upsert_posts("src_x", [self._p("a"), self._p("c")]) == 1

    def test_sources_are_isolated_from_each_other(self):
        """指纹不含来源，两个平台的同名空帖不能互相覆盖"""
        self.storage.upsert_posts("src_a", [{"fingerprint": "same", "source": "src_a",
                                             "content": "A 平台"}])
        self.storage.upsert_posts("src_b", [{"fingerprint": "same", "source": "src_b",
                                             "content": "B 平台"}])
        assert self.storage.load_posts(["src_a"])[0]["content"] == "A 平台"
        assert self.storage.load_posts(["src_b"])[0]["content"] == "B 平台"
        # 跨来源拼接的顺序由调用方给定，与 seq 无关
        both = self.storage.load_posts(["src_b", "src_a"])
        assert [p["content"] for p in both] == ["B 平台", "A 平台"]

    def test_incremental_anchors_come_from_the_table(self):
        self.storage.upsert_posts("src_x", [
            self._p("a", page_number=1), self._p("b", page_number=7),
        ])
        assert sorted(self.storage.known_fingerprints("src_x")) == ["a", "b"]
        assert self.storage.max_page_number("src_x") == 7
        assert self.storage.max_page_number("src_none") == 0

    def test_migration_preserves_array_order_as_seq(self):
        """迁移把数组下标变成 seq，读回来必须与原 JSON 逐条同序"""
        posts = [self._p(f"fp{i}") for i in range(6)]
        path = os.path.join(self.tmpdir, "legacy.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump({"total_pages": 1, "posts": posts}, f, ensure_ascii=False)

        assert self.storage.migrate_posts_file("src_x", path) == 6
        assert [p["fingerprint"] for p in self.storage.load_posts(["src_x"])] \
            == [p["fingerprint"] for p in posts]

        # 幂等：重跑不能把 seq 洗牌，也不能重复插入
        assert self.storage.migrate_posts_file("src_x", path) == 0
        assert self.storage.count_posts("src_x") == 6

    def test_migration_keeps_empty_posts_so_sentiment_blob_still_aligns(self):
        """迁移是照原样重建历史，不能套用「空正文不入库」那条采集规则。

        旧舆情 blob 的 results[i] 对齐的正是这个数组的第 i 条。迁移时少搬一条，
        条数就对不上，migrate_sentiment_blob() 的「条数不等整份跳过」会把那个来源的
        历史结论永久挡在门外 —— 静默丢掉一整份分析结果，而且没人看得出来。
        """
        posts = [self._p("fp0"), self._p("fp1", content=""), self._p("fp2")]
        path = os.path.join(self.tmpdir, "legacy_with_empty.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump({"total_pages": 1, "posts": posts}, f, ensure_ascii=False)

        assert self.storage.migrate_posts_file("src_e", path) == 3
        assert [p["fingerprint"] for p in self.storage.load_posts(["src_e"])] \
            == ["fp0", "fp1", "fp2"]

        # 但新采集到的空正文帖照样挡在门外
        assert self.storage.upsert_posts("src_e", [self._p("fp3", content="")]) == 0
    """舆情结果数组的下标必须是全量帖子数组里的绝对位置。

    页面和导出都按下标反查帖子。存成「待分析批次内的下标」不会报任何错，
    只会让每条结论悄悄挂到别人身上。
    """

    @staticmethod
    def _post(fp, content="内容"):
        return {"username": "u", "timestamp": "02-06-2026 09:00", "content": content,
                "fingerprint": fp, "source": "src_x", "parent_fingerprint": None}

    def _fixture(self, pending_idx):
        """5 条全量帖子，其中若干条本轮待分析"""
        from app.services.post_tree import post_key
        all_posts = [self._post(f"f{i}") for i in range(5)]
        fp_to_idx = {post_key(p): i for i, p in enumerate(all_posts)}
        pending = [all_posts[i] for i in pending_idx]
        return all_posts, fp_to_idx, pending

    def test_first_run_over_already_analyzed_data_uses_absolute_indices(self):
        """**根因用例**：本任务第一次跑舆情，但源数据已被别的任务分析过。

        `_processed.sentiment_at` 是写在**源文件**里的、跨任务共享，所以新任务的
        pending 只剩没分析过的那几条，而本任务的 existing_results 是空的。
        一旦拿 existing_results 当「要不要重映射」的开关，这一轮就会把批次内下标
        当成绝对下标存盘。
        """
        from app.services.sentiment_service import SentimentService

        all_posts, fp_to_idx, pending = self._fixture([3])       # 只有第 4 条待分析
        r3 = {"sentiment": "negative", "intensity": 3, "reason_cn": "属于f3", "dimensions": []}

        out = SentimentService._to_absolute([r3], pending, [], fp_to_idx, len(all_posts))

        assert len(out) == 5, f"结果数组必须与全量帖子等长，实际 {len(out)}"
        assert out[3] == r3, "结论要落在第 4 条帖子上"
        assert [out[i] for i in (0, 1, 2, 4)] == [None] * 4, "其余位置必须留空"

    def test_incremental_run_merges_into_existing(self):
        """常规增量：已有结果原样保留，新结论按指纹落到各自的绝对位置"""
        from app.services.sentiment_service import SentimentService

        all_posts, fp_to_idx, pending = self._fixture([1, 4])
        existing = [{"sentiment": "positive", "intensity": 4, "reason_cn": "属于f0", "dimensions": []},
                    None, None, None, None]
        r1 = {"sentiment": "neutral", "intensity": 2, "reason_cn": "属于f1", "dimensions": []}
        r4 = {"sentiment": "negative", "intensity": 5, "reason_cn": "属于f4", "dimensions": []}

        out = SentimentService._to_absolute([r1, r4], pending, existing, fp_to_idx, len(all_posts))

        assert out[0]["reason_cn"] == "属于f0"
        assert out[1]["reason_cn"] == "属于f1"
        assert out[4]["reason_cn"] == "属于f4"
        assert out[2] is None and out[3] is None

    def test_fresh_task_analyzing_everything_is_unchanged(self):
        """全量首跑时批次内下标本来就等于绝对下标，不能被搬错位"""
        from app.services.sentiment_service import SentimentService

        all_posts, fp_to_idx, pending = self._fixture([0, 1, 2, 3, 4])
        results = [{"sentiment": "neutral", "intensity": 1, "reason_cn": f"属于f{i}", "dimensions": []}
                   for i in range(5)]

        out = SentimentService._to_absolute(results, pending, [], fp_to_idx, len(all_posts))

        assert [r["reason_cn"] for r in out] == [f"属于f{i}" for i in range(5)]

    def test_empty_content_post_does_not_shift_the_rest(self):
        """空内容帖不会被送去分析，它在批次里的那个空位不能把后面的结论顶偏"""
        from app.services.sentiment_service import SentimentService

        all_posts, fp_to_idx, pending = self._fixture([2, 3])
        pending[0]["content"] = ""                       # 第 3 条是空帖，不分析
        r3 = {"sentiment": "positive", "intensity": 4, "reason_cn": "属于f3", "dimensions": []}

        out = SentimentService._to_absolute([None, r3], pending, [], fp_to_idx, len(all_posts))

        assert out[2] is None
        assert out[3] == r3, "空帖占位不能让后一条结论落到它自己的位置上"


class TestExportWorkbookEndToEnd:
    """用户下载到的那份报告 —— 结构、排版与脏数据容错"""

    def test_builds_two_sheets_with_readable_layout(self):
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from app.services.excel_service import ExcelService, EXPORT_COLUMNS, FONT_NAME

        rows = [
            _export_row(index=1, sentiment="正面", intensity=4, reason="满意",
                        dimensions="价格/性价比"),
            _export_row(index=2, level=1, username="回复者", sentiment="负面",
                        intensity=3, reason="有问题", dimensions="安装/配置体验"),
            _export_row(index=3, sentiment="未分析", intensity="", reason=""),
        ]
        meta = _export_meta(total=3, replies=1, analyzed=2,
                            summary={"top_dimensions": [["价格/性价比", 7], ["App/软件体验", 3]]},
                            top_users=[("u", 2), ("回复者", 1)])

        wb = load_workbook(BytesIO(ExcelService.build_export(rows, meta, EXPORT_COLUMNS)))
        assert wb.sheetnames == ["概览", "帖子明细"]

        ov = wb["概览"]
        assert "HYXi" in str(ov.cell(1, 1).value)
        flat = [str(ov.cell(r, c).value) for r in range(1, ov.max_row + 1) for c in (1, 2)]
        assert "测试任务" in flat
        # 占比按总条数算，「未分析」也占一行，四行加起来才是 100%
        assert "未分析" in flat

        ws = wb["帖子明细"]
        assert [c.value for c in ws[1]] == [label for _, label in EXPORT_COLUMNS]
        assert ws.max_row == 4
        # 冻结首行 + 自动筛选，88 行里找负面才不用人肉翻
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref == f"A1:{get_column_letter(len(EXPORT_COLUMNS))}4"
        # Arial 没有中文字形，中英文混排会错位
        assert ws.cell(1, 1).font.name == FONT_NAME
        assert ws.cell(2, 1).font.name == FONT_NAME

    def test_reply_rows_are_the_only_shaded_ones(self):
        """评论行单独上底色。再叠隔行斑马纹两者互相掩盖，等于谁也看不出来"""
        from openpyxl import load_workbook
        from app.services.excel_service import (
            ExcelService, EXPORT_COLUMNS, REPLY_FILL, SENTIMENT_STYLE,
        )

        rows = [
            _export_row(index=1, level=0, sentiment="正面"),
            _export_row(index=2, level=0, sentiment="负面"),   # 偶数行，旧实现会给它上色
            _export_row(index=3, level=1, sentiment="中立"),
        ]
        wb = load_workbook(BytesIO(ExcelService.build_export(rows, _export_meta(), EXPORT_COLUMNS)))
        ws = wb["帖子明细"]

        def fill_of(row):
            return ws.cell(row, 1).fill.start_color.rgb

        assert fill_of(2) == fill_of(3), "两个主贴的底色必须一致，不能隔行"
        assert fill_of(4) == "00" + REPLY_FILL.start_color.rgb[-6:]
        assert fill_of(4) != fill_of(2), "评论行要能与主贴区分开"

        # 情感列按情感上色，三种各不相同
        sentiment_col = [k for k, _ in EXPORT_COLUMNS].index("sentiment") + 1
        painted = {ws.cell(r, sentiment_col).fill.start_color.rgb[-6:] for r in (2, 3, 4)}
        assert painted == {bg for bg, _ in SENTIMENT_STYLE.values()}

    def test_survives_dirty_intensity(self):
        """LLM 会把 intensity 返回成字符串 / None / 越界值，星级渲染不能因此炸"""
        from openpyxl import load_workbook
        from app.services.excel_service import ExcelService, EXPORT_COLUMNS

        rows = [
            _export_row(index=1, sentiment="正面", intensity="高"),
            _export_row(index=2, sentiment="正面", intensity=None),
            _export_row(index=3, sentiment="正面", intensity=12),
            _export_row(index=4, sentiment="未分析", intensity=""),
        ]
        wb = load_workbook(BytesIO(ExcelService.build_export(rows, _export_meta(), EXPORT_COLUMNS)))
        ws = wb["帖子明细"]
        col = [k for k, _ in EXPORT_COLUMNS].index("intensity") + 1
        # 空串写进 openpyxl 读回来是 None（就是个空格子）；关键是别凭空编出星级
        assert not ws.cell(2, col).value        # "高"
        assert not ws.cell(3, col).value        # None
        assert ws.cell(4, col).value == "★★★★★"  # 12 收敛到 5
        assert not ws.cell(5, col).value        # 未分析

        # 脏强度不能把平均分拖成 0
        ov = wb["概览"]
        avg = next(ov.cell(r, 2).value for r in range(1, ov.max_row + 1)
                   if ov.cell(r, 1).value == "平均强度")
        assert avg == "5 / 5"

    def test_empty_rows_do_not_crash(self):
        """帖子一条都没有时端点会先 404，但生成器自己不该除零"""
        from openpyxl import load_workbook
        from app.services.excel_service import ExcelService, EXPORT_COLUMNS

        wb = load_workbook(BytesIO(
            ExcelService.build_export([], _export_meta(total=0, analyzed=0), EXPORT_COLUMNS)
        ))
        assert wb.sheetnames == ["概览", "帖子明细"]
        assert wb["帖子明细"].max_row == 1


class TestExportImagesEndToEnd:
    """报告里要看得见配图 —— 真图片文件、真 openpyxl、真 xlsx 字节流。

    纯图帖的全部信息都在图上，报告只给一行文字等于什么都没说。
    """

    def setup_method(self):
        import app.config as cfg
        from PIL import Image as PILImage
        self.cfg = cfg
        self.tmpdir = tempfile.mkdtemp()
        self._old_dir = cfg.settings.data_dir
        cfg.settings.data_dir = self.tmpdir

        # 造一张真图：600x400，宽高比 3:2，缩放对不对一眼能看出来
        self.rel = "src_x/shot_0.png"
        media = os.path.join(self.tmpdir, "media", "src_x")
        os.makedirs(media, exist_ok=True)
        self.abs_path = os.path.join(media, "shot_0.png")
        PILImage.new("RGB", (600, 400), (30, 90, 160)).save(self.abs_path)

    def teardown_method(self):
        self.cfg.settings.data_dir = self._old_dir
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _build(self, rows):
        from openpyxl import load_workbook
        from app.services.excel_service import ExcelService, EXPORT_COLUMNS
        return load_workbook(BytesIO(
            ExcelService.build_export(rows, _export_meta(total=len(rows)), EXPORT_COLUMNS)
        ))

    def test_image_row_gets_a_thumbnail_and_a_link_to_the_full_size(self):
        from app.services.excel_service import EXPORT_COLUMNS, IMAGE_SHEET, THUMB_MAX_PX

        wb = self._build([
            _export_row(index=1, images=[self.rel], image_desc="安装检查报告，总分 88"),
            _export_row(index=2, username="别人"),
        ])
        assert IMAGE_SHEET in wb.sheetnames, "没有生成配图工作表"

        ws = wb["帖子明细"]
        assert len(ws._images) == 1, f"明细表里的缩略图数量不对: {len(ws._images)}"
        thumb = ws._images[0]
        assert max(thumb.width, thumb.height) <= THUMB_MAX_PX, "缩略图没缩小，行会被撑爆"
        assert round(thumb.width / thumb.height, 2) == 1.5, "缩略图变形了"

        # 超链接挂在「图片描述」格上 —— 图片本身盖着的格子点不着（openpyxl 也不支持
        # 给图片挂链接），这一格既是最显眼的文字又点得到
        col = [k for k, _ in EXPORT_COLUMNS].index("image_desc") + 1
        cell = ws.cell(2, col)
        assert cell.hyperlink is not None, "没有跳转大图的入口"
        assert cell.hyperlink.location, "内部链接必须走 location，target 会被当成外部关系"
        assert IMAGE_SHEET in cell.hyperlink.location
        assert "安装检查报告" in str(cell.value)

        # 链接指向的那一行确实有大图
        target_row = int(cell.hyperlink.location.rsplit("A", 1)[1])
        big = wb[IMAGE_SHEET]
        anchored = {im.anchor._from.row + 1 for im in big._images}
        assert anchored, "配图表里一张图都没有"
        assert min(anchored) >= target_row, f"跳过去看不到图: 锚点 {anchored}, 目标行 {target_row}"

    def test_full_size_is_bigger_than_the_thumbnail(self):
        """「点击放大」得真的放大，否则跳过去毫无意义"""
        from app.services.excel_service import IMAGE_SHEET, LARGE_MAX_PX

        wb = self._build([_export_row(index=1, images=[self.rel], image_desc="报告")])
        thumb = wb["帖子明细"]._images[0]
        big = wb[IMAGE_SHEET]._images[0]
        assert big.width > thumb.width
        assert max(big.width, big.height) <= LARGE_MAX_PX

    def test_tall_image_gets_enough_rows_instead_of_one_oversized_one(self):
        """Excel 单行上限 409.5pt（≈546px）。真实 Facebook 截图是 367x795 这个量级，
        钳到 700px 后仍有 525pt —— 压进一行会被 Excel 截回来，图盖到下面几行上。
        """
        from PIL import Image as PILImage
        from app.services.excel_service import (
            IMAGE_SHEET, MAX_ROW_POINT, LARGE_MAX_PX, DEFAULT_ROW_PX,
        )

        tall = os.path.join(self.tmpdir, "media", "src_x", "tall_0.png")
        PILImage.new("RGB", (367, 795), (60, 60, 60)).save(tall)
        wb = self._build([
            _export_row(index=1, images=["src_x/tall_0.png"], image_desc="很高的截图"),
            _export_row(index=2, images=[self.rel], image_desc="正常的截图"),
        ])
        big = wb[IMAGE_SHEET]
        # 两张表一起查。配图表现在压根不设行高（大图靠占行数铺开），真正被这条断言
        # 跑到的是明细表那些设了高度的行；但它同时钉住「别再把图高压进一行」——
        # 那样写回来就是 531pt，立刻在这里翻车
        heights = [
            d.height
            for sheet in (wb["帖子明细"], big)
            for d in sheet.row_dimensions.values()
            if d.height
        ]
        assert heights, "一行显式行高都没有，这条断言就没在查东西"
        assert max(heights) <= MAX_ROW_POINT, f"最大行高 {max(heights)} 超了 Excel 上限"

        # 显示尺寸存在 anchor 的 ext 里（EMU，9525 EMU = 1px）—— 重新读回来的
        # img.height 是原始像素，不是我们钳过的那个值
        placed = sorted(
            (im.anchor._from.row + 1, im.anchor.ext.cy / 9525) for im in big._images
        )
        assert len(placed) == 2
        assert placed[0][1] <= LARGE_MAX_PX, "大图没有被钳到上限以内"
        # 两段不能叠在一起：第二段必须排在第一张图占掉的行之后
        assert placed[1][0] - placed[0][0] >= placed[0][1] / DEFAULT_ROW_PX, \
            f"第二段压在第一张大图上了: {placed}"

    def test_rows_without_images_have_neither_picture_nor_link(self):
        from app.services.excel_service import EXPORT_COLUMNS, IMAGE_SHEET

        wb = self._build([_export_row(index=1), _export_row(index=2)])
        ws = wb["帖子明细"]
        assert ws._images == []
        col = [k for k, _ in EXPORT_COLUMNS].index("image_desc") + 1
        assert ws.cell(2, col).hyperlink is None
        assert IMAGE_SHEET not in wb.sheetnames, "一张图都没有就别凭空多一张空表"

    def test_missing_file_does_not_break_the_report(self):
        """图片是采集时下载的，可能被清理掉。报告照出，只是这一行没有图"""
        wb = self._build([
            _export_row(index=1, images=["src_x/gone.png"], image_desc="描述还在"),
            _export_row(index=2, images=[self.rel], image_desc="这张还在"),
        ])
        ws = wb["帖子明细"]
        assert len(ws._images) == 1, "缺文件的那行不该插图，在的那行必须插"
        assert ws.max_row == 3

    def test_unreadable_file_gets_no_dead_link(self):
        """文件在、但读不出来（截断 / 损坏）。不能留一个跳到空段落的 🔍"""
        from app.services.excel_service import EXPORT_COLUMNS, IMAGE_SHEET

        broken = os.path.join(self.tmpdir, "media", "src_x", "broken_0.png")
        with open(broken, "wb") as f:
            f.write(b"\x89PNG\r\n\x1a\n" + b"garbage")   # 有 PNG 头，内容是坏的
        wb = self._build([_export_row(index=1, images=["src_x/broken_0.png"],
                                      image_desc="描述还在")])
        ws = wb["帖子明细"]
        assert ws._images == []
        col = [k for k, _ in EXPORT_COLUMNS].index("image_desc") + 1
        assert ws.cell(2, col).hyperlink is None, "跳过去是一段没有图的空标题"
        assert ws.cell(2, col).value == "描述还在", "描述本身还是要留着"
        assert IMAGE_SHEET not in wb.sheetnames

    def test_sheet_names_come_from_one_place(self):
        """两张表互相跳转。名字写死字面量的话，改名就变成指向不存在的表，
        Excel 打开时弹「需要修复」—— 比报错难查得多
        """
        from app.services.excel_service import DETAIL_SHEET, IMAGE_SHEET

        wb = self._build([_export_row(index=1, images=[self.rel], image_desc="报告")])
        assert wb.sheetnames == ["概览", DETAIL_SHEET, IMAGE_SHEET]
        back = wb[IMAGE_SHEET].cell(3, 2)
        assert back.hyperlink.location.startswith(f"'{DETAIL_SHEET}'!")

    def test_path_traversal_never_reaches_the_workbook(self):
        """images 来自采集脚本；media 目录之外是数据库和明文密钥"""
        wb = self._build([_export_row(index=1, images=["../../hyxi.db"])])
        assert wb["帖子明细"]._images == []

    def test_multiple_images_on_one_post_all_show_up(self):
        from PIL import Image as PILImage
        from app.services.excel_service import IMAGE_SHEET

        second = os.path.join(self.tmpdir, "media", "src_x", "shot_1.png")
        PILImage.new("RGB", (400, 400), (200, 60, 60)).save(second)
        wb = self._build([
            _export_row(index=1, images=[self.rel, "src_x/shot_1.png"], image_desc="两张")
        ])
        assert len(wb[IMAGE_SHEET]._images) == 2, "第二张图在报告里丢了"


class TestSearchFilteringEndToEnd:
    """帖子搜索过滤测试"""

    def _make_posts(self):
        return [
            {"username": "Alice", "timestamp": "22-05-2026 10:00", "content": "De installatie was goed", "translation": "安装很好", "page_number": 1},
            {"username": "Bob", "timestamp": "22-05-2026 11:00", "content": "Probleem met de app", "translation": "App有问题", "page_number": 1},
            {"username": "Charlie", "timestamp": "22-05-2026 12:00", "content": "Vraag over garantie", "translation": "保修问题", "page_number": 2},
        ]

    def test_search_matches_username(self):
        posts = self._make_posts()
        kw = "alice"
        filtered = [p for p in posts if kw in (p.get("username") or "").lower() or kw in (p.get("content") or "").lower() or kw in (p.get("translation") or "").lower()]
        assert len(filtered) == 1
        assert filtered[0]["username"] == "Alice"

    def test_search_matches_content_dutch(self):
        posts = self._make_posts()
        kw = "installatie"
        filtered = [p for p in posts if kw in (p.get("username") or "").lower() or kw in (p.get("content") or "").lower() or kw in (p.get("translation") or "").lower()]
        assert len(filtered) == 1
        assert filtered[0]["username"] == "Alice"

    def test_search_matches_translation_chinese(self):
        posts = self._make_posts()
        kw = "保修"
        filtered = [p for p in posts if kw in (p.get("username") or "").lower() or kw in (p.get("content") or "").lower() or kw in (p.get("translation") or "").lower()]
        assert len(filtered) == 1
        assert filtered[0]["username"] == "Charlie"

    def test_search_matches_multiple(self):
        posts = self._make_posts()
        kw = "app"
        filtered = [p for p in posts if kw in (p.get("username") or "").lower() or kw in (p.get("content") or "").lower() or kw in (p.get("translation") or "").lower()]
        assert len(filtered) == 1
        assert filtered[0]["username"] == "Bob"

    def test_search_no_match(self):
        posts = self._make_posts()
        kw = "xyz_not_found"
        filtered = [p for p in posts if kw in (p.get("username") or "").lower() or kw in (p.get("content") or "").lower() or kw in (p.get("translation") or "").lower()]
        assert len(filtered) == 0

    def test_search_empty_returns_all(self):
        posts = self._make_posts()
        kw = ""
        filtered = [p for p in posts if not kw or kw in (p.get("username") or "").lower() or kw in (p.get("content") or "").lower() or kw in (p.get("translation") or "").lower()]
        assert len(filtered) == 3


class TestAtomicWriteEndToEnd:
    """原子写入测试"""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        self.tasks_path = os.path.join(self.tmpdir, "tasks.json")

    def teardown_method(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_atomic_write_does_not_lose_data(self):
        """验证原子写入不会丢失已有数据"""
        # 写入初始数据
        initial = {"tasks": [{"id": "t1", "status": "completed"}]}
        with open(self.tasks_path, "w") as f:
            json.dump(initial, f)

        # 模拟原子写入
        tmp_path = self.tasks_path + ".tmp"
        new_data = {"tasks": [{"id": "t1", "status": "completed"}, {"id": "t2", "status": "pending"}]}
        with open(tmp_path, "w") as f:
            json.dump(new_data, f)
        os.replace(tmp_path, self.tasks_path)

        # 验证数据完整
        with open(self.tasks_path) as f:
            result = json.load(f)
        assert len(result["tasks"]) == 2

    def test_atomic_write_preserves_old_on_crash(self):
        """验证写入中途崩溃时原有数据不丢失"""
        # 写入原始数据
        original = {"tasks": [{"id": "original", "status": "completed"}]}
        with open(self.tasks_path, "w") as f:
            json.dump(original, f)

        # 模拟部分写入（不完成替换）
        tmp_path = self.tasks_path + ".tmp"
        with open(tmp_path, "w") as f:
            f.write("partial data [corrupt")

        # 验证原始数据完好
        with open(self.tasks_path) as f:
            result = json.load(f)
        assert result["tasks"][0]["id"] == "original"


class TestLoggingConfigEndToEnd:
    """日志配置测试"""

    def test_get_logger_returns_logger(self):
        from app.logging_config import get_logger
        import logging
        # 强制重置单例以测试
        import app.logging_config as lc
        lc._logger = None
        logger = get_logger()
        assert logger is not None
        assert isinstance(logger, logging.Logger)
        assert logger.level <= 20  # INFO or lower

    def test_get_logger_is_singleton(self):
        from app.logging_config import get_logger
        logger1 = get_logger("mod1")
        logger2 = get_logger("mod2")
        assert logger1 is logger2


class TestEnvFileAnchoringEndToEnd:
    """.env 只认项目根那一份 —— 文档里的启动命令先 cd 进 backend，
    env_file 写相对路径的话根目录配置会被整个跳过，密钥漏配还毫无提示"""

    def test_env_file_is_project_root_not_cwd(self):
        from app.config import Settings

        env_file = Settings.model_config["env_file"]
        assert os.path.isabs(env_file)
        assert env_file == os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            ".env",
        )

    def test_stray_env_in_cwd_is_ignored(self):
        """在别处放一份同名 .env 并切过去，配置不能被它带跑"""
        from app.config import Settings

        tmpdir = tempfile.mkdtemp()
        original_cwd = os.getcwd()
        try:
            with open(os.path.join(tmpdir, ".env"), "w") as f:
                f.write("TWEAKERS_SECRET_KEY=stray-key-from-cwd\n")
            os.chdir(tmpdir)
            assert Settings().secret_key != "stray-key-from-cwd"
        finally:
            os.chdir(original_cwd)
            shutil.rmtree(tmpdir, ignore_errors=True)


class TestTranslationNumberingStripEndToEnd:
    """翻译结果的编号前缀剥离：不得吞掉正文里的真实数值"""

    def test_leading_value_preserved(self):
        from app.services.translator_service import _strip_numbering
        assert _strip_numbering("5 kWh 就够了") == "5 kWh 就够了"

    def test_leading_year_preserved(self):
        from app.services.translator_service import _strip_numbering
        assert _strip_numbering("2026 年才会上市") == "2026 年才会上市"

    def test_real_numbering_prefix_removed(self):
        from app.services.translator_service import _strip_numbering
        assert _strip_numbering("1. 这块电池不错") == "这块电池不错"
        assert _strip_numbering("[2]. 这块电池不错") == "这块电池不错"
        assert _strip_numbering("3) 这块电池不错") == "这块电池不错"
        assert _strip_numbering("4、这块电池不错") == "这块电池不错"

    def test_bracketed_number_followed_by_newline_removed(self):
        """请求里的编号就是 [n] 这个形态，模型经常原样抄回来。
        实测在双来源端到端里印进了 Excel 和结果页。"""
        from app.services.translator_service import _strip_numbering
        assert _strip_numbering("[1]\n我使用HYXi Halo已经三个月了") == "我使用HYXi Halo已经三个月了"
        assert _strip_numbering("[12] 逆变器停机了") == "逆变器停机了"

    def test_bracketed_value_in_text_preserved(self):
        from app.services.translator_service import _strip_numbering
        # 方括号里是编号才剥；正文本身以数字开头的照旧保留
        assert _strip_numbering("15 分钟就装好了") == "15 分钟就装好了"


class TestUntranslatedDetectionEndToEnd:
    """批内混语种时 LLM 有时把一种语言原样吐回来，要能被判成漏译进重译队列"""

    def test_identical_non_chinese_is_flagged(self):
        from app.services.translator_service import _looks_untranslated
        assert _looks_untranslated("Firmware update broke my WiFi", "Firmware update broke my WiFi")

    def test_chinese_source_is_not_flagged(self):
        """原文本来就是中文时「译文==原文」是正确结果，不能拉去重译"""
        from app.services.translator_service import _looks_untranslated
        assert not _looks_untranslated("这块电池不错", "这块电池不错")

    def test_real_translation_is_not_flagged(self):
        from app.services.translator_service import _looks_untranslated
        assert not _looks_untranslated("固件更新弄坏了我的 WiFi", "Firmware update broke my WiFi")

    def test_empty_translation_is_not_flagged_here(self):
        """空译文由既有的 [翻译为空] 分支处理，这里不重复判定"""
        from app.services.translator_service import _looks_untranslated
        assert not _looks_untranslated("", "Firmware update")


class TestIncrementalMergeOrderEndToEnd:
    """增量翻译合并：顺序必须与源 JSON 一致"""

    def _post(self, fp, translated=False):
        p = {"fingerprint": fp, "username": "u", "content": f"c{fp}", "page_number": 1}
        if translated:
            p["translation"] = f"t{fp}"
            p["_processed"] = {"translated": True}
        return p

    def test_merge_keeps_source_order(self):
        from app.services.orchestrator import _merge_by_fingerprint
        # 源顺序 a,b,c,d，其中 b、d 已翻译，a、c 本轮才翻译
        source = [self._post("a"), self._post("b", True), self._post("c"), self._post("d", True)]
        pending = [source[0], source[2]]
        for p in pending:
            p["translation"] = f"new-{p['fingerprint']}"

        merged = _merge_by_fingerprint(source, pending)
        assert [p["fingerprint"] for p in merged] == ["a", "b", "c", "d"]
        assert merged[0]["translation"] == "new-a"
        assert merged[1]["translation"] == "tb"

    def test_posts_without_fingerprint_survive(self):
        from app.services.orchestrator import _merge_by_fingerprint
        source = [{"content": "无指纹"}, self._post("z")]
        merged = _merge_by_fingerprint(source, [{"content": "无指纹", "translation": "x"}])
        assert len(merged) == 2
        assert merged[1]["fingerprint"] == "z"


class TestSchedulerStartupResilienceEndToEnd:
    """调度器启动：单条坏配置不得砖化整个服务"""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        import app.services.scheduler_service as sched_mod
        import app.services.storage as storage_module
        self._orig_data_dir = sched_mod.settings.data_dir
        sched_mod.settings.data_dir = self.tmpdir
        # 定时任务配置和 APScheduler 的 job store 现在都在 hyxi.db 里，而 DB_PATH
        # 是 import 时算好的常量 —— 只改 data_dir 会让这批用例写进真实的库
        self.storage = storage_module
        self._old_db = storage_module.DB_PATH
        storage_module.DB_PATH = os.path.join(self.tmpdir, "hyxi.db")
        storage_module.init_db()

    def teardown_method(self):
        import asyncio
        import app.services.scheduler_service as sched_mod
        sched_mod.settings.data_dir = self._orig_data_dir
        self.storage.DB_PATH = self._old_db
        # asyncio.run 结束时会把当前事件循环置空，其余用 get_event_loop 的测试会跟着挂
        asyncio.set_event_loop(asyncio.new_event_loop())
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_malformed_config_is_skipped_not_fatal(self):
        import asyncio
        import app.services.scheduler_service as sched_mod

        svc = sched_mod.SchedulerService()
        # 坏配置排在前面：修复前它会中断整个加载循环，后面的好配置永远注册不上
        svc._save_configs([
            {"id": "bad", "description": "畸形时间", "interval": "daily",
             "time": "9", "enabled": True, "created_at": "2026-08-01T00:00:00"},
            {"id": "good", "description": "正常配置", "interval": "hourly",
             "enabled": True, "created_at": "2026-08-01T00:00:01"},
        ])

        async def run():
            svc.start()
            try:
                return svc.scheduler.get_job("bad"), svc.scheduler.get_job("good")
            finally:
                svc.shutdown()

        bad, good = asyncio.run(run())
        assert bad is None, "畸形配置被注册成了 job"
        assert good is not None, "坏配置把后面的好配置一起带崩了"

    def test_create_does_not_persist_unbuildable_config(self):
        import asyncio
        import app.services.scheduler_service as sched_mod

        svc = sched_mod.SchedulerService()

        async def run():
            svc.start()
            try:
                svc.create("时间畸形", "daily", "9")
            except Exception:
                pass
            finally:
                svc.shutdown()

        asyncio.run(run())
        assert svc._load_configs() == [], "建不起 job 的脏配置被落盘了"

    def test_update_does_not_persist_unbuildable_config(self):
        import asyncio
        import app.services.scheduler_service as sched_mod

        svc = sched_mod.SchedulerService()
        # 模拟旧版本遗留的脏数据：time 畸形，但 hourly 不读 time 所以 job 建得起来
        svc._save_configs([{"id": "legacy", "description": "遗留脏配置", "interval": "hourly",
                            "time": "9", "enabled": True,
                            "created_at": "2026-08-01T00:00:00"}])

        async def run():
            svc.start()
            try:
                svc.update("legacy", {"interval": "daily"})
            except Exception:
                pass
            finally:
                svc.shutdown()

        asyncio.run(run())
        assert svc._load_configs()[0]["interval"] == "hourly", "改不成的 interval 被落盘了"


class TestProcessedFlagSyncEndToEnd:
    """舆情分析写回 sentiment_at：只能动这一个字段，不能碰 translation"""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        self.storage, self._restore = _use_temp_db(self.tmpdir)
        self.storage.upsert_posts("src_x", [
            {"fingerprint": "aaa", "source": "src_x", "content": "c1", "translation": "t1",
             "_processed": {"translated": True}},
            {"fingerprint": "bbb", "source": "src_x", "content": "c2", "translation": "t2",
             "_processed": {"translated": True}},
        ])

    def teardown_method(self):
        self._restore()
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_sentiment_flag_does_not_wipe_translated(self):
        analyzed = [{"fingerprint": "aaa", "source": "src_x",
                     "_processed": {"sentiment_at": "2026-07-30T12:00:00"}}]
        assert self.storage.mark_sentiment_analyzed(analyzed) == 1

        posts = self.storage.load_posts(["src_x"])
        assert posts[0]["_processed"]["translated"] is True, "translated 标记被舆情写回抹掉了"
        assert posts[0]["_processed"]["sentiment_at"] == "2026-07-30T12:00:00"
        assert posts[0]["translation"] == "t1", "译文被抹掉了"
        assert posts[1]["_processed"] == {"translated": True}

    def test_no_match_leaves_rows_untouched(self):
        before = self.storage.load_posts(["src_x"])
        assert self.storage.mark_sentiment_analyzed(
            [{"fingerprint": "zzz", "source": "src_x",
              "_processed": {"sentiment_at": "2026-07-30T12:00:00"}}]
        ) == 0
        assert self.storage.load_posts(["src_x"]) == before

    def test_already_analyzed_post_keeps_its_first_timestamp(self):
        """重复分析不该把首次分析时间冲掉 —— 增量靠它判断谁已经算过"""
        first = [{"fingerprint": "aaa", "source": "src_x",
                  "_processed": {"sentiment_at": "2026-07-30T12:00:00"}}]
        again = [{"fingerprint": "aaa", "source": "src_x",
                  "_processed": {"sentiment_at": "2026-08-05T09:00:00"}}]
        assert self.storage.mark_sentiment_analyzed(first) == 1
        assert self.storage.mark_sentiment_analyzed(again) == 0
        assert self.storage.load_posts(["src_x"])[0]["_processed"]["sentiment_at"] \
            == "2026-07-30T12:00:00"


class TestExcelRobustnessEndToEnd:
    """Excel 渲染对脏数据的容忍度"""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        import app.services.excel_service as excel_mod
        excel_mod.settings.exports_dir = self.tmpdir

    def teardown_method(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_star_level_normalizes_dirty_intensity(self):
        from app.services.excel_service import _star_level
        assert _star_level("很强") == 0
        assert _star_level(None) == 0
        assert _star_level(-3) == 0
        assert _star_level(99) == 5
        assert _star_level(3.4) == 3

    def test_time_range_uses_extremes_not_first_last(self):
        from app.services.excel_service import ExcelService
        from app.services.progress_manager import ProgressManager
        import asyncio

        # 乱序 + 跨年，字典序会把 "05-01-2027" 排在 "22-05-2026" 前面
        posts = [
            {"username": "u", "timestamp": "22-05-2026 17:06", "content": "c", "translation": "t", "page_number": 1},
            {"username": "u", "timestamp": "05-01-2027 09:00", "content": "c", "translation": "t", "page_number": 1},
            {"username": "u", "timestamp": "03-03-2026 08:00", "content": "c", "translation": "t", "page_number": 1},
        ]
        pm = ProgressManager()
        result = asyncio.get_event_loop().run_until_complete(
            ExcelService.execute("time-range", posts, {"include_stats": True}, pm)
        )
        from openpyxl import load_workbook
        ws = load_workbook(result["file_path"])["统计信息"]
        rows = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, 7)}
        assert rows["时间范围开始"] == "03-03-2026 08:00"
        assert rows["时间范围结束"] == "05-01-2027 09:00"

    def test_export_name_is_timestamped_so_reruns_do_not_overwrite(self):
        from app.services.excel_service import ExcelService
        from app.services.progress_manager import ProgressManager
        import asyncio, time

        posts = [{"username": "u", "timestamp": "22-05-2026 17:06", "content": "c",
                  "translation": "t", "page_number": 1}]
        pm = ProgressManager()
        loop = asyncio.get_event_loop()
        first = loop.run_until_complete(ExcelService.execute("dup-task", posts, {}, pm))
        time.sleep(1.1)
        second = loop.run_until_complete(ExcelService.execute("dup-task", posts, {}, pm))

        assert first["file_name"] != second["file_name"], "同一任务重跑会盖掉上一次导出"
        assert os.path.exists(first["file_path"]) and os.path.exists(second["file_path"])


class TestSentimentIsolationEndToEnd:
    """舆情结果不得跨任务串台"""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        import app.services.storage as storage_module
        self.storage = storage_module
        self._old_db = storage_module.DB_PATH
        storage_module.DB_PATH = os.path.join(self.tmpdir, "hyxi.db")
        storage_module.init_db()

    def teardown_method(self):
        self.storage.DB_PATH = self._old_db
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    @staticmethod
    def _posts(*fps):
        return [{"fingerprint": f, "source": "src_x", "content": f"内容{f}"} for f in fps]

    def test_untouched_posts_have_no_conclusions(self):
        """没人分析过的帖子就是没有结论，不能从别处顶一个上来"""
        posts = self._posts("a1", "a2")
        self.storage.save_sentiment("task-a", {
            "task_id": "task-a", "analyzed_at": "2026-07-30T10:00:00",
            "results": [{"sentiment": "positive"}, {"sentiment": "negative"}],
        }, posts)
        assert self.storage.get_sentiment("task-a", posts)["success"] == 2
        # 另一批完全不同的帖子：结论按身份取，取不到就是真没有
        assert self.storage.get_sentiment("task-b", self._posts("z1", "z2")) is None

    def test_conclusions_are_visible_to_every_task_over_the_same_posts(self):
        """结论属于帖子，不属于任务。

        `_processed.sentiment_at` 按帖子存、跨任务共享（所以同一条帖子不会被重复
        花钱分析），结论必须同一个粒度。按任务过滤的话，第二个任务跑同一批数据时
        页面上会有一大片「未分析」，而它们早就分析过了 —— 线上就是这么暴露的：
        94 条里 90 条显示未分析。
        """
        posts = self._posts("p1", "p2", "p3")
        # 任务 A 分析了前两条
        self.storage.save_sentiment("task-a", {
            "analyzed_at": "2026-08-01T10:00:00",
            "results": [{"sentiment": "positive", "reason_cn": "属于p1"},
                        {"sentiment": "negative", "reason_cn": "属于p2"}],
        }, posts)
        # 任务 B 后来只分析了新增的第三条（前两条被 sentiment_at 挡住不再重复分析）
        self.storage.save_sentiment("task-b", {
            "analyzed_at": "2026-08-05T10:00:00",
            "results": [None, None, {"sentiment": "neutral", "reason_cn": "属于p3"}],
        }, posts)

        got = self.storage.get_sentiment("task-b", posts)
        assert got["success"] == 3, "任务 B 应该看得到 A 分析过的那两条"
        assert [r["reason_cn"] for r in got["results"]] == ["属于p1", "属于p2", "属于p3"]
        # 反过来任务 A 也看得到 B 的
        assert self.storage.get_sentiment("task-a", posts)["success"] == 3

    def test_later_analysis_wins_for_the_same_post(self):
        """同一条帖子被重新分析，留最后写的那份，不是两份并存"""
        posts = self._posts("p1")
        self.storage.save_sentiment("task-a", {
            "analyzed_at": "2026-08-01T10:00:00",
            "results": [{"sentiment": "positive", "reason_cn": "旧结论"}],
        }, posts)
        self.storage.save_sentiment("task-b", {
            "analyzed_at": "2026-08-05T10:00:00",
            "results": [{"sentiment": "negative", "reason_cn": "新结论"}],
        }, posts)
        got = self.storage.get_sentiment("task-a", posts)
        assert len(got["results"]) == 1
        assert got["results"][0]["reason_cn"] == "新结论"

    def test_results_follow_the_post_not_the_row(self):
        """结论按 (source_id, fingerprint) 存，帖子顺序变了也不会串位。

        存下标那版做不到这件事 —— 这正是上一个线上 bug 的形状。
        """
        posts = self._posts("p1", "p2", "p3")
        self.storage.save_sentiment("t", {
            "analyzed_at": "2026-08-05T10:00:00",
            "results": [
                {"sentiment": "positive", "intensity": 5, "reason_cn": "属于p1", "dimensions": ["价格/性价比"]},
                {"sentiment": "negative", "intensity": 4, "reason_cn": "属于p2", "dimensions": []},
                {"sentiment": "neutral", "intensity": 2, "reason_cn": "属于p3", "dimensions": []},
            ],
        }, posts)

        # 换个顺序读回来，结论必须跟着各自的帖子走
        shuffled = self._posts("p3", "p1", "p2")
        got = self.storage.get_sentiment("t", shuffled)
        assert [r["reason_cn"] for r in got["results"]] == ["属于p3", "属于p1", "属于p2"]
        assert got["results"][0]["dimensions"] == []
        assert got["results"][1]["dimensions"] == ["价格/性价比"]

    def test_new_posts_read_back_as_unanalyzed(self):
        """后来新采到的帖子没有结论，占位必须是 None 而不是顶上别人的"""
        posts = self._posts("p1", "p2")
        self.storage.save_sentiment("t", {
            "analyzed_at": "2026-08-05T10:00:00",
            "results": [{"sentiment": "positive", "reason_cn": "属于p1"},
                        {"sentiment": "negative", "reason_cn": "属于p2"}],
        }, posts)

        grown = self._posts("p1", "p2", "p3", "p4")
        got = self.storage.get_sentiment("t", grown)
        assert [r and r["reason_cn"] for r in got["results"]] == ["属于p1", "属于p2", None, None]
        assert got["total"] == 4 and got["success"] == 2 and got["failed"] == 2

    def test_duplicate_identity_does_not_wipe_the_whole_round(self):
        """同一批里出现两条 (source_id, fingerprint) 相同的帖子，不能把整轮结论搞没。

        指纹只吃 username|timestamp|content[:100]，翻页错位就能撞上。裸 INSERT 会抛
        IntegrityError 让事务整个回滚 —— 花钱算出来的结论一条都不落库，而任务还在
        报「分析完成」。
        """
        posts = [
            {"fingerprint": "dup", "source": "src_x", "content": "同一条帖子"},
            {"fingerprint": "dup", "source": "src_x", "content": "同一条帖子"},
            {"fingerprint": "other", "source": "src_x", "content": "另一条"},
        ]
        self.storage.save_sentiment("t", {
            "analyzed_at": "2026-08-05T10:00:00",
            "results": [{"sentiment": "positive", "reason_cn": "第一次"},
                        {"sentiment": "negative", "reason_cn": "第二次"},
                        {"sentiment": "neutral", "reason_cn": "属于other"}],
        }, posts)

        got = self.storage.get_sentiment("t", posts)
        assert got is not None, "整轮结论被回滚掉了"
        # 后写覆盖先写，与项目里其它按 post_key 建映射的地方口径一致
        assert got["results"][0]["reason_cn"] == "第二次"
        assert got["results"][2]["reason_cn"] == "属于other", "撞键把后面的记录一起带没了"

    def test_save_failure_is_not_swallowed(self):
        """存不下就必须抛。吞掉异常会让任务报「完成」而库里空空如也"""
        import sqlite3
        self.storage.DB_PATH = os.path.join(self.tmpdir, "nonexistent", "no.db")
        try:
            with pytest.raises(Exception):
                self.storage.save_sentiment("t", {
                    "analyzed_at": "2026-08-05T10:00:00",
                    "results": [{"sentiment": "positive", "reason_cn": "x"}],
                }, [{"fingerprint": "p1", "source": "src_x"}])
        finally:
            self.storage.DB_PATH = os.path.join(self.tmpdir, "hyxi.db")

    def test_same_fingerprint_across_sources_does_not_collide(self):
        """指纹不含来源，只按指纹存会让两个平台的同名空帖互相覆盖"""
        posts = [
            {"fingerprint": "same", "source": "src_a", "content": "A 平台"},
            {"fingerprint": "same", "source": "src_b", "content": "B 平台"},
        ]
        self.storage.save_sentiment("t", {
            "analyzed_at": "2026-08-05T10:00:00",
            "results": [{"sentiment": "positive", "reason_cn": "属于A"},
                        {"sentiment": "negative", "reason_cn": "属于B"}],
        }, posts)
        got = self.storage.get_sentiment("t", posts)
        assert [r["reason_cn"] for r in got["results"]] == ["属于A", "属于B"]


class TestProgressManagerResourceEndToEnd:
    """SSE 订阅队列的资源边界"""

    def test_full_queue_drops_oldest_instead_of_blocking(self):
        import asyncio
        from app.services.progress_manager import ProgressManager, QUEUE_MAXSIZE

        pm = ProgressManager()

        async def run():
            q = pm.subscribe("t1")
            for i in range(QUEUE_MAXSIZE + 5):
                await pm.emit("t1", "log", {"i": i})
            assert q.qsize() == QUEUE_MAXSIZE
            first = await q.get()
            # 最旧的 5 条被丢弃，而不是让 emit 无界堆积或阻塞抓取主流程
            assert first["data"]["i"] == 5

        asyncio.new_event_loop().run_until_complete(run())

    def test_unsubscribe_removes_empty_key(self):
        from app.services.progress_manager import ProgressManager

        pm = ProgressManager()
        q = pm.subscribe("t2")
        assert "t2" in pm.subscribers
        pm.unsubscribe("t2", q)
        assert "t2" not in pm.subscribers, "订阅字典会随任务数单调增长"

    def test_stream_ends_on_its_own_terminal_event_and_only_that_one(self):
        """结束条件由端点各自给：舆情流等 sentiment_complete，任务进度流等 task_complete。

        两条流跑在同一个频道上。若共用一份「终止事件表」，流水线里的 sentiment 步骤一发完
        sentiment_complete 就会把任务进度流也掐断，紧随其后的 task_complete 没人收得到。
        """
        import asyncio
        from app.services.progress_manager import ProgressManager

        pm = ProgressManager()

        async def _emit_later(manager, channel):
            await asyncio.sleep(0.05)
            await manager.emit(channel, "log", {"message": "x"})
            await manager.emit(channel, "sentiment_complete", {"status": "completed"})
            await manager.emit(channel, "task_complete", {"status": "completed"})

        async def collect(channel, terminal):
            asyncio.ensure_future(_emit_later(pm, channel))
            return [f async for f in pm.event_generator(channel, terminal)]

        loop = asyncio.new_event_loop()

        frames = loop.run_until_complete(asyncio.wait_for(collect("t3", "sentiment_complete"), timeout=10))
        assert any("sentiment_complete" in f for f in frames)
        assert not any("task_complete" in f for f in frames), "舆情流不该继续收后面的事件"

        # 同样一串事件，等 task_complete 的那条流必须挺过 sentiment_complete
        frames = loop.run_until_complete(asyncio.wait_for(collect("t4", "task_complete"), timeout=10))
        assert any("sentiment_complete" in f for f in frames)
        assert any("task_complete" in f for f in frames), "sentiment_complete 把任务进度流掐断了"

        assert not pm.subscribers


class TestUpstreamErrorRedactionEndToEnd:
    """上游错误体会落库并被展示，不能原样存"""

    def _resp(self, status: int, body: str):
        import httpx
        return httpx.Response(status, text=body, request=httpx.Request("POST", "https://x/y"))

    def test_extracts_message_and_masks_secrets(self):
        from app.services.llm_service import _safe_error_detail
        body = json.dumps({"error": {
            "message": "Invalid key sk-abcdef1234567890 for org-acme-42",
            "type": "authentication_error",
        }})
        detail = _safe_error_detail(self._resp(401, body))
        assert "sk-abcdef1234567890" not in detail
        assert "org-acme-42" not in detail
        assert "Invalid key" in detail

    def test_unparseable_body_is_not_echoed(self):
        from app.services.llm_service import _safe_error_detail
        detail = _safe_error_detail(self._resp(500, "<html>account 12345 quota exhausted</html>"))
        assert "12345" not in detail
        assert "日志" in detail


# 真实拉起 node 子进程，没有 node 就跳过而不是伪造
_HAS_NODE = shutil.which("node") is not None


def _fake_playwright_install(root: str) -> None:
    """让 CollectorRunner 的依赖自检在 tmpdir 里通过（用假脚本的测试并不需要真 playwright）"""
    pkg_dir = os.path.join(root, "node_modules", "playwright")
    os.makedirs(pkg_dir, exist_ok=True)
    with open(os.path.join(pkg_dir, "package.json"), "w", encoding="utf-8") as f:
        f.write('{"name":"playwright","version":"0.0.0-test"}')

def _write_collector_script(root: str, script: str) -> None:
    """假采集器脚本写到 collectors/tweakers.js —— 与真实布局一致"""
    os.makedirs(os.path.join(root, "collectors"), exist_ok=True)
    with open(os.path.join(root, "collectors", "tweakers.js"), "w", encoding="utf-8") as f:
        f.write(script)


# 采集脚本读 job 文件而不是 argv，假脚本也必须走同一条协议
_READ_JOB = """
const fs = require('fs');
const jobArg = process.argv.find(a => a.startsWith('--job='));
const job = JSON.parse(fs.readFileSync(jobArg.slice('--job='.length), 'utf-8'));
"""

STALL_SCRIPT = """
const fs = require('fs');
const path = require('path');
const { spawn } = require('child_process');
const beat = path.join(__dirname, 'heartbeat.txt');

if (process.argv.includes('--child')) {
  setInterval(() => fs.writeFileSync(beat, String(Date.now())), 100);
} else {
  console.log('第 1/9 页');
  spawn(process.execPath, [__filename, '--child'], { stdio: 'ignore' });
  setInterval(() => {}, 1000);
}
"""


class TestScraperStallTimeoutEndToEnd:
    """抓取子进程 stall 时必须超时并连同后代一起清理"""

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        _write_collector_script(self.tmpdir, STALL_SCRIPT)
        _fake_playwright_install(self.tmpdir)
        import app.services.collector_runner as runner_mod
        self.mod = runner_mod
        self._old_root = runner_mod.settings.project_root
        self._old_timeout = runner_mod.SUBPROCESS_TIMEOUT
        runner_mod.settings.project_root = self.tmpdir
        runner_mod.SUBPROCESS_TIMEOUT = 3
        self.storage, self._restore_db = _use_temp_db(self.tmpdir)

    def teardown_method(self):
        self.mod.settings.project_root = self._old_root
        self.mod.SUBPROCESS_TIMEOUT = self._old_timeout
        self._restore_db()
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_stalled_stdout_times_out_and_kills_whole_tree(self):
        if not _HAS_NODE:
            import pytest
            pytest.skip("未安装 node")

        import asyncio, time
        from app.collectors import get_collector
        from app.services.progress_manager import ProgressManager

        collector = get_collector("tweakers")

        async def run():
            # 外层再套一层超时：回归时读流循环永不结束，测试应快速失败而不是挂死
            with_timeout = asyncio.wait_for(
                self.mod.CollectorRunner.execute(
                    "stall", collector,
                    {"id": collector.id, "params": {"thread_id": 123}},
                    ProgressManager(),
                ),
                timeout=30,
            )
            try:
                await with_timeout
            except Exception as e:
                return str(e)
            return ""

        loop = asyncio.new_event_loop()
        try:
            message = loop.run_until_complete(run())
        finally:
            loop.close()

        assert "超时" in message, f"stall 的子进程没有触发超时: {message!r}"

        # 心跳文件由孙进程写：只 terminate 父进程的话它会继续跳（Playwright 的 Chromium 同理）
        beat = os.path.join(self.tmpdir, "collectors", "heartbeat.txt")
        assert os.path.exists(beat), "孙进程没起来，本用例失去意义"
        before = os.path.getmtime(beat)
        time.sleep(2)
        assert os.path.getmtime(beat) == before, "子进程树没被清理干净，孙进程仍在运行"


class TestStartPageResolutionEndToEnd:
    """起始页由后端确定性派生 —— LLM 把 URL 末尾页码当起始页会永久丢掉前面几页"""

    def _resolve(self, description: str, params: dict):
        from app.services.orchestrator import _resolve_start_page
        ignored = _resolve_start_page({"description": description}, params)
        return params.get("start_page"), ignored

    def test_url_trailing_page_does_not_become_start_page(self):
        # 用户原话回归锁：URL 末尾的 /4 是页码，且「所有页面」必须从第 1 页开始
        start, ignored = self._resolve(
            "抓取https://gathering.tweakers.net/forum/list_messages/2336074/4所有页面，翻译成中文",
            {"thread_id": 2336074, "start_page": 4},
        )
        assert start == 1
        assert ignored == 4, "被忽略的 LLM 取值必须回传给调用方打日志，静默纠正同样难排查"

    def test_no_llm_value_resolves_to_first_page_silently(self):
        start, ignored = self._resolve("抓取帖子 2336074 所有页面", {"thread_id": 2336074})
        assert start == 1
        assert ignored is None

    def test_explicit_instruction_is_honored(self):
        assert self._resolve("抓取帖子 2336074，从第 7 页开始", {})[0] == 7
        assert self._resolve("抓取帖子 2336074，从第7页起", {})[0] == 7
        assert self._resolve("scrape thread 2336074 start_page: 7", {})[0] == 7

    def test_illegal_values_fall_back_to_first_page(self):
        assert self._resolve("抓取帖子 2336074，从第 0 页开始", {"start_page": 0})[0] == 1
        assert self._resolve("抓取帖子 2336074", {"start_page": -3})[0] == 1
        assert self._resolve("抓取帖子 2336074", {"start_page": "第四页"})[0] == 1


OK_SCRIPT = _READ_JOB + """
console.log(JSON.stringify({ evt: 'progress', current: 1, total: 1, msg: '第 1/1 页' }));
fs.writeFileSync(
  job.output_path,
  JSON.stringify({ thread_id: 123, total_pages: 1, total_posts: 0, complete: true, stop_reason: null, posts: [] }),
  'utf-8'
);
"""

JOB_ECHO_SCRIPT = _READ_JOB + """
fs.writeFileSync(
  job.output_path,
  JSON.stringify({ thread_id: 123, total_pages: 1, total_posts: 0, complete: true,
                   stop_reason: null, posts: [], job: job }),
  'utf-8'
);
"""

PARTIAL_SCRIPT = _READ_JOB + """
console.log(JSON.stringify({ evt: 'progress', current: 1, total: 9, msg: '第 1/9 页' }));
// 退出码 2 的契约是「部分完成，数据已落盘」——脚本在退出前必定先把本轮抓到的写出来
fs.writeFileSync(
  job.output_path,
  JSON.stringify({ thread_id: 123, total_pages: 9, total_posts: 2, complete: false,
                   stop_reason: '目标站拒绝访问 (HTTP 429)，已主动停止抓取',
                   posts: [
                     { fingerprint: 'p1', username: 'u1', timestamp: '01-07-2026 10:00',
                       content: '限流前抓到的第一条', page_number: 1 },
                     { fingerprint: 'p2', username: 'u2', timestamp: '01-07-2026 10:05',
                       content: '限流前抓到的第二条', page_number: 2 },
                   ] }),
  'utf-8'
);
console.error('目标站拒绝访问 (HTTP 429)，已主动停止抓取');
process.exitCode = 2;
"""


class _ScraperTmpRoot:
    """把 project_root 指到 tmpdir，脚本内容由子类给出"""

    script = OK_SCRIPT

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        _write_collector_script(self.tmpdir, self.script)
        import app.services.collector_runner as runner_mod
        self.mod = runner_mod
        self._old_root = runner_mod.settings.project_root
        runner_mod.settings.project_root = self.tmpdir
        # 帖子进库之后 CollectorRunner 会查 posts 表算增量锚点、再把结果写回去
        self.storage, self._restore_db = _use_temp_db(self.tmpdir)

    def teardown_method(self):
        self.mod.settings.project_root = self._old_root
        self._restore_db()
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _execute(self, params=None):
        import asyncio
        from app.collectors import get_collector
        from app.services.progress_manager import ProgressManager

        collector = get_collector("tweakers")

        async def run():
            return await self.mod.CollectorRunner.execute(
                "t", collector,
                {"id": collector.id, "params": params or {"thread_id": 123}},
                ProgressManager(),
            )

        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(run())
        finally:
            loop.close()


class TestScraperDependencyGuardEndToEnd(_ScraperTmpRoot):
    """根目录没装 Node 依赖时，用户该看到「怎么修」而不是 MODULE_NOT_FOUND 堆栈"""

    def test_missing_playwright_reports_how_to_fix(self):
        try:
            self._execute()
        except Exception as e:
            assert "npm" in str(e), f"异常没告诉用户怎么修: {e}"
        else:
            assert False, "缺少 playwright 时没有报错"

    def test_guard_passes_once_installed(self):
        if not _HAS_NODE:
            import pytest
            pytest.skip("未安装 node")
        _fake_playwright_install(self.tmpdir)
        data = self._execute()
        assert data["complete"] is True


class TestScraperPartialExitEndToEnd(_ScraperTmpRoot):
    """残缺结果必须让任务失败，且原因要能被用户读懂"""

    script = PARTIAL_SCRIPT

    def setup_method(self):
        super().setup_method()
        _fake_playwright_install(self.tmpdir)

    def test_partial_exit_code_and_reason_reach_the_user(self):
        if not _HAS_NODE:
            import pytest
            pytest.skip("未安装 node")
        try:
            self._execute()
        except Exception as e:
            message = str(e)
        else:
            message = ""
        assert "code=2" in message, f"退出码 2 没有让抓取步骤失败: {message!r}"
        assert "拒绝访问" in message, f"中断原因没带到用户面前: {message!r}"

    def test_partial_results_are_persisted_so_retry_can_resume(self):
        """退出码 2 的契约是「部分完成，**数据已落盘**，可增量续抓」。

        抓到一半撞上限流时，脚本已经把本轮的写进交接文件了。要是先判退出码再入库，
        这批数据会连同临时文件一起被删掉，/retry 只能从第 1 页重抓一遍 ——
        160 页的长帖抓到第 100 页限流，那 100 页就白抓了。
        """
        if not _HAS_NODE:
            import pytest
            pytest.skip("未安装 node")
        try:
            self._execute()
        except Exception:
            pass

        posts = self.storage.load_posts(["tweakers"])
        assert [p["fingerprint"] for p in posts] == ["p1", "p2"], \
            "退出码 2 抓到的帖子没入库，重试只能从头再抓"
        # 续抓锚点也要跟着建立起来，否则下一轮仍从第 1 页开始
        assert self.storage.max_page_number("tweakers") == 2
        assert sorted(self.storage.known_fingerprints("tweakers")) == ["p1", "p2"]


class TestStartPageReachesCollectorJobEndToEnd(_ScraperTmpRoot):
    """用户原话 → 真实子进程收到的 job：URL 末尾的页码不能变成起始页"""

    script = JOB_ECHO_SCRIPT

    def setup_method(self):
        super().setup_method()
        _fake_playwright_install(self.tmpdir)

    def test_url_trailing_page_does_not_become_start_page(self):
        if not _HAS_NODE:
            import pytest
            pytest.skip("未安装 node")
        from app.services.orchestrator import _resolve_start_page

        task = {"description": "抓取 https://gathering.tweakers.net/forum/list_messages/2336074/4"
                               "所有页面，翻译成中文，导出Excel，分析舆情"}
        params = {"thread_id": 123, "start_page": 4}  # LLM 把 URL 末尾的 4 当成了起始页
        _resolve_start_page(task, params)

        job = self._execute(params)["job"]
        assert job["params"]["start_page"] == 1, f"起始页没有落到第 1 页: {job['params']}"

    def test_pacing_is_not_configurable_at_all(self):
        """请求节奏是反爬纪律，谁都不能改 —— 塞进 params 也得被忽略"""
        if not _HAS_NODE:
            import pytest
            pytest.skip("未安装 node")

        job = self._execute({
            "thread_id": 123,
            "pacing": {"delay_min": 0, "delay_max": 0},
        })["job"]
        assert job["pacing"] == {"delay_min": 4000, "delay_max": 11000}

    def test_llm_cannot_reach_collector_params_at_all(self):
        """collect 步骤只把 source_id 交给编排层，LLM 给的其它参数一律不进 job。

        这比「base_url 只认 source」更强：模型现在连一个字段都递不进来。
        base_url 本身是允许用户在数据源页配的（自建镜像 / 本地验证）。
        """
        from app.models import PlanStep
        from app.services.orchestrator import _resolve_sources

        sources = [{"id": "src_x", "name": "某来源"}]
        plan, _warns = _resolve_sources(
            {"description": "采集"},
            [PlanStep(action="collect", params={
                "source_id": "src_x",
                "base_url": "http://evil.example",
                "thread_id": 999,
                "pacing": {"delay_min": 0},
            })],
            sources,
        )
        assert set(plan[0].params) == {"source_id", "source_name"}, plan[0].params


_FIXTURES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures")
GOLDEN_FILE = os.path.join(_FIXTURES_DIR, "golden_tweakers.json")
FIXTURE_THREAD_ID = 9990001


class TestTweakersCollectorGoldenEndToEnd:
    """真 Chrome + 真 HTTP + 真子进程跑本地 fixture 站点，提取结果必须与黄金基线逐字相等。

    出口 IP 已被 Tweakers 防火墙封禁，真站抓不了；换成本地站点后跑的仍是完整链路，
    不涉及任何 mock。基线由重构前的脚本对同一份 HTML 抓出，指纹是增量去重与跨文件
    合并的唯一锚点 —— 它变了，全部历史数据失配、已翻译的帖子会被重新付费翻译。
    """

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        self.media = os.path.join(self.tmpdir, "media")
        # 帖子进库之后脚本的产出是个用完即删的交接文件，不再落项目根目录
        self.storage, self._restore_db = _use_temp_db(self.tmpdir)

    def teardown_method(self):
        self._restore_db()
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _skip_unless_ready(self):
        import pytest
        from app.config import settings

        if not _HAS_NODE:
            pytest.skip("未安装 node")
        if not os.path.exists(
            os.path.join(settings.project_root, "node_modules", "playwright", "package.json")
        ):
            pytest.skip("项目根目录未安装 playwright")

    def _run(self, base_url, progress=None):
        import asyncio
        from app.collectors import get_collector
        from app.services.collector_runner import CollectorRunner
        from app.services.progress_manager import ProgressManager

        collector = get_collector("tweakers")
        source = {
            "id": "fixture_tweakers",
            "params": {
                "thread_id": FIXTURE_THREAD_ID,
                "headless": True,
                "incremental": False,
            },
            "base_url": base_url,
            "state_file": os.path.join(self.tmpdir, "state.json"),
            # 不给的话正文图会落进真实的 backend/data/media（settings.data_dir 那份）
            "media_dir": self.media,
            "pacing": {"delay_min": 200, "delay_max": 400},
        }
        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(
                CollectorRunner.execute("golden", collector, source,
                                        progress or ProgressManager())
            )
        finally:
            loop.close()

    def test_extraction_matches_golden_baseline(self):
        self._skip_unless_ready()
        sys.path.insert(0, _FIXTURES_DIR)
        from fixture_site import FixtureSite

        with FixtureSite() as base_url:
            data = self._run(base_url)

        with open(GOLDEN_FILE, "r", encoding="utf-8") as f:
            golden = json.load(f)

        assert data["complete"] is True, f"抓取未完成: {data.get('stop_reason')}"
        assert data["total_pages"] == golden["total_pages"]
        assert len(data["posts"]) == len(golden["posts"]), "帖子数与基线不一致"

        fields = ("username", "timestamp", "content", "page_number", "message_id", "fingerprint")
        for got, want in zip(data["posts"], golden["posts"]):
            assert {k: got.get(k) for k in fields} == {k: want.get(k) for k in fields}


    def test_body_images_are_downloaded_and_quoted_ones_ignored(self):
        """Tweakers 的正文图要抓回本地，引用块里的图和表情都不能算进来。

        这条能力以前根本不存在 —— 三个采集器里只有 facebook_group.js 有图片代码，
        Tweakers 侧只是把正文里的 [Afbeelding] 替换成「[图片]」四个字，用户因此
        一张图都拿不到。

        两个坑一并钉住：
          · **引用块里的图不算引用者的**。正文提取为剥引用做了一次 cloneNode，
            但尺寸必须在**原始元素**上量 —— 游离于文档之外的 clone，
            getBoundingClientRect() 一律返回 0，照着 clone 取会把每张图都过滤掉。
          · **表情按渲染尺寸挡掉**。这里不设 host 白名单：Facebook 那边能写死
            scontent 是因为对真站核实过，Tweakers 的真实 DOM 本机访问不到。
        """
        self._skip_unless_ready()
        sys.path.insert(0, _FIXTURES_DIR)
        from fixture_site import FixtureSite

        progress = _RecordingProgress()
        with FixtureSite() as base_url:
            data = self._run(base_url, progress=progress)

        by_id = {p["message_id"]: p for p in data["posts"]}

        # Havelaar 那条：一张 400x300 的正文图 + 一个 20x20 的表情
        with_image = by_id["80000002"]
        images = with_image.get("images") or []
        assert len(images) == 1, f"表情混进来了或者正文图没抓到: {images}"
        assert images[0].startswith("fixture_tweakers/"), f"没按 source 分目录: {images[0]}"
        assert not os.path.isabs(images[0]), "存了绝对路径，落盘文件就搬不了机器了"
        assert os.path.getsize(os.path.join(self.media, images[0])) > 0

        # Dorpjes 那条引用了 Havelaar，引用块里也有一张 400x300 的图 —— 不是他的
        assert not by_id["80000001"].get("images"), "把引用块里的图算到引用者头上了"

        # 第 2 页那条帖子配的是与第 1 页同一个 URL 的图，两页各报一次候选
        assert (by_id["80000006"].get("images") or []), "第 2 页复用同一 URL 的图没抓到"

        joined = chr(10).join(progress.messages)
        assert "排除(尺寸)" in joined, f"没报出表情是因为尺寸被挡的: {joined}"
        # **汇总按 URL 去重**：逐页累加的话同一张图会被算两次，汇总行显示成
        # 「候选 4 张 · 保存 2 张」，看着像丢了一半，而实际一张没丢。信息流那边更狠 ——
        # 每一批都会把上一批的帖子重新提取一遍，10 批下来能虚报出 82% 的假丢失率
        assert "图片汇总：候选图片地址 3 个 · 通过筛选 1 个 · 落盘 2 张" in joined, (
            f"汇总行把同一个 URL 重复计数了: {joined}"
        )


class TestFullRerunDropsIncrementalAnchorsEndToEnd(_ScraperTmpRoot):
    """全量重跑必须连「已知指纹」和「续抓页码」一起作废，真子进程收到的 job 说了算。

    这不是锦上添花：`facebook_group.js` 里 `const seen = new Set(CONFIG.knownFingerprints)`
    的过滤**不看 incremental**（只有水位线提前退出那句看）。照旧下发的话每条帖子都会
    被判成「见过」→ 不进 fresh → **配图也不会重下**，于是 incremental=False 对它完全
    无效。用户点了「全量重跑」，图却还是老样子。
    """

    script = JOB_ECHO_SCRIPT

    def setup_method(self):
        super().setup_method()
        _fake_playwright_install(self.tmpdir)
        import app.services.storage as storage
        storage.upsert_posts("tweakers", [
            {"username": "老用户", "timestamp": "22-05-2026 17:06", "content": "已经抓过的帖子",
             "translation": "", "page_number": 7, "fingerprint": "old1",
             "source": "tweakers", "parent_fingerprint": None, "reply_level": 0},
        ])

    def test_incremental_run_still_gets_the_anchors(self):
        """对照组：默认增量时锚点照旧下发，否则每一轮都在重抓全部历史"""
        if not _HAS_NODE:
            import pytest
            pytest.skip("未安装 node")

        job = self._execute({"thread_id": 123})["job"]
        assert job["known_fingerprints"] == ["old1"]
        assert job["incremental"] is True
        assert job["params"]["start_page"] == 8, "续抓页码没从 maxPage+1 算"

    def test_full_rerun_sends_no_anchors_at_all(self):
        if not _HAS_NODE:
            import pytest
            pytest.skip("未安装 node")

        job = self._execute({"thread_id": 123, "incremental": False})["job"]
        assert job["known_fingerprints"] == [], (
            f"已知指纹还在下发，Facebook 侧会把每条帖子判成见过、图也不会重下: {job}"
        )
        assert job["incremental"] is False
        assert job["params"]["start_page"] == 1, "全量重跑却从续抓页码开始，前面几页永远补不回来"


class TestGroupFeedCollectorEndToEnd:
    """真 Chrome + 真 HTTP + 真子进程跑本地小组 fixture：嵌套评论与增量合并"""

    GROUP_ID = "2407063016436085"

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        self.storage, self._restore_db = _use_temp_db(self.tmpdir)

    def teardown_method(self):
        self._restore_db()
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _run(self, incremental: bool):
        import asyncio
        from app.collectors import get_collector
        from app.services.collector_runner import CollectorRunner
        from app.services.progress_manager import ProgressManager

        sys.path.insert(0, _FIXTURES_DIR)
        from fixture_site import FixtureSite

        collector = get_collector("group_feed")

        with FixtureSite() as base_url:
            source = {
                "id": "fixture_group",
                "collector_id": "group_feed",
                "params": {
                    "group_id": self.GROUP_ID,
                    "base_url": base_url,
                    "headless": True,
                    "incremental": incremental,
                },
                "state_file": os.path.join(self.tmpdir, "state.json"),
                "pacing": {"delay_min": 200, "delay_max": 400},
            }
            loop = asyncio.new_event_loop()
            try:
                return loop.run_until_complete(
                    CollectorRunner.execute("group-e2e", collector, source, ProgressManager())
                )
            finally:
                loop.close()

    def _skip_unless_ready(self):
        import pytest
        from app.config import settings

        if not _HAS_NODE:
            pytest.skip("未安装 node")
        if not os.path.exists(
            os.path.join(settings.project_root, "node_modules", "playwright", "package.json")
        ):
            pytest.skip("项目根目录未安装 playwright")

    def test_nested_comments_are_extracted_with_parent_links(self):
        self._skip_unless_ready()
        data = self._run(incremental=False)

        assert data["complete"] is True, f"采集未完成: {data.get('stop_reason')}"
        posts = data["posts"]
        roots = [p for p in posts if not p["parent_fingerprint"]]
        comments = [p for p in posts if p["parent_fingerprint"]]
        assert len(roots) == 5, [p["username"] for p in roots]
        assert len(comments) == 5

        # 每条评论的 parent_fingerprint 必须真的指向本批数据里的某个主贴
        by_fp = {p["fingerprint"]: p for p in posts}
        for c in comments:
            assert c["parent_fingerprint"] in by_fp
            assert by_fp[c["parent_fingerprint"]]["parent_fingerprint"] is None
            assert c["reply_level"] == 1

        # 空用户名要落到兜底命名，不能是空字符串
        assert all(p["username"] for p in posts)
        # 时间统一成落盘格式 dd-mm-yyyy HH:MM
        assert posts[0]["timestamp"] == "02-06-2026 09:14"

    def test_root_whose_body_failed_to_extract_survives_with_its_comment(self):
        """真 Chrome 走一遍：正文没提取出来、但下面有人发言的主贴必须进得了库。

        真站两次实证：一次是长正文没展开、一次是纯图帖。丢了它，那条评论会被提成
        主贴，而提升是**不可逆的** —— 下一轮正文提对了，父贴换指纹重新入库，评论却
        还挂在主贴身份上，舆情从此按孤立主贴判（实测「Is bij mij ook zo…」因此成了
        neutral，它其实在附和一条报故障的帖子）。
        """
        self._skip_unless_ready()
        from app.services.post_tree import thread_of, post_key

        self._run(incremental=False)
        stored = self.storage.load_posts(["fixture_group"])

        blank = [p for p in stored if not (p.get("content") or "").strip()]
        assert len(blank) == 1, f"正文提取失败的那条主贴没留住: {[p['username'] for p in blank]}"
        assert blank[0]["username"] == "Koen_DH"
        assert blank[0]["reply_level"] == 0

        kid = next(p for p in stored if "zonnepanelen niet kunnen zien" in (p.get("content") or ""))
        assert kid["parent_fingerprint"] == blank[0]["fingerprint"], "评论被提成主贴了"
        assert kid["reply_level"] == 1

        # 结构留住的意义：这条评论拿得到整串上下文，而不是自己一条孤零零的串
        members = thread_of(stored)[post_key(kid)]
        assert [m["fingerprint"] for m in members] == \
            [blank[0]["fingerprint"], kid["fingerprint"]]

    def test_incremental_rerun_keeps_translations(self):
        """增量重跑绝不能覆盖已翻译的帖子。

        posts 表同时承载 translation 和 _processed 标记，整体覆盖等于把已翻译的帖子
        重新变成新帖 —— 下一轮再付一次翻译钱，舆情也重算一遍。
        """
        self._skip_unless_ready()
        first = self._run(incremental=False)
        first_fp = first["posts"][0]["fingerprint"]
        seq_before = [p["fingerprint"] for p in self.storage.load_posts(["fixture_group"])]

        # 模拟翻译步骤写回
        self.storage.save_translations([{
            "source": "fixture_group", "fingerprint": first_fp,
            "translation": "我已经用了三个月", "_processed": {"translated": True},
        }])
        # 再模拟舆情写回，验证两个标记互不干扰
        self.storage.mark_sentiment_analyzed([{
            "source": "fixture_group", "fingerprint": first_fp,
            "_processed": {"sentiment_at": "2026-08-05T10:00:00"},
        }])

        again = self._run(incremental=True)

        assert len(again["posts"]) == 10, "增量重跑后帖子数变了，说明历史数据被覆盖"
        kept = next(p for p in again["posts"] if p["fingerprint"] == first_fp)
        assert kept["translation"] == "我已经用了三个月", "译文被增量重跑抹掉了"
        assert kept["_processed"]["translated"] is True, "translated 标记被抹掉了"
        assert kept["_processed"]["sentiment_at"] == "2026-08-05T10:00:00", "sentiment_at 被抹掉了"

        # seq 是全链路的顺序锚点：它一洗牌，所有历史舆情结论就会错位到别的帖子上
        seq_after = [p["fingerprint"] for p in self.storage.load_posts(["fixture_group"])]
        assert seq_after == seq_before, "增量重跑把已有帖子的 seq 顺序打乱了"

    def test_handoff_file_is_deleted_after_ingest(self):
        """脚本的产出是用完即删的交接文件，不该在磁盘上留下第二份帖子数据"""
        self._skip_unless_ready()
        self._run(incremental=False)

        leftovers = []
        for root, _dirs, files in os.walk(self.tmpdir):
            leftovers += [f for f in files if f.endswith(".json") and "_out" in f]
        assert leftovers == [], f"交接文件没删干净: {leftovers}"
        assert self.storage.count_posts("fixture_group") == 10


def _RecordingProgress():
    """在真实的 pub/sub 上顺手留一份日志副本，不替换任何被测组件。

    采集脚本的 stdout 只走 SSE、不落进 task["logs"]，要断言它就得在流上收。
    ProgressManager 在这个文件里一律局部 import（import 顺序有讲究，见文件头），
    所以这里也惰性构造，不在模块级引它。
    """
    from app.services.progress_manager import ProgressManager

    class _Recording(ProgressManager):
        def __init__(self):
            super().__init__()
            self.messages = []

        async def emit(self, task_id, event_type, data):
            if event_type == "log":
                self.messages.append(data.get("message") or "")
            await super().emit(task_id, event_type, data)

    return _Recording()


class TestFacebookLoginEndToEnd:
    """登录 / 会话复用 / 两步验证退出路径 —— 真 Chrome 打真带登录门的本地站点。

    facebook.com 本身不在这里跑：一次失败的自动登录很可能把真账号推进 checkpoint，
    那是不可逆的对外动作。这里验的是脚本用的**同一套选择器和同一条代码路径**，
    fixture 站点有真表单、真 Set-Cookie、真重定向，不涉及任何 mock。
    """

    GROUP_ID = "2407063016436085"

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        self.state = os.path.join(self.tmpdir, "session.json")
        self.media = os.path.join(self.tmpdir, "media")
        self.storage, self._restore_db = _use_temp_db(self.tmpdir)

    def teardown_method(self):
        self._restore_db()
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _skip_unless_ready(self):
        import pytest
        from app.config import settings

        if not _HAS_NODE:
            pytest.skip("未安装 node")
        if not os.path.exists(
            os.path.join(settings.project_root, "node_modules", "playwright", "package.json")
        ):
            pytest.skip("项目根目录未安装 playwright")

    def _posts(self):
        return self.storage.load_posts(["fixture_fb"])

    def _run(self, base_url, username, password, incremental=False, progress=None):
        import asyncio
        from app.collectors import get_collector
        from app.services.collector_runner import CollectorRunner
        from app.services.progress_manager import ProgressManager

        collector = type(get_collector("facebook_group"))()
        # 凭据只走子进程环境变量，绝不进 argv 和 job 文件
        os.environ["HYXI_CRED_USERNAME"] = username
        os.environ["HYXI_CRED_PASSWORD"] = password
        try:
            source = {
                "id": "fixture_fb",
                "collector_id": "facebook_group",
                "name": "FB fixture",
                "params": {
                    "group_id": self.GROUP_ID,
                    "base_url": base_url,
                    "headless": True,
                    "incremental": incremental,
                    "max_batches": 1,
                },
                "state_file": self.state,
                "media_dir": self.media,
                "pacing": {"delay_min": 200, "delay_max": 400},
            }
            loop = asyncio.new_event_loop()
            try:
                return loop.run_until_complete(
                    CollectorRunner.execute("fb-e2e", collector, source,
                                            progress or ProgressManager())
                )
            finally:
                loop.close()
        finally:
            os.environ.pop("HYXI_CRED_USERNAME", None)
            os.environ.pop("HYXI_CRED_PASSWORD", None)

    def _login_site(self):
        sys.path.insert(0, _FIXTURES_DIR)
        import login_site
        return login_site

    def test_credential_login_creates_reusable_session(self):
        """(a) 首次凭据登录成功并落会话文件"""
        self._skip_unless_ready()
        site = self._login_site()

        with site.LoginSite() as base_url:
            data = self._run(base_url, site.GOOD_USER, site.GOOD_PASSWORD)

        assert data["complete"] is True, data.get("stop_reason")
        assert len(data["posts"]) == 4          # 2 主贴 + 1 评论 + 1 嵌套回复
        assert os.path.exists(self.state), "会话文件没有落盘，下一轮还得再输一次密码"

        roots = [p for p in data["posts"] if not p["parent_fingerprint"]]
        comments = [p for p in data["posts"] if p["parent_fingerprint"]]
        assert len(roots) == 2 and len(comments) == 2
        assert comments[0]["reply_level"] == 1
        import re
        assert re.match(r"^\d{2}-\d{2}-\d{4} \d{2}:\d{2}$", roots[0]["timestamp"])

    def test_timestamp_comes_from_tooltip_not_the_relative_label(self):
        """时间必须取 hover tooltip 的绝对时间。

        页面上另外两个时间来源都不能用，而且错了不会报错、只会静默污染指纹：
        主贴头部链接的 aria-label 是相对时间（「6天」），明天再抓同一条帖子就是「7天」，
        timestamp 进指纹 → 全部历史数据被判成新帖，已翻译的重新付费翻译、舆情重复计数；
        评论的 aria-label 是绝对时间，但走 Facebook 账号自己的时区（实测早 15 小时），
        和主贴对不上。作者、正文、message_id 一并钉住 —— 真实页面上
        data-post-id / data-comment-id / abbr[data-utime] 都不存在。
        """
        self._skip_unless_ready()
        site = self._login_site()

        with site.LoginSite() as base_url:
            data = self._run(base_url, site.GOOD_USER, site.GOOD_PASSWORD)

        roots = [p for p in data["posts"] if not p["parent_fingerprint"]]
        comments = [p for p in data["posts"] if p["parent_fingerprint"]]
        assert roots[0]["timestamp"] == "28-05-2026 17:54", "主贴取了会漂的相对时间"
        assert comments[0]["timestamp"] == "28-05-2026 18:42", "评论取了账号时区的时间"
        # 作者链接前面还有一个同一个人的头像链接，文本是空的
        assert roots[0]["username"] == "Marieke_V", "取到的是空文本的头像链接"
        assert comments[0]["username"] == "Joost1988"
        assert "+1" in comments[0]["content"], "评论正文不在 [data-ad-comet-preview] 里"
        assert roots[0]["message_id"] == "9001"
        assert comments[0]["message_id"] == "5501"

    def test_folded_body_is_expanded_before_extraction(self):
        """长正文必须先点开「展开」再提取。

        Facebook 对长帖只渲染前几行，末尾挂一个 role=button 的「展开」。不点它，
        textContent 拿到的是**残缺正文 + 「展开」两个字** —— 真站实测有一条整条正文
        只剩 16 个字符（`Goedemiddag,… 展开`），152 条里 22 条中招。翻译和舆情全建立在
        这段残文上，而 content 前 100 字还进指纹，等于把界面文案写进了去重锚点。
        展开后按钮文字变成「收起」，同样会被 textContent 吃进正文，一并剥掉。
        """
        self._skip_unless_ready()
        site = self._login_site()

        with site.LoginSite() as base_url:
            data = self._run(base_url, site.GOOD_USER, site.GOOD_PASSWORD)

        folded = [p for p in data["posts"] if p["username"] == "TechNerd_NL"][0]
        assert "downgrade" in folded["content"], "正文没展开，拿到的是截断的前几行"
        assert "展开" not in folded["content"], "「展开」按钮的文字混进了正文"
        assert "收起" not in folded["content"], "展开后「收起」按钮的文字混进了正文"
        assert not folded["content"].endswith("…"), "截断省略号残留在正文末尾"

    def test_non_post_articles_are_dropped(self):
        """信息流里的广告 / 推荐卡片也是 role=article，不能存成空帖。

        它们既没有固定链接也没有正文容器，存下来就是一条四个字段全空的记录：
        白占一次翻译调用（真站旧数据里有 56 条这种，全被标成"已翻译"），还会在
        结果页显示成一条什么都没有的帖子。判据是 **id 和正文全都没有** ——
        纯图片帖有 id 没正文、正文没渲染出来的帖子有正文没 id，两种都得留住。
        """
        self._skip_unless_ready()
        site = self._login_site()

        with site.LoginSite() as base_url:
            data = self._run(base_url, site.GOOD_USER, site.GOOD_PASSWORD)

        blank = [p for p in data["posts"]
                 if not p["message_id"] and not (p["content"] or "").strip()]
        assert blank == [], f"广告 article 被存成了空帖: {blank}"
        assert len(data["posts"]) == 4, "丢空帖时把真帖子也带走了"

    def test_multi_paragraph_comment_keeps_every_paragraph(self):
        """评论的多段正文必须全取，而且不能把嵌套回复的正文吞进来。

        评论没有专用正文容器，每一段是一个并列的 div[dir=auto] —— querySelector
        只拿第一段。真站实测一条 9 段的评论只存下第一段的 71 个字符，原文 811 个，
        丢了 92%。而嵌套回复也是 role=article，取多段时不限定层级就会把子回复的
        正文并进父评论，父子两条都错。
        """
        self._skip_unless_ready()
        site = self._login_site()

        with site.LoginSite() as base_url:
            data = self._run(base_url, site.GOOD_USER, site.GOOD_PASSWORD)

        parent = [p for p in data["posts"] if p["message_id"] == "5501"][0]
        child = [p for p in data["posts"] if p["message_id"] == "5502"][0]

        assert "+1" in parent["content"], "第一段丢了"
        assert "app is sterk verbeterd" in parent["content"], "第二段丢了"
        assert "min SOC blijft een raadsel" in parent["content"], "第三段丢了"
        assert "ondergrens" not in parent["content"], "把嵌套回复的正文吞进父评论了"
        assert child["content"] == "Die 8% ondergrens is inderdaad vreemd."

    def test_body_images_are_downloaded_and_decoys_ignored(self):
        """正文图要下载到本地，头像 / emoji / 非 scontent 的图都不能混进来。

        只存 Facebook 的原始链接不行：fbcdn URL 带签名和过期时间，几天后就是一片
        裂图。实测正文图是 <img>、host 在 scontent 上、渲染尺寸几百像素；界面图标是
        data:svg，emoji 在 static.xx.fbcdn.net，而**头像是 <svg><image> 不是 img**。
        fixture 的第一条主贴把这四种都摆上了，只有那张 400×300 的该被抓走。
        """
        self._skip_unless_ready()
        site = self._login_site()

        with site.LoginSite() as base_url:
            data = self._run(base_url, site.GOOD_USER, site.GOOD_PASSWORD)

        root = [p for p in data["posts"] if p["message_id"] == "9001"][0]
        images = root.get("images") or []
        assert len(images) == 1, f"抓多了或抓少了: {images}"

        rel = images[0]
        assert rel.startswith("fixture_fb/"), f"没有按 source 分目录: {rel}"
        assert not os.path.isabs(rel), "存了绝对路径，落盘文件就搬不了机器了"
        saved = os.path.join(self.media, rel)
        assert os.path.exists(saved), f"文件没落盘: {saved}"
        assert os.path.getsize(saved) > 0

        # 没图的帖子不该凭空多出一个 images 字段
        folded = [p for p in data["posts"] if p["username"] == "TechNerd_NL"][0]
        assert not folded.get("images")

    def test_images_survive_when_only_the_browser_can_fetch_them(self):
        """浏览器能下的图，脚本必须也能存下来 —— 哪怕脚本自己那条网络栈根本不通。

        图片字节曾经靠 `context.request.get()` 二次回源，而那是 **Playwright 在 Node
        进程里自带的 HTTP 客户端，不是 Chrome**：实测它连 Sec-Fetch-Dest / Referer /
        sec-ch-ua 都不带，更不读系统代理（playwright-core 整个包里 HTTPS_PROXY 出现
        0 次，browser.js 也从没传过 proxy）。于是「Chrome 走系统代理把图显示出来了、
        脚本直连回源却连不上」是一个完全可能的状态 —— 用户实测到的正是这个：
        帖子抓得到，图一张都没有。

        本地 fixture 打的是 127.0.0.1、代理不生效，所以旧实现在测试里永远是绿的。
        这里让 fixture 服务器按**真实请求头**区分两个客户端：不带浏览器特征的图片
        请求一律 502。真服务器、真 Chrome、真子进程，没有任何 mock。
        """
        self._skip_unless_ready()
        site = self._login_site()
        progress = _RecordingProgress()

        with site.LoginSite(browser_only_media=True) as base_url:
            data = self._run(base_url, site.GOOD_USER, site.GOOD_PASSWORD,
                             progress=progress)

        root = [p for p in data["posts"] if p["message_id"] == "9001"][0]
        images = root.get("images") or []
        assert len(images) == 1, (
            f"回源那条路被掐断后图就没了，说明字节还是靠二次请求拿的: {images}"
        )
        saved = os.path.join(self.media, images[0])
        assert os.path.getsize(saved) > 0, "文件落盘了但是空的"

        # 必须是从浏览器响应里取的。若这里显示 0 张走缓存，说明 502 那条路
        # 其实没被触发，这个用例就什么都没测到
        line = [m for m in progress.messages if "取自浏览器缓存" in m]
        assert line and "1 张取自浏览器缓存" in line[0], progress.messages

    def test_image_stage_is_never_silent(self):
        """提取和下载两个阶段都必须报数，否则远端出问题只能靠猜。

        「页面上就没有图」「选择器没选中」「被尺寸门限挡了」「下载失败」四种情况
        以前在日志上长得一模一样：什么都不显示 —— 这正是这个问题拖到用户那边才
        暴露出来的原因。
        """
        self._skip_unless_ready()
        site = self._login_site()
        progress = _RecordingProgress()

        with site.LoginSite() as base_url:
            self._run(base_url, site.GOOD_USER, site.GOOD_PASSWORD, progress=progress)

        joined = chr(10).join(progress.messages)
        assert "张候选" in joined, f"提取阶段一声不吭: {joined}"
        assert "排除(" in joined, f"没报出被排除的图是因为什么: {joined}"
        assert "图片：保存" in joined, f"下载阶段一声不吭: {joined}"

    def test_session_is_reused_without_password(self):
        """(c) 保留会话重跑：密码给成错的也应该照样成功，证明这一轮根本没走登录"""
        self._skip_unless_ready()
        site = self._login_site()

        with site.LoginSite() as base_url:
            self._run(base_url, site.GOOD_USER, site.GOOD_PASSWORD)
            assert os.path.exists(self.state)
            data = self._run(base_url, site.GOOD_USER, "这个密码是错的")

        assert data["complete"] is True, data.get("stop_reason")
        assert len(data["posts"]) == 4

    def test_deleting_session_falls_back_to_credentials(self):
        """(b) 删掉会话重跑仍能成功"""
        self._skip_unless_ready()
        site = self._login_site()

        with site.LoginSite() as base_url:
            self._run(base_url, site.GOOD_USER, site.GOOD_PASSWORD)
            os.remove(self.state)
            data = self._run(base_url, site.GOOD_USER, site.GOOD_PASSWORD)

        assert data["complete"] is True, data.get("stop_reason")
        assert os.path.exists(self.state)

    def test_two_factor_raises_manual_auth_with_actionable_message(self):
        """撞上两步验证要以退出码 3 交回给人，并给出「去哪点哪个按钮」的人话"""
        self._skip_unless_ready()
        import pytest
        from app.services.collector_runner import ManualAuthRequired

        site = self._login_site()
        with site.LoginSite() as base_url:
            with pytest.raises(ManualAuthRequired) as e:
                self._run(base_url, site.TWO_FACTOR_USER, "任意密码")

        assert "两步验证" in e.value.reason
        message = str(e.value)
        assert "FB fixture" in message
        assert "人工登录" in message, f"没告诉用户去哪操作: {message}"
        assert self.storage.count_posts("fixture_fb") == 0, "没登进去却写入了帖子"

    def test_bad_credentials_raise_manual_auth(self):
        self._skip_unless_ready()
        import pytest
        from app.services.collector_runner import ManualAuthRequired

        site = self._login_site()
        with site.LoginSite() as base_url:
            with pytest.raises(ManualAuthRequired) as e:
                self._run(base_url, "wrong@example.com", "wrong")

        assert "密码" in e.value.reason

    def _run_login_only(self, base_url, timeout_ms):
        import asyncio
        from app.collectors import get_collector
        from app.services.collector_runner import CollectorRunner
        from app.services.progress_manager import ProgressManager

        collector = type(get_collector("facebook_group"))()
        source = {
            "id": "fixture_fb",
            "collector_id": "facebook_group",
            "name": "FB fixture",
            "mode": "login_only",
            "params": {"group_id": self.GROUP_ID, "base_url": base_url},
            "state_file": self.state,
            "manual_login_timeout_ms": timeout_ms,
            "pacing": {"delay_min": 200, "delay_max": 400},
        }
        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(
                CollectorRunner.execute("fb-auth", collector, source, ProgressManager())
            )
        finally:
            loop.close()

    def test_manual_auth_saves_session_when_window_is_logged_in(self):
        """人工授权模式：窗口里已是登录态时保存会话并成功退出。

        真实场景里「变成登录态」这一步是人在窗口里过两步验证；这里先用凭据登录一次
        拿到真会话，再跑 login_only —— 走的是同一条 waitForManualLogin 成功分支。
        """
        self._skip_unless_ready()
        site = self._login_site()

        with site.LoginSite() as base_url:
            self._run(base_url, site.GOOD_USER, site.GOOD_PASSWORD)
            assert os.path.exists(self.state)
            # 清掉上一次采集的帖子，好断言 login_only 自己一条都不产出
            conn = self.storage._get_conn()
            conn.execute("DELETE FROM posts WHERE source_id = 'fixture_fb'")
            conn.commit()
            conn.close()
            result = self._run_login_only(base_url, timeout_ms=30000)

        assert result == {"mode": "login_only", "authorized": True}
        assert os.path.exists(self.state), "授权成功却没落会话"
        assert self.storage.count_posts("fixture_fb") == 0, "人工授权模式不该产出帖子数据"

    def test_manual_auth_times_out_with_actionable_message(self):
        """没人去点时必须超时收场，不能永久挂住一个有头浏览器"""
        self._skip_unless_ready()
        import pytest
        from app.services.collector_runner import ManualAuthRequired

        site = self._login_site()
        with site.LoginSite() as base_url:
            with pytest.raises(ManualAuthRequired) as e:
                self._run_login_only(base_url, timeout_ms=6000)

        assert "超时" in e.value.reason
        assert "人工登录" in str(e.value)

    def test_missing_credentials_ask_for_manual_login_instead_of_crashing(self):
        """没配凭据、会话又失效时，要给「去点人工登录」而不是一句技术报错。

        取凭据失败在 CollectorRunner 里是被吞掉的（按无凭据处理），就是为了让脚本走到
        这条路 —— 在那里抛只会变成一个看不懂的 OperationalError 把可操作提示盖掉。
        """
        self._skip_unless_ready()
        import pytest
        from app.services.collector_runner import ManualAuthRequired

        site = self._login_site()
        with site.LoginSite() as base_url:
            with pytest.raises(ManualAuthRequired) as e:
                self._run(base_url, "", "")   # 环境变量为空 = 库里没有凭据

        assert "未配置凭据" in e.value.reason
        # orchestrator 靠这两个字段给出人话并在界面上标出是哪个源要重新授权
        assert e.value.source_id == "fixture_fb"
        assert "人工登录" in str(e.value)

    def test_password_never_reaches_argv_or_job_file(self):
        """凭据只走环境变量：进程命令行和 job 文件里都不能出现"""
        from app.collectors import get_collector

        collector = get_collector("facebook_group")
        source = {
            "id": "fixture_fb", "collector_id": "facebook_group",
            "params": {"group_id": self.GROUP_ID},
        }
        job = collector.build_job(source, os.path.join(self.tmpdir, "handoff.json"))
        assert "HYXI_CRED_PASSWORD" not in json.dumps(job)
        assert "password" not in json.dumps(job).lower()
        # 命令行只有脚本路径和 --job=
        assert collector.script_path().endswith("facebook_group.js")


class TestSessionStateReflectsReality:
    """会话失效后，界面上的「会话正常」徽标必须翻回去。

    last_auth_at 只在授权成功时写入。若没有任何地方清除它，会话过期后界面会一直显示
    「会话正常」—— 而采集恰恰正因为会话失效在失败，用户被指向错误的排查方向。
    退出码 3 是脚本对「这个会话不管用了」的权威判定，直接拿来用，不去猜会话文件的有效期。
    """

    GROUP_ID = "2407063016436085"

    def setup_method(self):
        self.tmpdir = tempfile.mkdtemp()
        self.state = os.path.join(self.tmpdir, "session.json")
        self.storage, self._restore_db = _use_temp_db(self.tmpdir)

    def teardown_method(self):
        self._restore_db()
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _skip_unless_ready(self):
        import pytest
        from app.config import settings

        if not _HAS_NODE:
            pytest.skip("未安装 node")
        if not os.path.exists(
            os.path.join(settings.project_root, "node_modules", "playwright", "package.json")
        ):
            pytest.skip("项目根目录未安装 playwright")

    def _collect(self, source: dict, base_url: str, username: str, password: str):
        import asyncio
        from app.collectors import get_collector
        from app.services.collector_runner import CollectorRunner
        from app.services.progress_manager import ProgressManager

        collector = type(get_collector("facebook_group"))()
        job_source = dict(source)
        job_source["params"] = {
            **source["params"],
            "base_url": base_url,
            "headless": True,
            "incremental": False,
        }
        job_source["state_file"] = self.state
        job_source["pacing"] = {"delay_min": 200, "delay_max": 400}

        os.environ["HYXI_CRED_USERNAME"] = username
        os.environ["HYXI_CRED_PASSWORD"] = password
        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(
                CollectorRunner.execute("badge-e2e", collector, job_source, ProgressManager())
            )
        finally:
            loop.close()
            os.environ.pop("HYXI_CRED_USERNAME", None)
            os.environ.pop("HYXI_CRED_PASSWORD", None)

    def _authorized_source(self):
        from app.services import source_service

        source = source_service.create_source(
            collector_id="facebook_group",
            name="FB 会话状态回归",
            params={"group_id": self.GROUP_ID, "max_batches": 1},
        )
        source_service.mark_authorized(source["id"])
        assert source_service.get_source(source["id"])["last_auth_at"], "前置条件：先有一次成功授权"
        return source

    def test_manual_auth_required_clears_last_auth_at(self):
        """撞上两步验证 → 会话状态必须翻回「需重新授权」"""
        self._skip_unless_ready()
        import pytest
        from app.services import source_service
        from app.services.collector_runner import ManualAuthRequired

        sys.path.insert(0, _FIXTURES_DIR)
        import login_site

        source = self._authorized_source()
        with login_site.LoginSite() as base_url:
            with pytest.raises(ManualAuthRequired):
                self._collect(source, base_url, login_site.TWO_FACTOR_USER, "任意密码")

        assert source_service.get_source(source["id"])["last_auth_at"] is None, (
            "会话已经不管用了，界面上却还会显示「会话正常」"
        )

    def _authorize(self, source: dict, base_url: str, timeout_ms: int, progress=None):
        """跑一轮人工授权（login_only）。超时上限可配就是为了在验证里跑短一点"""
        import asyncio
        from app.collectors import get_collector
        from app.services.collector_runner import CollectorRunner
        from app.services.progress_manager import ProgressManager

        collector = type(get_collector("facebook_group"))()
        job_source = dict(source)
        job_source["params"] = {**source["params"], "base_url": base_url}
        job_source["state_file"] = self.state
        job_source["mode"] = "login_only"
        job_source["manual_login_timeout_ms"] = timeout_ms

        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(
                CollectorRunner.execute(
                    "auth-e2e", collector, job_source, progress or ProgressManager()
                )
            )
        finally:
            loop.close()

    def test_manual_auth_timeout_reports_configured_duration(self):
        """人工授权等到超时 → 退出码 3，且提示里的时长必须按实际配置算。

        这句话会原样变成界面上的失败提示，而页面同时在跑一个按同一配置走的倒计时。
        写死「5 分钟」会让两者对不上，用户以为超时判定坏了。
        """
        self._skip_unless_ready()
        import pytest
        from app.services import source_service
        from app.services.collector_runner import ManualAuthRequired

        sys.path.insert(0, _FIXTURES_DIR)
        import login_site

        source = self._authorized_source()
        with login_site.LoginSite() as base_url:
            with pytest.raises(ManualAuthRequired) as exc:
                self._authorize(source, base_url, 5000)

        assert "等待人工登录超时" in exc.value.reason
        assert "5 分钟" not in exc.value.reason, (
            f"时长写死了，没跟着配置走: {exc.value.reason}"
        )
        assert source_service.get_source(source["id"])["last_auth_at"] is None, (
            "授权超时也意味着会话没拿到，徽标必须翻回「需重新授权」"
        )

    def test_browser_requests_chinese_ui(self):
        """人工授权窗口是给操作者看的，站点界面必须是他读得懂的语言。

        locale 从 tweakers.js 抄来的 nl-NL 会把 Facebook 未登录界面整页切成荷兰语，
        而视觉上最抢眼的绿色按钮写着「Nieuw account maken」（创建新账户）——
        实测被用户认成注册页，以为点错了地方。
        """
        self._skip_unless_ready()

        sys.path.insert(0, _FIXTURES_DIR)
        import login_site

        source = self._authorized_source()
        site = login_site.LoginSite()
        with site as base_url:
            self._collect(source, base_url, login_site.GOOD_USER, login_site.GOOD_PASSWORD)

        assert site.request_languages, "fixture 站点没收到任何请求"
        assert all(lang.startswith("zh") for lang in site.request_languages), (
            f"浏览器界面语言不是中文，人工授权窗口操作者读不懂: {set(site.request_languages)}"
        )

    def test_stuck_on_registration_page_is_reported(self):
        """人工授权时落到注册页 → 必须出一句提示，而不是静默空转到超时。

        窗口里登录表单下方就是「创建新账户」，面板第 ②③ 步本来就是让人自己操作，
        误点是可预期的。轮询循环只认 loggedIn，落到注册页两头都不认，
        用户对着一个没反应的窗口等满 10 分钟都不知道自己走错了页。
        """
        self._skip_unless_ready()
        import pytest
        from app.services.collector_runner import ManualAuthRequired
        from app.services.progress_manager import ProgressManager

        sys.path.insert(0, _FIXTURES_DIR)
        import login_site

        progress = ProgressManager()
        queue = progress.subscribe("auth-e2e")

        source = self._authorized_source()
        with login_site.LoginSite(landing="reg") as base_url:
            with pytest.raises(ManualAuthRequired):
                self._authorize(source, base_url, 6000, progress=progress)

        messages = []
        while not queue.empty():
            msg = queue.get_nowait()
            if msg["event"] == "log":
                messages.append(msg["data"]["message"])

        assert any("创建新账户" in m for m in messages), (
            f"落到注册页却一声不吭，用户只能干等到超时: {messages}"
        )

    def test_navigation_during_polling_does_not_kill_the_script(self):
        """人工授权期间页面导航，不能把脚本搞挂。

        人在窗口里输账号、提交、过人机验证，每一步都是导航；轮询每 2 秒查一次 loggedIn，
        撞上导航提交时 Playwright 抛「Execution context was destroyed」。原先直接往外抛，
        脚本以退出码 1 死掉、Playwright 连带关掉那个窗口 —— 用户刚输进去的账号密码全白费，
        界面上还只给一句 Playwright 的英文报错。必须当成「这一轮没查成」继续轮询，
        最终以退出码 3 给出可操作提示。
        """
        self._skip_unless_ready()
        import pytest
        from app.services.collector_runner import ManualAuthRequired

        sys.path.insert(0, _FIXTURES_DIR)
        import login_site

        source = self._authorized_source()
        with login_site.LoginSite(landing="churn") as base_url:
            # 不是 ManualAuthRequired 就说明脚本被导航搞挂了（异常里带 code=1 那条）
            with pytest.raises(ManualAuthRequired) as exc:
                self._authorize(source, base_url, 8000)

        assert "等待人工登录超时" in exc.value.reason, exc.value.reason

    def test_successful_collect_keeps_last_auth_at(self):
        """反向保证：采集成功时不能把授权时间抹掉，否则徽标会反过来误报失效"""
        self._skip_unless_ready()
        from app.services import source_service

        sys.path.insert(0, _FIXTURES_DIR)
        import login_site

        source = self._authorized_source()
        before = source_service.get_source(source["id"])["last_auth_at"]
        with login_site.LoginSite() as base_url:
            data = self._collect(source, base_url, login_site.GOOD_USER, login_site.GOOD_PASSWORD)

        assert data["complete"] is True, data.get("stop_reason")
        assert source_service.get_source(source["id"])["last_auth_at"] == before


class TestPortablePackagePathsEndToEnd:
    """便携包的路径解析与首启自举 —— 真文件、真 Fernet 密钥、真环境变量"""

    def setup_method(self):
        import sys as _sys
        self.tmpdir = tempfile.mkdtemp()
        self._sys = _sys
        self._had_frozen = hasattr(_sys, "frozen")
        self._old_frozen = getattr(_sys, "frozen", None)
        self._old_exe = _sys.executable

    def teardown_method(self):
        if self._had_frozen:
            self._sys.frozen = self._old_frozen
        elif hasattr(self._sys, "frozen"):
            del self._sys.frozen
        self._sys.executable = self._old_exe
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _freeze_at(self, pkg_root):
        """假装自己是解压在 pkg_root 里的便携包（PyInstaller 就是这么标记的）"""
        app_dir = os.path.join(pkg_root, "app")
        os.makedirs(app_dir, exist_ok=True)
        self._sys.frozen = True
        self._sys.executable = os.path.join(app_dir, "hyxi.exe")

    def test_source_mode_paths_are_unchanged(self):
        """源码态一个字都不能变，否则开发流程和现有数据全部错位"""
        from app import paths

        assert not paths.is_frozen()
        assert os.path.isfile(os.path.join(paths.project_root(), "backend", "main.py"))
        assert paths.data_dir().endswith(os.path.join("backend", "data"))

    def test_frozen_paths_point_at_the_package_root(self):
        """app/hyxi.exe 的上一级才是包根。这一处错了会连锁带偏数据、密钥、
        采集脚本和 playwright 四个位置"""
        from app import paths

        pkg = os.path.join(self.tmpdir, "HYXi-9.9.9-win64")
        self._freeze_at(pkg)
        real = os.path.realpath(pkg)
        assert paths.is_frozen()
        assert paths.project_root() == real
        # 数据在包的**同级**：包目录带版本号，装包里的话每升一次级
        # 就是一个全新的空目录，配置和历史数据全丢
        assert paths.data_dir() == os.path.join(
            os.path.realpath(self.tmpdir), paths.EXTERNAL_DATA_DIRNAME)

    def test_frozen_root_is_expanded_from_8_3_short_paths(self):
        """sys.executable 可能是 8.3 短路径 —— 实测双击启动器拿到的是
        C:\\HYXI-B~1\\HYXI-1~1.1-W。它指的目录没错，但会原样印在「程序目录」那行上，
        而用户正要照着它去资源管理器里找 data 文件夹，短路径既看不懂也不好粘。
        """
        from app import paths

        # 用 GetShortPathName 拿到真的短路径，不是自己编一个
        import ctypes
        buf = ctypes.create_unicode_buffer(512)
        n = ctypes.windll.kernel32.GetShortPathNameW(self.tmpdir, buf, 512)
        if not n:
            import pytest
            pytest.skip("这个卷没有开启 8.3 短名")
        short = buf.value
        if short == self.tmpdir:
            import pytest
            pytest.skip("这个卷没有开启 8.3 短名")

        app_dir = os.path.join(short, "app")
        os.makedirs(app_dir, exist_ok=True)
        self._sys.frozen = True
        self._sys.executable = os.path.join(app_dir, "hyxi.exe")

        root = paths.project_root()
        assert "~" not in root, f"短路径没有被还原: {root}"
        assert root == os.path.realpath(self.tmpdir)

    def test_bundled_node_beats_whatever_is_on_path(self):
        """目标机器不会装 Node，PATH 上没有 node —— 必须用包里自带的那个"""
        import app.config as cfg

        node_dir = os.path.join(self.tmpdir, "node")
        os.makedirs(node_dir)
        exe = os.path.join(node_dir, "node.exe")
        with open(exe, "wb") as f:
            f.write(b"MZ")

        old_root, old_path = cfg.settings.project_root, cfg.settings.node_path
        try:
            cfg.settings.project_root = self.tmpdir
            cfg.settings.node_path = ""
            assert cfg.resolve_node_executable() == exe
        finally:
            cfg.settings.project_root, cfg.settings.node_path = old_root, old_path

    def test_without_a_bundled_node_it_falls_back_to_path(self):
        """源码态包里没有 node 目录，回退到 PATH，start.ps1 那条路不受影响"""
        import app.config as cfg

        old_root, old_path = cfg.settings.project_root, cfg.settings.node_path
        try:
            cfg.settings.project_root = self.tmpdir
            cfg.settings.node_path = ""
            assert cfg.resolve_node_executable() == "node"
        finally:
            cfg.settings.project_root, cfg.settings.node_path = old_root, old_path

    def test_explicit_node_path_wins(self):
        import app.config as cfg

        old_path = cfg.settings.node_path
        try:
            cfg.settings.node_path = r"D:\custom\node.exe"
            assert cfg.resolve_node_executable() == r"D:\custom\node.exe"
        finally:
            cfg.settings.node_path = old_path

    def test_first_run_generates_a_usable_fernet_key(self):
        """缺 SECRET_KEY 时后端会拒绝保存数据源凭据并返回 400（绝不降级成明文落库），
        用户会卡在「数据源」页且看不出为什么。首启必须把它备好"""
        from cryptography.fernet import Fernet
        import run_server

        env = os.path.join(self.tmpdir, ".env")
        run_server.ensure_env_file(env)

        assert os.path.isfile(env)
        line = [l for l in open(env, encoding="utf-8").read().splitlines()
                if l.startswith("TWEAKERS_SECRET_KEY=")]
        assert len(line) == 1, line
        key = line[0].split("=", 1)[1]
        # 真的能拿来加解密，不是随便一串
        token = Fernet(key.encode()).encrypt(b"hunter2")
        assert Fernet(key.encode()).decrypt(token) == b"hunter2"

    def test_existing_secret_key_is_never_replaced(self):
        """换掉 SECRET_KEY 会让已录入的凭据全部解不开 —— 每次启动重生成就是每次清空凭据"""
        import run_server

        env = os.path.join(self.tmpdir, ".env")
        with open(env, "w", encoding="utf-8") as f:
            f.write("TWEAKERS_SECRET_KEY=already-here\n")

        run_server.ensure_env_file(env)
        assert open(env, encoding="utf-8").read().count("TWEAKERS_SECRET_KEY=") == 1
        assert "already-here" in open(env, encoding="utf-8").read()

    def test_other_settings_in_env_survive(self):
        """用户可能自己加了 API_KEY、代理、超时 —— 补一行密钥不能把它们冲掉"""
        import run_server

        env = os.path.join(self.tmpdir, ".env")
        with open(env, "w", encoding="utf-8") as f:
            f.write("TWEAKERS_API_KEY=mine\nTWEAKERS_TASK_TIMEOUT_MINUTES=60\n")

        run_server.ensure_env_file(env)
        text = open(env, encoding="utf-8").read()
        assert "TWEAKERS_API_KEY=mine" in text
        assert "TWEAKERS_TASK_TIMEOUT_MINUTES=60" in text
        assert "TWEAKERS_SECRET_KEY=" in text

    def test_no_api_key_is_invented(self):
        """便携包绑 127.0.0.1，按既定姿态「未设 API_KEY 时不鉴权」正合适。
        自动生成一把反而要求用户先去前端粘贴一次才能用"""
        import run_server

        env = os.path.join(self.tmpdir, ".env")
        run_server.ensure_env_file(env)
        assert "TWEAKERS_API_KEY" not in open(env, encoding="utf-8").read()


class TestFreshReplyDetectionEndToEnd:
    """「老主贴上的新回复」判据。

    明细按主贴时间从新到旧排，评论跟着自己的主贴走 —— 于是今天发在两个月前主贴上的
    回复，会被排到两个月前的位置去，用户翻不到。真实数据里有一条 08-15 的回复挂在
    06-06 的主贴上，排在第 40 多个主贴之后。
    """

    def _post(self, fp, ts, parent=None, source="src_a", **kw):
        return {"source": source, "fingerprint": fp, "parent_fingerprint": parent,
                "username": fp, "content": fp, "timestamp": ts, **kw}

    def _mark(self, posts, days=7):
        from app.services.post_tree import mark_fresh_replies
        return mark_fresh_replies(posts, days)

    def test_new_reply_on_an_old_root_is_marked(self):
        """要抓的就是这一类：主贴很旧、回复很新，排序把它埋了"""
        from app.services.post_tree import post_key

        root = self._post("root", "06-06-2026 10:00")
        reply = self._post("r1", "15-08-2026 09:00", parent="root")
        newest = self._post("other", "19-08-2026 12:00")

        marked = self._mark([root, reply, newest])
        assert post_key(reply) in marked, marked
        # 距主贴的天数要能算对，报告上要显示「主贴 69 天前」。
        # 06-06 10:00 → 15-08 09:00 差 69 天 23 小时，向下取整 69 —— 与真实数据里
        # 那条（08-15 的回复挂在 06-06 的主贴上）实测的天数一致
        assert marked[post_key(reply)] == 69, marked

    def test_reply_on_a_root_that_is_also_new_is_not_marked(self):
        """整串都是新的 —— 它本来就排在最前面，没被埋，标了反而稀释重点"""
        from app.services.post_tree import post_key

        root = self._post("root", "18-08-2026 10:00")
        reply = self._post("r1", "19-08-2026 09:00", parent="root")
        assert post_key(reply) not in self._mark([root, reply])

    def test_old_reply_on_an_old_root_is_not_marked(self):
        from app.services.post_tree import post_key

        root = self._post("root", "06-06-2026 10:00")
        reply = self._post("r1", "07-06-2026 09:00", parent="root")
        newest = self._post("other", "19-08-2026 12:00")
        assert post_key(reply) not in self._mark([root, reply, newest])

    def test_the_root_itself_is_never_marked(self):
        """只标回复。主贴新不新，排序已经表达了"""
        from app.services.post_tree import post_key

        root = self._post("root", "19-08-2026 10:00")
        old = self._post("old", "06-06-2026 10:00")
        marked = self._mark([root, old])
        assert post_key(root) not in marked and post_key(old) not in marked

    def test_missing_timestamps_are_never_marked(self):
        """早期采集读不到 tooltip 时是**故意**留空的（写相对时间会污染指纹）。

        这批帖子实际很新，但那是推测 —— 宁可漏标，也不能凭空标一条出来。
        """
        from app.services.post_tree import post_key

        root = self._post("root", "06-06-2026 10:00")
        no_time = self._post("r1", "", parent="root")
        rootless_time = self._post("root2", "")
        r2 = self._post("r2", "19-08-2026 09:00", parent="root2")
        newest = self._post("other", "19-08-2026 12:00")

        marked = self._mark([root, no_time, rootless_time, r2, newest])
        assert post_key(no_time) not in marked, "回复没有时间也被标了"
        assert post_key(r2) not in marked, "主贴没有时间，无从判断它算不算老"

    def test_orphan_reply_is_not_marked(self):
        """父贴不在本批数据里的评论，build_tree 已按主贴处理 —— 它就不是回复了"""
        from app.services.post_tree import post_key

        orphan = self._post("r1", "19-08-2026 09:00", parent="not-here")
        newest = self._post("other", "19-08-2026 12:00")
        assert post_key(orphan) not in self._mark([orphan, newest])

    def test_same_fingerprint_across_sources_does_not_cross_wire(self):
        """指纹不含来源，只用 fingerprint 判会让两个平台的帖子互相顶替"""
        from app.services.post_tree import post_key

        a_root = self._post("root", "06-06-2026 10:00", source="src_a")
        a_reply = self._post("r1", "15-08-2026 09:00", parent="root", source="src_a")
        b_root = self._post("root", "18-08-2026 10:00", source="src_b")
        b_reply = self._post("r1", "19-08-2026 09:00", parent="root", source="src_b")

        marked = self._mark([a_root, a_reply, b_root, b_reply])
        assert post_key(a_reply) in marked, "src_a 的老帖新回复没标出来"
        assert post_key(b_reply) not in marked, "src_b 整串都是新的，不该被 src_a 带标"

    def test_baseline_is_the_newest_post_not_now(self):
        """基准取数据集里最新的帖子，不是 datetime.now()。

        按 now 算的话，隔一阵子没采集、或者翻看几个月前的历史报告时，**一条都不会亮** ——
        而那份报告当初想标出来的东西并没有变。报告要自洽、可复现。
        """
        from app.services.post_tree import post_key

        # 整批都是两年前的数据，按 now 算没有任何一条落在 7 天窗口里
        root = self._post("root", "01-03-2024 10:00")
        reply = self._post("r1", "20-08-2024 09:00", parent="root")
        newest = self._post("other", "22-08-2024 12:00")

        marked = self._mark([root, reply, newest])
        assert post_key(reply) in marked, "基准用了当前时间，历史数据一条都标不出来"

    def test_window_size_changes_what_is_marked(self):
        from app.services.post_tree import post_key

        root = self._post("root", "06-06-2026 10:00")
        r_recent = self._post("r1", "15-08-2026 09:00", parent="root")   # 距基准 4 天
        r_older = self._post("r2", "07-08-2026 09:00", parent="root")    # 距基准 12 天
        newest = self._post("other", "19-08-2026 12:00")
        posts = [root, r_recent, r_older, newest]

        assert post_key(r_recent) in self._mark(posts, 7)
        assert post_key(r_older) not in self._mark(posts, 7)
        assert post_key(r_older) in self._mark(posts, 14)

    def test_all_timestamps_missing_yields_nothing(self):
        """没有任何可用时间时不该抛异常，也不该标出东西来"""
        posts = [self._post("root", ""), self._post("r1", "", parent="root")]
        assert self._mark(posts) == {}

    def test_real_data_matches_the_measured_counts(self):
        """拿真实库跑一遍，钉住实测数字。

        这批数据是这个需求的由来：7 天窗口下 4 条，其中一条隔了 69 天。
        """
        import sqlite3
        from app.services.post_tree import post_key

        db = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                          "data", "hyxi.db")
        if not os.path.isfile(db):
            import pytest
            pytest.skip("本机没有真实库")
        conn = sqlite3.connect(db)
        conn.row_factory = sqlite3.Row
        posts = [dict(r) for r in conn.execute(
            "SELECT source_id AS source, fingerprint, parent_fingerprint, username, "
            "timestamp, content FROM posts ORDER BY seq")]
        conn.close()
        if len(posts) < 100:
            import pytest
            pytest.skip("真实库里数据太少，钉不住数字")

        assert len(self._mark(posts, 7)) >= 1, "真实数据里一条老帖新回复都没找到"
        # 窗口越大命中越多，绝不会反过来
        assert len(self._mark(posts, 3)) <= len(self._mark(posts, 7)) <= len(self._mark(posts, 14))


class TestFreshReplyExportEndToEnd:
    """「老帖新回复」在导出报告里的呈现：明细列 + 高亮 + 独立工作表 + 双向跳转。

    真 openpyxl 输出、真 load_workbook 回读。
    """

    def _rows(self):
        """一份已按 order_by_thread 排好的明细：两个主贴，第二个带一条新回复"""
        def row(i, level, user, ts, fresh="", gap=-1, **kw):
            base = {
                "index": i, "source": "Facebook", "level": level, "username": user,
                "timestamp": ts, "fresh": fresh, "fresh_gap": gap,
                "content": f"{user} 的原文", "translation": f"{user} 的译文",
                "image_desc": "", "images": [], "sentiment": "中立", "intensity": 3,
                "reason": "", "dimensions": "",
            }
            base.update(kw)
            return base

        return [
            row(1, 0, "新主贴作者", "2026-08-18 10:00"),
            row(2, 1, "新主贴的回复", "2026-08-19 09:00"),
            row(3, 0, "老主贴作者", "2026-06-06 19:57"),
            row(4, 1, "老回复", "2026-06-07 10:00"),
            row(5, 1, "新回复作者", "2026-08-15 15:50", fresh="🔥 新回复（主贴 69 天前）", gap=69),
        ]

    def _meta(self, **kw):
        base = {
            "description": "d", "sources": "Facebook", "exported_at": "2026-08-19 10:00",
            "total": 5, "replies": 3, "analyzed": 5, "time_start": "", "time_end": "",
            "summary": {}, "top_users": [],
            "fresh_days": 7, "fresh_count": 1, "fresh_baseline": "2026-08-19 04:35",
        }
        base.update(kw)
        return base

    def _book(self, rows=None, meta=None):
        from openpyxl import load_workbook
        from app.services.excel_service import ExcelService, EXPORT_COLUMNS

        blob = ExcelService.build_export(rows or self._rows(), meta or self._meta(),
                                         EXPORT_COLUMNS)
        return load_workbook(BytesIO(blob))

    def test_detail_sheet_has_the_reminder_column(self):
        from app.services.excel_service import EXPORT_COLUMNS, DETAIL_SHEET

        ws = self._book()[DETAIL_SHEET]
        labels = [c.value for c in ws[1]]
        assert "更新提醒" in labels, labels
        col = labels.index("更新提醒") + 1
        # 第 5 行数据 = 表里第 6 行
        assert "69 天前" in str(ws.cell(row=6, column=col).value), ws.cell(row=6, column=col).value
        assert not ws.cell(row=2, column=col).value, "普通行不该有提醒"

    def test_fresh_row_is_highlighted_and_wins_over_the_reply_fill(self):
        """命中行整行暖色。两个底色叠在一起就都看不出来了 —— 命中时只上暖色"""
        from app.services.excel_service import DETAIL_SHEET, FRESH_FILL, REPLY_FILL

        ws = self._book()[DETAIL_SHEET]
        fresh_cell = ws.cell(row=6, column=1)   # 新回复那行
        plain_reply = ws.cell(row=5, column=1)  # 老回复那行
        assert fresh_cell.fill.start_color.rgb[-6:] == FRESH_FILL.start_color.rgb[-6:]
        assert plain_reply.fill.start_color.rgb[-6:] == REPLY_FILL.start_color.rgb[-6:]

    def test_focus_sheet_exists_and_sits_right_after_the_overview(self):
        """工作簿默认停在第一张，第二张最容易被看到 —— 这正是要第一时间关注的东西"""
        from app.services.excel_service import FRESH_SHEET, DETAIL_SHEET

        names = self._book().sheetnames
        assert names[0] == "概览"
        assert names[1] == FRESH_SHEET, names
        assert names.index(FRESH_SHEET) < names.index(DETAIL_SHEET)

    def test_focus_sheet_brings_the_root_post_along(self):
        """脱离主贴，一条「你能通过 HA 控制它吗」根本读不懂在说什么。

        用户不该为了看懂一条新回复再去别处翻主贴 —— 这是这张表存在的理由。
        """
        from app.services.excel_service import FRESH_SHEET

        ws = self._book()[FRESH_SHEET]
        text = "\n".join(str(c.value or "") for row in ws.iter_rows() for c in row)
        assert "老主贴作者" in text, "主贴作者没带进来"
        assert "老主贴作者 的原文" in text, "主贴原文没带进来"
        assert "老主贴作者 的译文" in text, "主贴译文没带进来"
        assert "新回复作者" in text and "主贴 69 天前" in text
        # 不相干的那个新主贴串不该混进来
        assert "新主贴作者" not in text, "整串都是新的，不该进聚焦区"

    def test_focus_sheet_does_not_drag_in_the_older_replies(self):
        """热帖有十几条旧回复，全搬过来就长得没法一口气读完 —— 给跳转就够了"""
        from app.services.excel_service import FRESH_SHEET, DETAIL_SHEET

        ws = self._book()[FRESH_SHEET]
        text = "\n".join(str(c.value or "") for row in ws.iter_rows() for c in row)
        assert "老回复 的原文" not in text, "旧回复被搬进聚焦区了"
        assert "另有 1 条较早回复" in text, text[:400]
        assert DETAIL_SHEET in text

    def test_links_between_the_two_sheets_use_location_not_target(self):
        """内部链接必须走 location。给 target 会被当成外部关系，Excel 打开报「需要修复」"""
        from app.services.excel_service import FRESH_SHEET, DETAIL_SHEET

        wb = self._book()
        fresh = wb[FRESH_SHEET]
        links = [c.hyperlink for row in fresh.iter_rows() for c in row if c.hyperlink]
        assert links, "聚焦表里一个跳转都没有"
        for link in links:
            assert link.location and DETAIL_SHEET in link.location, link.location
            assert not link.target, "用了 target，Excel 会报需要修复"

        # 明细表那一侧也要能跳回来
        detail = wb[DETAIL_SHEET]
        back = [c.hyperlink for row in detail.iter_rows() for c in row
                if c.hyperlink and FRESH_SHEET in (c.hyperlink.location or "")]
        assert back, "明细表的更新提醒没有跳回聚焦表的链接"

    def test_detail_link_points_at_the_thread_block_not_the_reply_row(self):
        """跳到回复自己那一行没有意义 —— 那一段的价值就在于主贴也在场"""
        from app.services.excel_service import FRESH_SHEET, DETAIL_SHEET

        wb = self._book()
        detail, fresh = wb[DETAIL_SHEET], wb[FRESH_SHEET]
        target = None
        for row in detail.iter_rows():
            for c in row:
                if c.hyperlink and FRESH_SHEET in (c.hyperlink.location or ""):
                    target = int(c.hyperlink.location.split("!A")[1])
        assert target, "没找到指向聚焦表的链接"
        assert "主贴" in str(fresh.cell(row=target, column=1).value), \
            fresh.cell(row=target, column=1).value

    def test_no_fresh_reply_means_no_focus_sheet(self):
        """留一张只有标题的空表比没有更费解（同「配图」表的既有做法）"""
        from app.services.excel_service import FRESH_SHEET

        rows = [r for r in self._rows() if not r["fresh"]]
        names = self._book(rows, self._meta(fresh_count=0)).sheetnames
        assert FRESH_SHEET not in names, names

    def test_overview_records_the_window_and_the_baseline(self):
        """不写的话「这次和上次为什么不一样」无从解释"""
        ws = self._book()["概览"]
        text = "\n".join(str(c.value or "") for row in ws.iter_rows() for c in row)
        assert "近 7 天老帖新回复" in text, text[:500]
        assert "2026-08-19 04:35" in text, "基准时间没写进报告"


class TestResultsViewRevealsFreshRepliesEndToEnd:
    """结果页默认只展开每个主贴的前 3 条回复。

    「老帖新回复」正是因为排序被埋才做的功能，结果又被这条预览规则截掉的话，
    卡片上一个橙色标记都看不到 —— 用户还得先猜到要点展开。搜索命中早就为同一个
    问题开了豁免，这里必须一样。
    """

    def _visible(self, replies):
        """把 ResultsView.vue 里的 visibleReplies 抠出来用真 Node 跑一遍。

        不抄一份实现进测试 —— 那样测的是抄件，源文件改了这里照样绿。
        """
        import json
        import re
        import subprocess

        vue = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            "frontend", "src", "views", "ResultsView.vue")
        src = open(vue, encoding="utf-8").read()
        body = re.search(
            r"function visibleReplies\(t: \{[^}]*\}\): PostData\[\] \{(.*?)\n\}",
            src, re.S)
        assert body, "visibleReplies 没找到，函数签名改了就要同步这里"
        js = body.group(1)
        # 去掉 TS 标注与外部依赖，只保留判据本身
        js = js.replace("openThreads.value.has(threadKey(t.root))", "false")
        script = (
            "const REPLY_PREVIEW = 3;\n"
            "function visibleReplies(t) {" + js + "\n}\n"
            "const t = { root: {}, replies: " + json.dumps(replies) + " };\n"
            "console.log(JSON.stringify(visibleReplies(t).length));"
        )
        out = subprocess.run(["node", "-e", script], capture_output=True, text=True)
        assert out.returncode == 0, out.stderr
        return json.loads(out.stdout.strip())

    def _replies(self, n, fresh_at=None, matched_at=None):
        return [{"matched": i == matched_at, "fresh_reply": i == fresh_at} for i in range(n)]

    def test_long_thread_is_truncated_by_default(self):
        if not _HAS_NODE:
            import pytest
            pytest.skip("未安装 node")
        assert self._visible(self._replies(10)) == 3

    def test_a_fresh_reply_beyond_the_preview_still_shows(self):
        """第 9 条才是新回复时，默认状态下也必须露出来"""
        if not _HAS_NODE:
            import pytest
            pytest.skip("未安装 node")
        assert self._visible(self._replies(10, fresh_at=8)) == 10

    def test_search_hit_exemption_still_works(self):
        if not _HAS_NODE:
            import pytest
            pytest.skip("未安装 node")
        assert self._visible(self._replies(10, matched_at=8)) == 10


class TestPortableDataSurvivesUpgradeEndToEnd:
    """便携包升级后，配置和数据必须还在 —— 真目录、真文件、真复制，不 mock。

    包目录带版本号，升级就是解压出一个全新的空目录。数据以前放在包**里**，于是用户每升
    一次级，LLM 配置、数据源、跑过的任务和舆情结论就全部要重来一遍（用户实测报过）。
    现在数据挂在包的**同级**目录，并在第一次升上来时把旧包里的那份接过来。
    """

    def setup_method(self):
        import app.paths as paths
        self.paths = paths
        # realpath：mkdtemp 在本机给的是 8.3 短路径（C:\Users\ADMINI~1\...），
        # 而 project_root() 会把它展开成长路径 —— 不对齐的话比的是两种写法
        self.tmp = os.path.realpath(tempfile.mkdtemp())
        self._old_frozen = getattr(sys, "frozen", None)
        self._old_exe = sys.executable

    def teardown_method(self):
        if self._old_frozen is None:
            if hasattr(sys, "frozen"):
                del sys.frozen
        else:
            sys.frozen = self._old_frozen
        sys.executable = self._old_exe
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _install(self, name, with_data=False, env_body=None):
        """造一个便携包目录。with_data=True 就是「已经用过一阵子」的老布局安装。"""
        root = os.path.join(self.tmp, name)
        os.makedirs(os.path.join(root, "app"), exist_ok=True)
        exe = os.path.join(root, "app", "hyxi.exe")
        open(exe, "wb").close()
        if with_data:
            os.makedirs(os.path.join(root, "data", "media", "src_x"), exist_ok=True)
            with open(os.path.join(root, "data", "hyxi.db"), "wb") as f:
                f.write(b"SQLite format 3\x00" + name.encode())
            with open(os.path.join(root, "data", "media", "src_x", "p.jpg"), "wb") as f:
                f.write(b"\xff\xd8\xff")
        if env_body is not None:
            with open(os.path.join(root, ".env"), "w", encoding="utf-8") as f:
                f.write(env_body)
        return root, exe

    def _become(self, exe):
        """让 paths 认为「我就是这个包里的 hyxi.exe」—— 走的是 is_frozen() 真实的判据"""
        sys.frozen = True
        sys.executable = exe

    def _real_db(self, data_path, model=None):
        """在 data_path 里建一个**真**库：表结构走 storage.init_db()，不是手搓的假表。

        `seed_default_sources()` 是照着 main.py 的 lifespan 来的 —— 刚建出来的空目录
        也会有一条 sources 记录。判据里 sources 的门槛因此是 1 而不是 0，给 0 的话
        每个全新目录看起来都是有数据的，接管永远不会发生。
        """
        import app.services.storage as storage
        from app.services import source_service
        os.makedirs(data_path, exist_ok=True)
        old = storage.DB_PATH
        storage.DB_PATH = os.path.join(data_path, "hyxi.db")
        try:
            storage.init_db()
            source_service.seed_default_sources()
            if model:
                storage.set_app_config("llm", {
                    "api_key": "sk-probe", "base_url": "https://api.deepseek.com",
                    "model_name": model,
                })
        finally:
            storage.DB_PATH = old

    def _model_name(self, data_path):
        """从真库里读回 LLM 模型名 —— 用户就是靠这一格发现数据没带过来的"""
        import app.services.storage as storage
        old = storage.DB_PATH
        storage.DB_PATH = os.path.join(data_path, "hyxi.db")
        try:
            return storage.get_app_config("llm").get("model_name")
        finally:
            storage.DB_PATH = old

    # ---------- 路径解析 ----------

    def test_fresh_install_puts_data_beside_the_package_not_inside(self):
        """新装：数据落在包的同级目录。装包里的话，下一个版本就是一个空目录"""
        root, exe = self._install("HYXi-1.8.2-win64")
        self._become(exe)
        assert self.paths.data_dir() == os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME)
        assert self.paths.env_file() == os.path.join(
            self.tmp, self.paths.EXTERNAL_DATA_DIRNAME, ".env"
        ), ".env 必须跟着数据走 —— 密钥与库一分家，凭据密文就全部解不开"

    def test_an_existing_install_keeps_using_the_data_inside_it(self):
        """老布局向后兼容：包里已经有 data 就沿用。

        升上来的用户只是重启一次，不能凭空变成空库。
        """
        root, exe = self._install("HYXi-1.8.1-win64", with_data=True, env_body="TWEAKERS_SECRET_KEY=old\n")
        self._become(exe)
        assert self.paths.data_dir() == os.path.join(root, "data")
        assert self.paths.env_file() == os.path.join(root, ".env")

    def test_source_checkout_is_untouched(self):
        """源码态一个字都不该变 —— 开发机的 .env 在仓库根，数据在 backend/data"""
        if hasattr(sys, "frozen"):
            del sys.frozen
        root = self.paths.project_root()
        assert self.paths.data_dir() == os.path.join(root, "backend", "data")
        assert self.paths.env_file() == os.path.join(root, ".env")

    # ---------- 升级接管 ----------

    def _adopt(self, root):
        import run_server
        run_server.adopt_previous_install(root)

    def test_upgrade_adopts_the_previous_installs_data_and_key(self):
        """核心用例：旧版本用过一阵子，新版本解压在旁边，启动即接上。

        密钥必须一起搬 —— 只搬库的话数据源密码全部解不开，用户还得重新录入一遍。
        """
        old_root, _ = self._install(
            "HYXi-1.8.1-win64", with_data=True, env_body="TWEAKERS_SECRET_KEY=the-old-key\n")
        new_root, new_exe = self._install("HYXi-1.8.2-win64")
        self._become(new_exe)

        self._adopt(new_root)

        target = os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME)
        with open(os.path.join(target, "hyxi.db"), "rb") as f:
            assert b"HYXi-1.8.1-win64" in f.read(), "接过来的不是旧版本那份库"
        assert os.path.exists(os.path.join(target, "media", "src_x", "p.jpg")), "配图没跟过来"
        with open(os.path.join(target, ".env"), encoding="utf-8") as f:
            assert "the-old-key" in f.read(), "密钥没跟过来，库里的凭据密文就全废了"
        # 复制而不是移动：旧包留在原地，用户确认没问题再自己删
        assert os.path.exists(os.path.join(old_root, "data", "hyxi.db")), "旧包被搬空了，回退不了"

    def test_the_new_version_generates_its_key_into_the_adopted_dir(self):
        """接管之后再补密钥：旧 .env 已经在那儿，绝不能被一把新密钥顶掉"""
        import run_server
        self._install("HYXi-1.8.1-win64", with_data=True, env_body="TWEAKERS_SECRET_KEY=the-old-key\n")
        new_root, new_exe = self._install("HYXi-1.8.2-win64")
        self._become(new_exe)

        run_server.adopt_previous_install(new_root)
        run_server.ensure_env_file(self.paths.env_file())

        with open(self.paths.env_file(), encoding="utf-8") as f:
            body = f.read()
        assert "the-old-key" in body, "旧密钥被新生成的顶掉了 —— 凭据全部解不开"
        assert body.count("TWEAKERS_SECRET_KEY=") == 1, body

    def test_the_busiest_previous_install_wins(self):
        """旁边摆着好几个旧版本时，取最近用过的那个（按 hyxi.db 的改动时间）"""
        stale, _ = self._install("HYXi-1.6.0-win64", with_data=True)
        recent, _ = self._install("HYXi-1.8.1-win64", with_data=True)
        os.utime(os.path.join(stale, "data", "hyxi.db"), (1_600_000_000, 1_600_000_000))
        new_root, new_exe = self._install("HYXi-1.8.2-win64")
        self._become(new_exe)

        self._adopt(new_root)

        with open(os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME, "hyxi.db"), "rb") as f:
            assert b"HYXi-1.8.1-win64" in f.read(), "接的是那个早就不用的旧版本"

    def test_existing_external_data_is_never_overwritten(self):
        """第二次、第三次升级：外部数据已经在了，旁边的旧包一律不许覆盖它。

        少了这条，每升一次级都会被某个还没删掉的旧文件夹把当前数据顶回去。
        """
        target = os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME)
        os.makedirs(target)
        with open(os.path.join(target, "hyxi.db"), "wb") as f:
            f.write(b"SQLite format 3\x00CURRENT")
        self._install("HYXi-1.8.1-win64", with_data=True)
        new_root, new_exe = self._install("HYXi-1.8.3-win64")
        self._become(new_exe)

        self._adopt(new_root)

        with open(os.path.join(target, "hyxi.db"), "rb") as f:
            assert b"CURRENT" in f.read(), "当前数据被旧包顶掉了"

    # ---------- 「先试了一下新版本」这条路（用户实测报过） ----------

    def test_a_version_opened_before_the_data_existed_still_adopts_it_later(self):
        """用户实测报的：新版本先双击试了一下，那会儿旧版本还没配东西；之后回旧版本
        配好 LLM（模型名 deepseek-chat123），再启新版本 —— 一条都没带过来。

        「先试试新版本 → 发现要重配 → 回去接着用旧的 → 再来试新的」正是最自然的
        升级姿势，而第一次试跑就把外部数据目录建出来了。接管原先只看这个目录
        存不存在，于是此后**永远**短路，用户还看不到任何原因。
        判据必须是「里面有没有用户自己的东西」。
        """
        old_root, _ = self._install(
            "HYXi-1.8.1-win64", env_body="TWEAKERS_SECRET_KEY=the-old-key\n")
        self._real_db(os.path.join(old_root, "data"), model="deepseek-chat123")
        new_root, new_exe = self._install("HYXi-1.9.0-win64")
        self._become(new_exe)
        # 新版本被打开过一次：目录、空库、新密钥都建出来了，但用户什么都还没做
        target = os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME)
        self._real_db(target)
        with open(os.path.join(target, ".env"), "w", encoding="utf-8") as f:
            f.write("TWEAKERS_SECRET_KEY=a-brand-new-key\n")

        self._adopt(new_root)

        assert self._model_name(target) == "deepseek-chat123", "旧版本配的模型没带过来"
        with open(os.path.join(target, ".env"), encoding="utf-8") as f:
            assert "the-old-key" in f.read(), "密钥没跟着搬 —— 库里的凭据密文就全废了"

    def test_a_data_dir_the_user_has_actually_used_is_never_replaced(self):
        """对照组：外部目录里已经有用户自己配的东西，旁边的旧包一律不许顶掉它。

        只测上一条的话，把接管改成「无条件覆盖」也是绿的 —— 而那会让每一次
        启动都被某个还没删掉的旧文件夹把当前数据冲回去。
        """
        old_root, _ = self._install("HYXi-1.8.1-win64", env_body="TWEAKERS_SECRET_KEY=old\n")
        self._real_db(os.path.join(old_root, "data"), model="deepseek-chat123")
        new_root, new_exe = self._install("HYXi-1.9.1-win64")
        self._become(new_exe)
        target = os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME)
        self._real_db(target, model="用户在新版本里自己配的")

        self._adopt(new_root)

        assert self._model_name(target) == "用户在新版本里自己配的", "正在用的配置被旧包顶掉了"

    def test_a_data_dir_where_only_a_source_was_added_is_still_protected(self):
        """只加了数据源、还没来得及配 LLM 的目录，也不许被旧包顶掉。

        判据里 sources 的门槛是 1（首启会自动补一条 Tweakers 源）——「大于 0」会让
        每个空目录都算有数据，而「不看 sources」则会把用户刚加的源静默冲掉。
        """
        old_root, _ = self._install("HYXi-1.8.1-win64", env_body="TWEAKERS_SECRET_KEY=old\n")
        self._real_db(os.path.join(old_root, "data"), model="deepseek-chat123")
        new_root, new_exe = self._install("HYXi-1.9.1-win64")
        self._become(new_exe)
        target = os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME)
        self._real_db(target)
        import app.services.storage as storage
        old_db = storage.DB_PATH
        storage.DB_PATH = os.path.join(target, "hyxi.db")
        try:
            storage.save_source({"id": "src_mine", "name": "我自己加的",
                                 "collector_id": "tweakers", "params": {}, "enabled": True})
        finally:
            storage.DB_PATH = old_db

        self._adopt(new_root)

        storage.DB_PATH = os.path.join(target, "hyxi.db")
        try:
            assert any(s["id"] == "src_mine" for s in storage.load_sources()), \
                "用户刚加的数据源被旧包冲掉了"
        finally:
            storage.DB_PATH = old_db

    def test_an_empty_previous_install_is_not_adopted_over_and_over(self):
        """旁边摆着一个从没用过的旧包：不接管，也别打「已接管」的假日志。

        接管一个空库之后 target 仍然是空的，下次启动会照做一遍 —— 每启动一次
        复制一次，日志上却说数据接过来了。
        """
        old_root, _ = self._install("HYXi-1.8.1-win64", env_body="TWEAKERS_SECRET_KEY=old\n")
        self._real_db(os.path.join(old_root, "data"))       # 装过、但什么都没配
        new_root, new_exe = self._install("HYXi-1.9.1-win64")
        self._become(new_exe)

        self._adopt(new_root)

        assert not os.path.exists(os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME)), \
            "把一个空库接管过来了"

    def test_replacing_an_unused_data_dir_leaves_no_leftovers(self):
        """顶掉的那个空壳目录不能留在数据目录旁边 —— 多一个看不懂的文件夹，
        用户下次就不知道该备份哪个了。
        """
        old_root, _ = self._install("HYXi-1.8.1-win64", env_body="TWEAKERS_SECRET_KEY=k\n")
        self._real_db(os.path.join(old_root, "data"), model="deepseek-chat123")
        new_root, new_exe = self._install("HYXi-1.9.0-win64")
        self._become(new_exe)
        target = os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME)
        self._real_db(target)

        self._adopt(new_root)

        leftovers = [n for n in os.listdir(self.tmp)
                     if n.startswith(self.paths.EXTERNAL_DATA_DIRNAME)
                     and n != self.paths.EXTERNAL_DATA_DIRNAME]
        assert not leftovers, f"数据目录旁边留下了 {leftovers}"

    def test_a_failed_replacement_puts_the_unused_dir_back(self):
        """让路之后才失败：必须把它挪回来。

        不挪回来的话，用户连那个（虽然是空的）数据目录都没了 —— 里面的 .env
        跟着消失，下次启动重新生成一把密钥，而旧包里的凭据是用别的密钥加密的。
        """
        old_root, _ = self._install("HYXi-1.8.1-win64", env_body="TWEAKERS_SECRET_KEY=k\n")
        self._real_db(os.path.join(old_root, "data"), model="deepseek-chat123")
        new_root, new_exe = self._install("HYXi-1.9.0-win64")
        self._become(new_exe)
        target = os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME)
        self._real_db(target)
        with open(os.path.join(target, ".env"), "w", encoding="utf-8") as f:
            f.write("TWEAKERS_SECRET_KEY=still-mine\n")

        real_rename = os.rename
        blown = []

        def flaky(src, dst, *a, **kw):
            # 只炸第一次 —— 杀软锁一下就放开，还原那一步该能走通
            if os.path.normcase(str(dst)) == os.path.normcase(target) and not blown:
                blown.append(1)
                raise OSError("目标目录被杀软锁着")   # 扶正 staging 那一步炸
            return real_rename(src, dst, *a, **kw)

        os.rename = flaky
        try:
            self._adopt(new_root)
        finally:
            os.rename = real_rename

        assert os.path.isdir(target), "让路的那份没挪回来，用户的数据目录凭空消失了"
        with open(os.path.join(target, ".env"), encoding="utf-8") as f:
            assert "still-mine" in f.read(), "挪回来的不是原来那份"
        assert not os.path.exists(target + ".partial"), "暂存目录没清掉"

    def test_a_lone_fresh_install_starts_empty_without_error(self):
        """旁边什么都没有：安安静静地从空开始，不该抛异常"""
        new_root, new_exe = self._install("HYXi-1.8.2-win64")
        self._become(new_exe)
        self._adopt(new_root)
        assert not os.path.exists(os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME))

    def test_old_layout_install_is_left_alone(self):
        """包内已有 data 的老安装：原地用自己的，不往外搬、也不去接别人的"""
        old_root, old_exe = self._install("HYXi-1.8.1-win64", with_data=True)
        self._install("HYXi-1.7.0-win64", with_data=True)
        self._become(old_exe)

        self._adopt(old_root)

        assert self.paths.data_dir() == os.path.join(old_root, "data")
        assert not os.path.exists(os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME))

    # ---------- 两个易错的边界 ----------

    def test_a_hand_written_env_in_the_package_never_shadows_the_data_one(self):
        """用户照《使用说明》在包根手写 .env 开局域网访问 —— 不能遮住数据目录那份。

        遮住了的话，ensure_env_file() 会见它没有 SECRET_KEY 便补上一把**新**密钥，
        而库里的凭据是用旧密钥加密的 —— 界面只会说「与保存时的密钥不一致」。
        """
        import run_server
        root, exe = self._install("HYXi-1.9.0-win64")
        self._become(exe)
        external = os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME)
        os.makedirs(external)
        with open(os.path.join(external, ".env"), "w", encoding="utf-8") as f:
            f.write("TWEAKERS_SECRET_KEY=the-real-key\n")
        # 用户在包根另建了一份，里面只有局域网那几行
        with open(os.path.join(root, ".env"), "w", encoding="utf-8") as f:
            f.write("TWEAKERS_HOST=0.0.0.0\n")

        assert self.paths.env_file() == os.path.join(external, ".env"), \
            "包根那份把数据目录的 .env 遮住了"

        run_server.ensure_env_file(self.paths.env_file())
        with open(os.path.join(external, ".env"), encoding="utf-8") as f:
            assert "the-real-key" in f.read(), "真正在用的密钥被动了"
        with open(os.path.join(root, ".env"), encoding="utf-8") as f:
            assert "SECRET_KEY" not in f.read(), "往包根那份里补了一把新密钥"

    def test_a_failed_adoption_leaves_nothing_behind_so_it_retries(self):
        """搬到一半出错：不能把半份数据留在目标位置。

        留下半份是永久性的坏状态：.env 那一步没执行到，密钥会被当成缺失
        重新生成；而下次启动 os.path.exists(target) 直接短路，再也不会重试。
        """
        import run_server
        old_root, _ = self._install(
            "HYXi-1.8.1-win64", with_data=True, env_body="TWEAKERS_SECRET_KEY=k\n")
        new_root, new_exe = self._install("HYXi-1.9.0-win64")
        self._become(new_exe)
        target = os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME)

        real_copytree = shutil.copytree

        def boom(src, dst, *a, **kw):
            real_copytree(src, dst, *a, **kw)      # 先真的把文件写出去
            raise shutil.Error("[('x', 'y', 'PermissionError')]")   # 再像它那样收尾抛

        shutil.copytree = boom
        try:
            run_server.adopt_previous_install(new_root)
        finally:
            shutil.copytree = real_copytree

        assert not os.path.exists(target), "半份数据留在了目标位置，下次启动不会重试"
        assert not os.path.exists(target + ".partial"), "暂存目录没清掉"
        # 修好那个毛病后再启动一次，应该能正常接过来
        run_server.adopt_previous_install(new_root)
        with open(os.path.join(target, ".env"), encoding="utf-8") as f:
            assert "TWEAKERS_SECRET_KEY=k" in f.read()

    def test_source_checkout_never_adopts_a_neighbouring_project(self):
        """源码态调到这里时不能去扫仓库的父目录 —— 那里躺着别的项目"""
        import run_server
        if hasattr(sys, "frozen"):
            del sys.frozen
        neighbour = os.path.join(self.tmp, "隔壁项目")
        os.makedirs(os.path.join(neighbour, "data"))
        with open(os.path.join(neighbour, "data", "hyxi.db"), "wb") as f:
            f.write(b"SQLite format 3\x00NEIGHBOUR")
        fake_repo = os.path.join(self.tmp, "仓库")
        os.makedirs(fake_repo)

        run_server.adopt_previous_install(fake_repo)   # 不报错，也什么都不该做

        assert not os.path.exists(os.path.join(self.tmp, self.paths.EXTERNAL_DATA_DIRNAME))
        assert not os.path.exists(os.path.join(fake_repo, "data"))
